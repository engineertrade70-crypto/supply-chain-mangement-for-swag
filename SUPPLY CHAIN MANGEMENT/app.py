"""
SWAG Product Comparison Dashboard
Version 29.0 — Full fix: multi-company analytics, bilingual stability,
               canonical columns, pagination, localization consistency
"""

import io
import re
import hashlib
import time
import xmlrpc.client
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import altair as alt
import pandas as pd
import streamlit as st

try:
    import plotly.express as px
    import plotly.graph_objects as go
    _HAS_PLOTLY = True
except ImportError:
    _HAS_PLOTLY = False

st.set_page_config(
    page_title="SWAG Dashboard",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

* { margin: 0; padding: 0; box-sizing: border-box; }
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.stApp {
    background: linear-gradient(135deg, #0a0c10 0%, #121417 100%);
    min-height: 100vh;
}

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f1115 0%, #0a0c10 100%) !important;
    border-right: 1px solid rgba(255,255,255,0.05);
    box-shadow: 4px 0 20px rgba(0,0,0,0.3);
}
section[data-testid="stSidebar"] *,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] div { color: #e8edf2 !important; }
section[data-testid="stSidebar"] input {
    background: #1a1e24 !important;
    color: #e8edf2 !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
}

@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(30px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}
@keyframes slideInLeft {
    from { opacity: 0; transform: translateX(-30px); }
    to   { opacity: 1; transform: translateX(0); }
}
@keyframes shimmer {
    0%   { background-position: -200% center; }
    100% { background-position:  200% center; }
}

.login-orb {
    width: 100px; height: 100px; border-radius: 50%;
    background: linear-gradient(135deg, #2c3e50, #3498db);
    display: flex; align-items: center; justify-content: center;
    font-size: 2.5rem; margin: 0 auto 20px;
    animation: fadeInUp 0.6s ease forwards;
    box-shadow: 0 8px 32px rgba(52,152,219,0.3);
}
.login-title {
    font-size: 2rem; font-weight: 700;
    background: linear-gradient(90deg, #3498db, #2ecc71);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    text-align: center; margin-bottom: 8px;
    animation: fadeInUp 0.7s ease forwards;
}
.login-subtitle {
    color: #8e9aaf !important; font-size: 0.9rem; text-align: center;
    animation: fadeInUp 0.8s ease forwards; margin-bottom: 28px;
}
.login-card {
    background: rgba(20,24,30,0.8); backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.1); border-radius: 24px;
    padding: 32px 36px; width: 100%;
    animation: fadeInUp 0.9s ease forwards;
}

.dash-header { text-align: center; padding: 20px 0 28px; }
.dash-title {
    font-size: 2.2rem; font-weight: 800;
    background: linear-gradient(90deg, #3498db, #2ecc71, #3498db);
    background-size: 200% auto;
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    animation: shimmer 4s linear infinite;
}
.dash-subtitle { color: #8e9aaf; font-size: 0.9rem; margin-top: 4px; }

.kpi-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px; margin: 20px 0 24px;
}
.kpi-card {
    background: linear-gradient(135deg, #1a1e24, #14181e);
    border: 1px solid rgba(255,255,255,0.08); border-radius: 20px;
    padding: 20px; position: relative; overflow: hidden;
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
    animation: fadeInUp 0.5s ease forwards;
}
.kpi-card:hover {
    transform: translateY(-4px);
    border-color: rgba(52,152,219,0.3);
    box-shadow: 0 12px 32px rgba(0,0,0,0.3);
}
.kpi-card::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #3498db, #2ecc71);
    border-radius: 20px 20px 0 0;
}
.kpi-icon { font-size: 1.8rem; margin-bottom: 12px; display: block; opacity: 0.9; }
.kpi-value {
    font-size: 1.8rem; font-weight: 800; line-height: 1.2; margin-bottom: 6px;
    background: linear-gradient(90deg, #fff, #3498db);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    letter-spacing: -0.5px;
}
.kpi-label { font-size: 0.75rem; color: #8e9aaf; font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
.kpi-trend { font-size: 0.7rem; margin-top: 8px; color: #2ecc71; }

.section-header {
    background: linear-gradient(135deg, #1a1e24, #14181e);
    border-left: 4px solid #3498db; border-radius: 12px;
    padding: 12px 20px; margin: 24px 0 20px;
    font-weight: 700; font-size: 1rem; color: #e8edf2;
    letter-spacing: -0.2px; animation: slideInLeft 0.4s ease forwards;
}

.chart-card {
    background: linear-gradient(135deg, #1a1e24, #14181e);
    border: 1px solid rgba(255,255,255,0.05); border-radius: 20px;
    padding: 20px; margin-bottom: 20px;
    transition: all 0.3s ease; animation: fadeInUp 0.5s ease forwards;
}
.chart-card:hover {
    border-color: rgba(52,152,219,0.2);
    box-shadow: 0 8px 24px rgba(0,0,0,0.2);
}
.chart-title {
    font-size: 0.9rem; font-weight: 600; color: #8e9aaf;
    margin-bottom: 16px; display: flex; align-items: center;
    gap: 8px; letter-spacing: 0.3px;
}
.chart-title span.accent {
    width: 3px; height: 16px;
    background: linear-gradient(180deg, #3498db, #2ecc71);
    border-radius: 2px; display: inline-block;
}

.table-wrapper {
    width: 100%; overflow-x: auto; border-radius: 20px;
    background: linear-gradient(135deg, #1a1e24, #14181e);
    border: 1px solid rgba(255,255,255,0.05); margin: 16px 0;
    animation: fadeInUp 0.5s ease forwards;
}
.premium-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
.premium-table thead tr {
    background: linear-gradient(135deg, #0f1115, #0a0c10);
    border-bottom: 1px solid rgba(52,152,219,0.3);
}
.premium-table thead th {
    color: #3498db; font-weight: 600; padding: 14px 16px;
    text-align: center; font-size: 0.8rem; text-transform: uppercase;
    letter-spacing: 0.5px; position: sticky; top: 0;
    background: #0f1115; z-index: 10;
}
.premium-table tbody tr {
    transition: all 0.2s ease;
    border-bottom: 1px solid rgba(255,255,255,0.05);
}
.premium-table tbody tr:hover {
    background: rgba(52,152,219,0.1); transform: scale(1.01);
}
.premium-table tbody td { padding: 12px 16px; text-align: center; color: #e8edf2; }
.premium-table tbody td:first-child { font-weight: 600; color: #3498db; }

.pagination-bar {
    display: flex; align-items: center; justify-content: center;
    gap: 12px; padding: 12px 0; margin-top: 4px;
    font-size: 0.85rem; color: #8e9aaf;
}
.page-info {
    background: rgba(52,152,219,0.12); border: 1px solid rgba(52,152,219,0.3);
    border-radius: 8px; padding: 4px 14px; color: #3498db; font-weight: 600;
}

[dir="rtl"] .premium-table thead th { text-align: right; }
[dir="rtl"] .section-header { border-left: none; border-right: 4px solid #3498db; }

.stButton button {
    background: linear-gradient(135deg, #2c3e50, #1a1e24) !important;
    border: 1px solid rgba(52,152,219,0.3) !important;
    border-radius: 12px !important; color: #e8edf2 !important;
    font-weight: 600 !important; transition: all 0.3s ease !important;
}
.stButton button:hover {
    transform: translateY(-2px);
    border-color: #3498db !important;
    box-shadow: 0 4px 12px rgba(52,152,219,0.2) !important;
}
.stButton button[kind="primary"] {
    background: linear-gradient(90deg, #3498db, #2ecc71) !important;
    border: none !important; color: white !important;
}

.stTextInput input, .stNumberInput input,
.stTextArea textarea, .stDateInput input {
    background: #1a1e24 !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    border-radius: 12px !important; color: #e8edf2 !important;
}
.stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
    border-color: #3498db !important;
    box-shadow: 0 0 0 2px rgba(52,152,219,0.2) !important;
}

.stTabs [data-baseweb="tab-list"] {
    background: #1a1e24; border-radius: 12px; padding: 4px; gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    color: #8e9aaf !important; border-radius: 10px !important;
    font-size: 0.8rem !important; font-weight: 600 !important; padding: 8px 20px !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg, #3498db, #2ecc71) !important; color: white !important;
}

.info-banner {
    background: rgba(52,152,219,0.1); border-left: 3px solid #3498db;
    border-radius: 12px; padding: 12px 16px; margin: 16px 0;
    font-size: 0.85rem; color: #8e9aaf;
}
.warn-banner {
    background: rgba(241,196,15,0.1); border-left: 3px solid #f1c40f;
    border-radius: 12px; padding: 12px 16px; margin: 16px 0;
    font-size: 0.85rem; color: #f1c40f;
}
.alert-banner {
    background: rgba(231,76,60,0.1); border-left: 3px solid #e74c3c;
    border-radius: 12px; padding: 12px 16px; margin: 16px 0;
    font-size: 0.85rem; color: #e74c3c; animation: pulse 2s infinite;
}
@keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.7; } }

::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: #0a0c10; }
::-webkit-scrollbar-thumb { background: #2c3e50; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #3498db; }

.divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(52,152,219,0.3), transparent);
    margin: 24px 0;
}
.stats-card {
    background: #1a1e24; border-radius: 16px; padding: 16px;
    text-align: center; border: 1px solid rgba(255,255,255,0.05);
}
.stats-value { font-size: 1.5rem; font-weight: 800; color: #3498db; }
.stats-label { font-size: 0.7rem; color: #8e9aaf; margin-top: 4px; }
.footer {
    text-align: center; padding: 24px 0; color: #5a6e8a;
    font-size: 0.7rem; border-top: 1px solid rgba(255,255,255,0.05); margin-top: 40px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_KEYS = ["SWAG", "LAROUCHE", "DIFFC", "FASHION_LIMITS"]

# ─────────────────────────────────────────────────────────────────────────────
# CANONICAL COLUMN NAMES  ← always English internally
# ─────────────────────────────────────────────────────────────────────────────
# These are the internal schema names used in ALL DataFrames throughout the app.
# Translation to Arabic happens only at display time.
C_SYSTEM        = "System"
C_MODEL         = "Model Code"
C_PRODUCT       = "Product"
C_SALE_PRICE    = "Sale Price"
C_ON_HAND       = "On Hand"
C_BRANCH        = "Branch"
C_LOCATION      = "Location"
C_REFERENCE     = "Reference"
C_TYPE          = "Type"
C_STATE         = "State"
C_FROM          = "From"
C_TO            = "To"
C_QTY           = "Qty"
C_SCHEDULED     = "Scheduled"
C_SOLD          = "Sold(30d)"
C_VEL           = "Daily Vel"
C_DAYS_LEFT     = "Days Left"
C_SUGGEST       = "Suggest"
C_PRIORITY      = "Priority"
C_DATE          = "Date"
C_PO            = "PO"
C_SO            = "SO"
C_VENDOR        = "Vendor"
C_CUSTOMER      = "Customer"
C_BRAND_CAT     = "Brand Category"
C_CATEGORY      = "Category"
C_UNIT_PRICE    = "Unit Price"
C_SUBTOTAL      = "Subtotal"
C_QTY_PURCHASED = "Qty Purchased"

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE
# ─────────────────────────────────────────────────────────────────────────────
def get_lang():
    return st.session_state.get("lang", "EN")

def t(en, ar):
    """UI translation only — never use for DataFrame column names."""
    return ar if get_lang() == "AR" else en

def get_system_name(key):
    cfg = st.secrets.get(key, {})
    return cfg.get("name_ar", cfg.get("name", key)) if get_lang() == "AR" else cfg.get("name", key)

def get_dir():
    return "rtl" if get_lang() == "AR" else "ltr"

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN DISPLAY LABEL MAP  ← translate canonical → UI label
# ─────────────────────────────────────────────────────────────────────────────
_COL_LABELS_EN = {
    C_SYSTEM: "System", C_MODEL: "Model Code", C_PRODUCT: "Product",
    C_SALE_PRICE: "Sale Price", C_ON_HAND: "On Hand", C_BRANCH: "Branch",
    C_LOCATION: "Location", C_REFERENCE: "Reference", C_TYPE: "Type",
    C_STATE: "State", C_FROM: "From", C_TO: "To", C_QTY: "Qty",
    C_SCHEDULED: "Scheduled", C_SOLD: "Sold(30d)", C_VEL: "Daily Vel",
    C_DAYS_LEFT: "Days Left", C_SUGGEST: "Suggest", C_PRIORITY: "Priority",
    C_DATE: "Date", C_PO: "PO", C_SO: "SO", C_VENDOR: "Vendor",
    C_CUSTOMER: "Customer", C_BRAND_CAT: "Brand Category", C_CATEGORY: "Category",
    C_UNIT_PRICE: "Unit Price", C_SUBTOTAL: "Subtotal",
    C_QTY_PURCHASED: "Qty Purchased",
}
_COL_LABELS_AR = {
    C_SYSTEM: "النظام", C_MODEL: "رمز الموديل", C_PRODUCT: "المنتج",
    C_SALE_PRICE: "سعر البيع", C_ON_HAND: "متوفر", C_BRANCH: "الفرع",
    C_LOCATION: "الموقع", C_REFERENCE: "المرجع", C_TYPE: "النوع",
    C_STATE: "الحالة", C_FROM: "من", C_TO: "إلى", C_QTY: "الكمية",
    C_SCHEDULED: "المجدول", C_SOLD: "مباع(30ي)", C_VEL: "معدل/يوم",
    C_DAYS_LEFT: "أيام متبقية", C_SUGGEST: "المقترح", C_PRIORITY: "الأولوية",
    C_DATE: "التاريخ", C_PO: "أمر الشراء", C_SO: "أمر البيع",
    C_VENDOR: "المورد", C_CUSTOMER: "العميل", C_BRAND_CAT: "الفئة التجارية",
    C_CATEGORY: "الفئة", C_UNIT_PRICE: "سعر الوحدة", C_SUBTOTAL: "المجموع",
    C_QTY_PURCHASED: "الكمية المشتراة",
}

def col_label(canonical):
    """Return the UI display label for a canonical column name."""
    if get_lang() == "AR":
        return _COL_LABELS_AR.get(canonical, canonical)
    return _COL_LABELS_EN.get(canonical, canonical)

def df_for_display(df):
    """
    Return a copy of df with columns renamed to current UI language for display only.
    Internal DataFrames always stay in canonical English.
    """
    if df is None or df.empty:
        return df
    label_map = _COL_LABELS_AR if get_lang() == "AR" else _COL_LABELS_EN
    return df.rename(columns={k: v for k, v in label_map.items() if k in df.columns})

# ─────────────────────────────────────────────────────────────────────────────
# SAFE NUMERIC HELPER
# ─────────────────────────────────────────────────────────────────────────────
def _to_num(series):
    """Safely coerce a pandas Series to float, replacing errors/NaN with 0."""
    return pd.to_numeric(series, errors="coerce").fillna(0)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
_DEF = {
    "authenticated"      : False,
    "user_email"         : "",
    "lang"               : "EN",
    "last_run"           : None,
    "total_df"           : None,
    "branch_df"          : None,
    "transfers_df"       : None,
    "reorder_df"         : None,
    "sys_stats"          : {},
    "search_exact"       : False,
    "low_stock_thresh"   : 5,
    "price_history"      : {},
    "show_transfers"     : False,
    "show_reorder"       : False,
    "reorder_mode"       : "days_cover",
    "reorder_target_days": 30,
    "reorder_max_level"  : 100,
    "reorder_point"      : 10,
    "pdf_codes"          : None,
    "pdf_mode"           : "total",
    "po_analytics_df"    : None,
    "pc_purch_df"        : None,
    "pc_stock_df"        : None,
    "pc_last_code"       : "",
    "salesanalyticsdf"   : None,
    "analytics_view"     : "purchase",
    # pagination state
    "page_total"         : 0,
    "page_branch"        : 0,
    "page_transfers"     : 0,
    "page_reorder"       : 0,
    "page_po"            : 0,
    "page_sales"         : 0,
}
for k, v in _DEF.items():
    if k not in st.session_state:
        st.session_state[k] = v

PAGE_SIZE = 50

# ─────────────────────────────────────────────────────────────────────────────
# SESSION LOGIN RESTORE
# ─────────────────────────────────────────────────────────────────────────────
_COOKIE_SECRET = "swag_2025_secure"

def _make_token(email):
    return hashlib.sha256(f"{_COOKIE_SECRET}_{email}".encode()).hexdigest()[:32]

def _verify_token(email, token):
    return bool(email and token and token == _make_token(email))

def restore_session():
    if st.session_state.get("authenticated"):
        return
    try:
        params = st.query_params
        email  = params.get("u", "")
        token  = params.get("t", "")
        if email and token and _verify_token(email, token):
            st.session_state.authenticated = True
            st.session_state.user_email    = email
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# XML-RPC
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def _proxy(url, ep):
    return xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/{ep}", allow_none=True)

@st.cache_data(ttl=28800, show_spinner=False)
def _auth(url, db, user, key):
    try:
        uid = _proxy(url, "common").authenticate(db, user, key, {})
        return uid or None
    except Exception:
        return None

def _x(url, db, uid, key, model, method, domain, kw):
    return _proxy(url, "object").execute_kw(db, uid, key, model, method, domain, kw)

# ─────────────────────────────────────────────────────────────────────────────
# DOMAIN BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def _domain(codes, exact):
    if exact:
        return [["default_code", "in", codes]]
    if len(codes) == 1:
        return [["default_code", "=like", f"{codes[0]}%"]]
    parts = [["default_code", "=like", f"{c}%"] for c in codes]
    return ["|"] * (len(parts) - 1) + parts

# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSING
# ─────────────────────────────────────────────────────────────────────────────
_RE_BRACKET = re.compile(r'\[([A-Za-z0-9\-_()]{3,30})\]')
_RE_SR_LINE = re.compile(
    r'(?:^|\s)([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9()]{1,10})?)\s+.{0,80}?\d+\.?\d*\s+SR',
    re.MULTILINE)
_RE_GENERAL = re.compile(
    r'\b([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9]{1,4})?(?:\([^)]{1,15}\))?)\b')
_EXCLUDE = frozenset([
    'SR','VAT','TAX','PCS','QTY','NO','REF','INV','PO','SO',
    'DO','ID','EN','AR','PDF','AED','SAR','USD','KWD','OMR',
    'BHD','JOD','EGP','TRY'
])

def _valid(code):
    c = code.strip().upper()
    return (bool(re.search(r'[A-Z]', c)) and bool(re.search(r'\d', c))
            and 4 <= len(c) <= 25 and c not in _EXCLUDE)

def extract_base_model(code):
    code = re.sub(r'\([^)]*\)', '', code)
    for s in ['-2XL','-3XL','-4XL','-XXL','-XL','-L','-M','-S','-XS','-2X','-3X']:
        if code.upper().endswith(s.upper()):
            code = code[:-len(s)]; break
    return re.sub(r'-\d{2,3}$', '', code).strip()

def get_unique_base_models(raw):
    seen, out = set(), []
    for item in raw:
        b = extract_base_model(item["code"])
        if b and b not in seen:
            seen.add(b)
            out.append({"sequence": item["sequence"], "code": b})
    return out

@st.cache_data(show_spinner=False)
def parse_invoice_pdf_cached(file_bytes):
    try:
        from pypdf import PdfReader
    except ImportError:
        return []
    text = ""
    for page in PdfReader(io.BytesIO(file_bytes)).pages:
        text += (page.extract_text() or "") + "\n"
    if not text.strip():
        return []
    raw = (_RE_BRACKET.findall(text)
           + [m.group(1) for m in _RE_SR_LINE.finditer(text)]
           + _RE_GENERAL.findall(text))
    seen, out = set(), []
    seq = 1
    for c in raw:
        u = c.strip().upper()
        if _valid(u) and u not in seen:
            seen.add(u)
            out.append({"sequence": seq, "code": u})
            seq += 1
    return out

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS  (unchanged from original — write canonical columns to Excel)
# ─────────────────────────────────────────────────────────────────────────────
def _style_worksheet(ws, df_clean, lang="EN"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
    from openpyxl.chart import BarChart, Reference
    if lang == "AR":
        ws.sheet_view.rightToLeft = True
    hdr_fill     = PatternFill("solid", fgColor="2C3E50")
    hdr_font     = Font(bold=True, color="FFFFFF", size=11, name="Inter")
    hdr_align    = Alignment(horizontal="center", vertical="center")
    thin         = Side(border_style="thin", color="D0D0D0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", fgColor="1A1E24")
    zero_fill    = PatternFill("solid", fgColor="FFE0E0")
    zero_font    = Font(color="CC0000", bold=True, name="Inter")
    normal_font  = Font(name="Inter", size=10, color="E8EDF2")
    num_align    = Alignment(horizontal="right",  vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    total_fill   = PatternFill("solid", fgColor="3498DB")
    total_font   = Font(bold=True, name="Inter", color="FFFFFF")
    max_row = ws.max_row
    max_col = ws.max_column
    ws.row_dimensions[1].height = 28
    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    col_names = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    on_hand_col = sale_price_col = loc_col = branch_col = model_col = None
    for i, name in enumerate(col_names, 1):
        # check both canonical and translated names for robustness
        if name in (C_ON_HAND, "متوفر"):       on_hand_col    = i
        if name in (C_SALE_PRICE, "سعر البيع"): sale_price_col = i
        if name in (C_LOCATION, "الموقع"):      loc_col        = i
        if name in (C_BRANCH, "الفرع"):         branch_col     = i
        if name in (C_MODEL, "رمز الموديل"):    model_col      = i
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        is_zero = False
        if on_hand_col:
            val = ws.cell(row=row[0].row, column=on_hand_col).value
            is_zero = (val is None or
                       str(val).strip() in ['0','Not Available','غير متوفر','—','-',''] or
                       val == 0)
        for cell in row:
            cell.border = border
            cell.font   = zero_font if is_zero else normal_font
            if is_zero:   cell.fill = zero_fill
            elif cell.row % 2 == 0: cell.fill = alt_fill
            if isinstance(cell.value, (int, float)): cell.alignment = num_align
            else: cell.alignment = center_align
        ws.row_dimensions[row[0].row].height = 18
    for col_num in range(1, max_col + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for r in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in r:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if on_hand_col and max_row > 1:
        col_letter = get_column_letter(on_hand_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            DataBarRule(start_type="min", end_type="max", color="3498DB"))
    if sale_price_col and max_row > 1:
        col_letter = get_column_letter(sale_price_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            ColorScaleRule(start_type="min", start_color="63BE7B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="F8696B"))
    if on_hand_col and max_row > 1:
        col_letter     = get_column_letter(on_hand_col)
        low_stock_fill = PatternFill("solid", fgColor="FFF2CC")
        low_stock_font = Font(color="7F6000", bold=True, name="Inter")
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            CellIsRule(operator="lessThanOrEqual", formula=["3"],
                       fill=low_stock_fill, font=low_stock_font))
    total_row = max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font      = total_font
    ws.cell(row=total_row, column=1).fill      = total_fill
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    if on_hand_col:
        col = get_column_letter(on_hand_col)
        ws.cell(row=total_row, column=on_hand_col,
                value=f"=SUM({col}2:{col}{max_row})")
        ws.cell(row=total_row, column=on_hand_col).font      = total_font
        ws.cell(row=total_row, column=on_hand_col).fill      = total_fill
        ws.cell(row=total_row, column=on_hand_col).alignment = Alignment(horizontal="center")
    ws.row_dimensions[total_row].height = 20
    ws.sheet_properties.tabColor = "3498DB"
    footer_row = total_row + 2
    ws.cell(row=footer_row, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SWAG Dashboard")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9, name="Inter")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = "1:1"
    ws.print_area             = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.oddHeader.center.text  = "SWAG Product Report"
    ws.oddHeader.center.font  = "Inter,Bold"
    ws.oddFooter.center.text  = "Page &P of &N  |  Generated: &D"
    ws.sheet_view.zoomScale   = 85
    if loc_col:
        ws.column_dimensions[get_column_letter(loc_col)].width = 35
        for row_num in range(2, max_row + 1):
            ws.cell(row=row_num, column=loc_col).alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="left")
            ws.row_dimensions[row_num].height = 28

def _excel_generic(df, sheet_name="Data"):
    """Generic Excel export using canonical column names."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    lang = st.session_state.get("lang", "EN")
    buf  = io.BytesIO()
    clean = df.drop(columns=["_status"], errors="ignore").copy()
    # translate column headers for display in Excel
    clean = clean.rename(columns={k: v for k, v in
                                   (_COL_LABELS_AR if lang=="AR" else _COL_LABELS_EN).items()
                                   if k in clean.columns})
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        clean.to_excel(w, index=False, sheet_name=sheet_name[:31])
        _style_worksheet(w.sheets[sheet_name[:31]], clean, lang=lang)
    return buf.getvalue()

def to_csv(df):
    return df.drop(columns=["_status"], errors="ignore").to_csv(index=False).encode("utf-8-sig")

def to_excel(df):
    return _excel_generic(df, "Data")

def to_excel_bulk(df):
    lang = st.session_state.get("lang", "EN")
    buf  = io.BytesIO()
    lmap = _COL_LABELS_AR if lang == "AR" else _COL_LABELS_EN
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        def _ws(data, name):
            c = data.drop(columns=["_status"], errors="ignore").copy()
            c = c.rename(columns={k: v for k, v in lmap.items() if k in c.columns})
            c.to_excel(w, index=False, sheet_name=name[:31])
            _style_worksheet(w.sheets[name[:31]], c, lang=lang)
        _ws(df, t("All Systems", "كل الأنظمة"))
        if C_SYSTEM in df.columns:
            for key in SYSTEM_KEYS:
                nm  = get_system_name(key)
                sub = df[df[C_SYSTEM] == key]
                if not sub.empty:
                    _ws(sub, nm)
    return buf.getvalue()

def to_excel_purchase(df):
    return _excel_generic(df, "SWAG Purchase")

def to_excel_sales(df):
    return _excel_generic(df, "SWAG Sales")

def dl_name(tag, ext):
    return f"swag_{tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"

# ─────────────────────────────────────────────────────────────────────────────
# FETCH ALL DATA (stock comparison — unchanged logic, canonical columns)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=180, show_spinner=False)
def fetch_all_data(
    codes_tuple, exact=False,
    need_branch=False, need_transfers=False, need_reorder=False,
    reorder_mode="days_cover", target_days=30,
    max_level=100, reorder_point=10,
):
    DAYS  = 30
    dfrom = (datetime.now() - timedelta(days=DAYS)).strftime("%Y-%m-%d 00:00:00")
    codes = list(codes_tuple)
    dom   = _domain(codes, exact)
    SM    = {"draft":"Draft","waiting":"Waiting","confirmed":"Confirmed","assigned":"Ready"}

    def _one(key):
        cfg = st.secrets.get(key)
        R   = {"key": key, "total":[], "branch":[], "transfers":[], "reorder":[]}
        if not cfg:
            R["total"].append({C_SYSTEM:key, C_MODEL:"—", C_PRODUCT:"No config",
                               C_SALE_PRICE:0.0, C_ON_HAND:0, "_status":"ERROR"})
            return R
        uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
        if not uid:
            R["total"].append({C_SYSTEM:key, C_MODEL:"—", C_PRODUCT:"⚠️ Auth failed",
                               C_SALE_PRICE:0.0, C_ON_HAND:0, "_status":"ERROR"})
            return R
        u=cfg["url"]; db=cfg["db"]; ak=cfg["api_key"]
        try:
            prods = _x(u,db,uid,ak,"product.product","search_read",[dom],
                       {"fields":["id","display_name","default_code","qty_available","list_price"],
                        "limit":2000,"order":"default_code asc"})
            if not prods:
                R["total"].append({C_SYSTEM:key, C_MODEL:"—", C_PRODUCT:"Not found",
                                   C_SALE_PRICE:0.0, C_ON_HAND:0, "_status":"NOT_FOUND"})
                return R
            pids = [p["id"] for p in prods]
            pmap = {p["id"]:p for p in prods}
            for p in prods:
                R["total"].append({
                    C_SYSTEM:key,
                    C_MODEL: p.get("default_code") or "—",
                    C_PRODUCT: p.get("display_name") or "",
                    C_SALE_PRICE: float(p.get("list_price") or 0),
                    C_ON_HAND: int(p.get("qty_available") or 0),
                    "_status": "OK"})
            if need_branch:
                internal_locs = _x(u,db,uid,ak,"stock.location","search_read",
                                   [[["usage","=","internal"],["active","=",True]]],
                                   {"fields":["id"],"limit":10000})
                internal_ids  = {l["id"] for l in internal_locs}
                qs = _x(u,db,uid,ak,"stock.quant","search_read",
                        [[["product_id","in",pids],
                          ["location_id","in",list(internal_ids)],
                          ["quantity",">",0]]],
                        {"fields":["product_id","location_id","quantity"],"limit":5000})
                for q in qs:
                    pid = q["product_id"][0] if isinstance(q.get("product_id"),list) else None
                    loc = q.get("location_id") or [None,"—"]
                    ln  = loc[1] if isinstance(loc,list) else str(loc)
                    pm  = pmap.get(pid,{})
                    R["branch"].append({
                        C_SYSTEM: key,
                        C_BRANCH: ln.split("/")[0].strip(),
                        C_MODEL:  pm.get("default_code") or "—",
                        C_LOCATION: ln,
                        C_SALE_PRICE: float(pm.get("list_price") or 0),
                        C_ON_HAND: int(q.get("quantity") or 0),
                        "_status": "OK"})
            if need_transfers:
                mvs = _x(u,db,uid,ak,"stock.move","search_read",
                         [[["product_id","in",pids],
                           ["state","in",["draft","waiting","confirmed","assigned"]]]],
                         {"fields":["picking_id","product_id","product_uom_qty"],"limit":2000})
                if mvs:
                    pkids = list({m["picking_id"][0] for m in mvs
                                  if isinstance(m.get("picking_id"),list)})
                    if pkids:
                        pks   = _x(u,db,uid,ak,"stock.picking","search_read",
                                   [[["id","in",pkids]]],
                                   {"fields":["id","name","picking_type_id","state",
                                              "location_id","location_dest_id","scheduled_date"]})
                        pkmap = {p["id"]:p for p in pks}
                        for mv in mvs:
                            pr = mv.get("picking_id")
                            if not isinstance(pr,list): continue
                            pk = pkmap.get(pr[0],{})
                            def _n(f,_p=pk):
                                v=_p.get(f); return v[1] if isinstance(v,list) else (v or "—")
                            sd = pk.get("scheduled_date") or "—"
                            if sd != "—":
                                try: sd=datetime.strptime(sd,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
                                except: pass
                            pid2 = mv["product_id"][0] if isinstance(mv.get("product_id"),list) else None
                            pm2  = pmap.get(pid2,{})
                            R["transfers"].append({
                                C_SYSTEM:    key,
                                C_REFERENCE: pk.get("name") or "—",
                                C_TYPE:      _n("picking_type_id"),
                                C_STATE:     SM.get(pk.get("state",""),pk.get("state","")),
                                C_FROM:      _n("location_id"),
                                C_TO:        _n("location_dest_id"),
                                C_MODEL:     pm2.get("default_code") or "—",
                                C_QTY:       int(mv.get("product_uom_qty") or 0),
                                C_SCHEDULED: sd,
                                "_status":   "OK"})
            if need_reorder:
                sl = _x(u,db,uid,ak,"sale.order.line","search_read",
                        [[["product_id","in",pids],
                          ["order_id.state","in",["sale","done"]],
                          ["order_id.date_order",">=",dfrom]]],
                        {"fields":["product_id","product_uom_qty"],"limit":10000})
                sm2 = {}
                for l in sl:
                    pid = l["product_id"][0] if isinstance(l.get("product_id"),list) else None
                    if pid: sm2[pid] = sm2.get(pid,0)+float(l.get("product_uom_qty") or 0)
                for p in prods:
                    pid  = p["id"]; cq=int(p.get("qty_available") or 0)
                    sold = sm2.get(pid,0); vel=round(sold/DAYS,2)
                    dl   = str(round(cq/vel,1)) if vel>0 else "∞"
                    sg   = max(0,round(target_days*vel-cq)) if reorder_mode=="days_cover" else max(0,max_level-cq)
                    pr2  = ("🔴 Critical" if cq<=0 else "🟡 Low" if cq<=reorder_point else "🟢 OK")
                    R["reorder"].append({
                        C_SYSTEM:    key,
                        C_MODEL:     p.get("default_code") or "—",
                        C_PRODUCT:   p.get("display_name") or "",
                        C_ON_HAND:   cq,
                        C_SOLD:      int(sold),
                        C_VEL:       vel,
                        C_DAYS_LEFT: dl,
                        C_SUGGEST:   sg,
                        C_PRIORITY:  pr2,
                        "_status":   "OK"})
        except Exception as e:
            R["total"].append({C_SYSTEM:key, C_MODEL:"—", C_PRODUCT:f"❌ {e}",
                               C_SALE_PRICE:0.0, C_ON_HAND:0, "_status":"ERROR"})
        return R

    at=[]; ab=[]; atr=[]; ar=[]
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_one,k):k for k in SYSTEM_KEYS}
        for f in as_completed(futs):
            r = f.result()
            at.extend(r["total"]); ab.extend(r["branch"])
            atr.extend(r["transfers"]); ar.extend(r["reorder"])

    def _df(rows, cols):
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)
    return {
        "total"    : _df(at,  [C_SYSTEM,C_MODEL,C_PRODUCT,C_SALE_PRICE,C_ON_HAND,"_status"]),
        "branch"   : _df(ab,  [C_SYSTEM,C_BRANCH,C_MODEL,C_LOCATION,C_SALE_PRICE,C_ON_HAND,"_status"]),
        "transfers": _df(atr, [C_SYSTEM,C_REFERENCE,C_TYPE,C_STATE,C_FROM,C_TO,C_MODEL,C_QTY,C_SCHEDULED,"_status"]),
        "reorder"  : _df(ar,  [C_SYSTEM,C_MODEL,C_PRODUCT,C_ON_HAND,C_SOLD,C_VEL,C_DAYS_LEFT,C_SUGGEST,C_PRIORITY,"_status"]),
    }

# ─────────────────────────────────────────────────────────────────────────────
# FETCH PURCHASE HISTORY FOR ONE SYSTEM  ← NEW canonical-column version
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_purchase_history_for_system(system_key, model_code, date_from, date_to):
    """
    Returns a DataFrame with canonical English column names.
    Columns: System, Date, PO, Vendor, Brand Category, Category,
             Model Code, Product, Qty, Unit Price, Subtotal
    """
    empty_cols = [C_SYSTEM, C_DATE, C_PO, C_VENDOR, C_BRAND_CAT, C_CATEGORY,
                  C_MODEL, C_PRODUCT, C_QTY, C_UNIT_PRICE, C_SUBTOTAL]
    empty_df = pd.DataFrame(columns=empty_cols)
    cfg = st.secrets.get(system_key)
    if not cfg:
        return empty_df
    uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not uid:
        return empty_df
    u=cfg["url"]; db=cfg["db"]; ak=cfg["api_key"]
    try:
        date_from_dt = f"{date_from} 00:00:00"
        date_to_dt   = f"{date_to} 23:59:59"
        line_domain = [
            ["order_id.state","in",["purchase","done"]],
            ["order_id.date_order",">=",date_from_dt],
            ["order_id.date_order","<=",date_to_dt],
        ]
        if model_code and str(model_code).strip():
            line_domain.append(["product_id.default_code","=",str(model_code).strip()])
        lines = _x(u,db,uid,ak,"purchase.order.line","search_read",
                   [line_domain],
                   {"fields":["order_id","product_id","product_qty","price_unit"],
                    "limit":10000,"order":"order_id desc"})
        if not lines:
            return empty_df
        order_ids   = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"),list)})
        product_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"),list)})
        orders = _x(u,db,uid,ak,"purchase.order","search_read",
                    [[["id","in",order_ids]]],
                    {"fields":["id","name","partner_id","date_order"],"limit":len(order_ids)+10})
        order_map = {o["id"]:o for o in orders}
        products = _x(u,db,uid,ak,"product.product","search_read",
                      [[["id","in",product_ids]]],
                      {"fields":["id","default_code","display_name","categ_id","product_tmpl_id"],
                       "limit":len(product_ids)+10})
        prod_map = {p["id"]:p for p in products}
        tmpl_ids = list({p["product_tmpl_id"][0] for p in products
                         if isinstance(p.get("product_tmpl_id"),list)})
        tmpl_map = {}
        if tmpl_ids:
            try:
                tmpls = _x(u,db,uid,ak,"product.template","search_read",
                           [[["id","in",tmpl_ids]]],
                           {"fields":["id","x_brand_category_id"],"limit":len(tmpl_ids)+10})
                tmpl_map = {t_["id"]:t_ for t_ in tmpls}
            except Exception:
                tmpl_map = {}
        rows = []
        for line in lines:
            oid = line["order_id"][0] if isinstance(line.get("order_id"),list) else None
            pid = line["product_id"][0] if isinstance(line.get("product_id"),list) else None
            order = order_map.get(oid,{}); prod = prod_map.get(pid,{})
            raw_date = order.get("date_order") or ""
            try: date_str = datetime.strptime(raw_date,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except: date_str = raw_date[:10] if raw_date else ""
            partner = order.get("partner_id")
            vendor  = str(partner[1]) if isinstance(partner,list) and len(partner)>1 else (str(partner) if partner else "")
            categ    = prod.get("categ_id")
            category = str(categ[1]) if isinstance(categ,list) and len(categ)>1 else (str(categ) if categ else "")
            brand_category = ""
            tmpl_ref = prod.get("product_tmpl_id")
            if isinstance(tmpl_ref,list) and tmpl_ref:
                tmpl = tmpl_map.get(tmpl_ref[0],{})
                bc   = tmpl.get("x_brand_category_id")
                if isinstance(bc,list): brand_category = str(bc[1]) if len(bc)>1 else ""
                elif bc: brand_category = str(bc)
            qty        = float(line.get("product_qty") or 0)
            unit_price = float(line.get("price_unit") or 0)
            subtotal   = round(qty * unit_price, 2)
            rows.append({
                C_SYSTEM:    system_key,
                C_DATE:      date_str,
                C_PO:        str(order.get("name") or ""),
                C_VENDOR:    vendor,
                C_BRAND_CAT: brand_category,
                C_CATEGORY:  category,
                C_MODEL:     str(prod.get("default_code") or ""),
                C_PRODUCT:   str(prod.get("display_name") or ""),
                C_QTY:       qty,
                C_UNIT_PRICE: unit_price,
                C_SUBTOTAL:  subtotal,
            })
        if not rows:
            return empty_df
        df = pd.DataFrame(rows)
        # keep numeric columns numeric
        for c in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
            df[c] = _to_num(df[c])
        # stringify text columns
        for c in [col for col in df.columns if col not in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]]:
            df[c] = df[c].fillna("").astype(str)
        return df.sort_values(by=C_DATE, ascending=False).reset_index(drop=True)
    except Exception:
        return empty_df

# ─────────────────────────────────────────────────────────────────────────────
# FETCH PURCHASE HISTORY FOR ALL SYSTEMS
# ─────────────────────────────────────────────────────────────────────────────
def fetch_all_systems_purchase_history(model_code, date_from, date_to,
                                        system_keys=None, progress_placeholder=None):
    """
    Fetch purchase history from all (or specified) systems in parallel.
    Returns merged DataFrame with canonical columns + System column.
    """
    keys = system_keys or SYSTEM_KEYS
    results = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(fetch_purchase_history_for_system, k, model_code, date_from, date_to): k
                for k in keys}
        for f in as_completed(futs):
            try:
                df = f.result()
                if df is not None and not df.empty:
                    results.append(df)
            except Exception:
                pass
    if not results:
        return pd.DataFrame(columns=[C_SYSTEM, C_DATE, C_PO, C_VENDOR, C_BRAND_CAT,
                                      C_CATEGORY, C_MODEL, C_PRODUCT, C_QTY,
                                      C_UNIT_PRICE, C_SUBTOTAL])
    merged = pd.concat(results, ignore_index=True)
    for c in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
        merged[c] = _to_num(merged[c])
    return merged.sort_values(by=C_DATE, ascending=False).reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# FETCH SALES HISTORY FOR ONE SYSTEM  ← NEW canonical-column version
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_sales_history_for_system(system_key, model_code, date_from, date_to):
    """
    Returns a DataFrame with canonical English column names.
    Columns: System, Date, SO, Customer, Brand Category, Category,
             Model Code, Product, Qty, Unit Price, Subtotal
    """
    empty_cols = [C_SYSTEM, C_DATE, C_SO, C_CUSTOMER, C_BRAND_CAT, C_CATEGORY,
                  C_MODEL, C_PRODUCT, C_QTY, C_UNIT_PRICE, C_SUBTOTAL]
    empty_df = pd.DataFrame(columns=empty_cols)
    cfg = st.secrets.get(system_key)
    if not cfg:
        return empty_df
    uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not uid:
        return empty_df
    u=cfg["url"]; db=cfg["db"]; ak=cfg["api_key"]
    try:
        date_from_dt = f"{date_from} 00:00:00"
        date_to_dt   = f"{date_to} 23:59:59"
        line_domain = [
            ["order_id.state","in",["sale","done"]],
            ["order_id.date_order",">=",date_from_dt],
            ["order_id.date_order","<=",date_to_dt],
        ]
        if model_code and str(model_code).strip():
            line_domain.append(["product_id.default_code","=",str(model_code).strip()])
        lines = _x(u,db,uid,ak,"sale.order.line","search_read",
                   [line_domain],
                   {"fields":["order_id","product_id","product_uom_qty","price_unit"],
                    "limit":20000,"order":"order_id desc"})
        if not lines:
            return empty_df
        order_ids   = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"),list)})
        product_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"),list)})
        orders = _x(u,db,uid,ak,"sale.order","search_read",
                    [[["id","in",order_ids]]],
                    {"fields":["id","name","partner_id","date_order"],"limit":len(order_ids)+10})
        order_map = {o["id"]:o for o in orders}
        products = _x(u,db,uid,ak,"product.product","search_read",
                      [[["id","in",product_ids]]],
                      {"fields":["id","default_code","display_name","categ_id","product_tmpl_id"],
                       "limit":len(product_ids)+10})
        prod_map = {p["id"]:p for p in products}
        tmpl_ids = list({p["product_tmpl_id"][0] for p in products
                         if isinstance(p.get("product_tmpl_id"),list)})
        tmpl_map = {}
        if tmpl_ids:
            try:
                tmpls = _x(u,db,uid,ak,"product.template","search_read",
                           [[["id","in",tmpl_ids]]],
                           {"fields":["id","x_brand_category_id"],"limit":len(tmpl_ids)+10})
                tmpl_map = {t_["id"]:t_ for t_ in tmpls}
            except Exception:
                tmpl_map = {}
        rows = []
        for line in lines:
            oid = line["order_id"][0] if isinstance(line.get("order_id"),list) else None
            pid = line["product_id"][0] if isinstance(line.get("product_id"),list) else None
            order = order_map.get(oid,{}); prod = prod_map.get(pid,{})
            raw_date = order.get("date_order") or ""
            try: date_str = datetime.strptime(raw_date,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except: date_str = raw_date[:10] if raw_date else ""
            partner  = order.get("partner_id")
            customer = str(partner[1]) if isinstance(partner,list) and len(partner)>1 else (str(partner) if partner else "")
            categ    = prod.get("categ_id")
            category = str(categ[1]) if isinstance(categ,list) and len(categ)>1 else (str(categ) if categ else "")
            brand_category = ""
            tmpl_ref = prod.get("product_tmpl_id")
            if isinstance(tmpl_ref,list) and tmpl_ref:
                tmpl = tmpl_map.get(tmpl_ref[0],{})
                bc   = tmpl.get("x_brand_category_id")
                if isinstance(bc,list): brand_category = str(bc[1]) if len(bc)>1 else ""
                elif bc: brand_category = str(bc)
            qty        = float(line.get("product_uom_qty") or 0)
            unit_price = float(line.get("price_unit") or 0)
            subtotal   = round(qty * unit_price, 2)
            rows.append({
                C_SYSTEM:    system_key,
                C_DATE:      date_str,
                C_SO:        str(order.get("name") or ""),
                C_CUSTOMER:  customer,
                C_BRAND_CAT: brand_category,
                C_CATEGORY:  category,
                C_MODEL:     str(prod.get("default_code") or ""),
                C_PRODUCT:   str(prod.get("display_name") or ""),
                C_QTY:       qty,
                C_UNIT_PRICE: unit_price,
                C_SUBTOTAL:  subtotal,
            })
        if not rows:
            return empty_df
        df = pd.DataFrame(rows)
        for c in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
            df[c] = _to_num(df[c])
        for c in [col for col in df.columns if col not in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]]:
            df[c] = df[c].fillna("").astype(str)
        return df.sort_values(by=C_DATE, ascending=False).reset_index(drop=True)
    except Exception:
        return empty_df

# ─────────────────────────────────────────────────────────────────────────────
# FETCH SALES HISTORY FOR ALL SYSTEMS
# ─────────────────────────────────────────────────────────────────────────────
def fetch_all_systems_sales_history(model_code, date_from, date_to,
                                     system_keys=None):
    """
    Fetch sales history from all (or specified) systems in parallel.
    Returns merged DataFrame with canonical columns + System column.
    """
    keys = system_keys or SYSTEM_KEYS
    results = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(fetch_sales_history_for_system, k, model_code, date_from, date_to): k
                for k in keys}
        for f in as_completed(futs):
            try:
                df = f.result()
                if df is not None and not df.empty:
                    results.append(df)
            except Exception:
                pass
    if not results:
        return pd.DataFrame(columns=[C_SYSTEM, C_DATE, C_SO, C_CUSTOMER, C_BRAND_CAT,
                                      C_CATEGORY, C_MODEL, C_PRODUCT, C_QTY,
                                      C_UNIT_PRICE, C_SUBTOTAL])
    merged = pd.concat(results, ignore_index=True)
    for c in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
        merged[c] = _to_num(merged[c])
    return merged.sort_values(by=C_DATE, ascending=False).reset_index(drop=True)

# ─────────────────────────────────────────────────────────────────────────────
# KEEP OLD FETCH NAMES AS ALIASES (backward compat)
# ─────────────────────────────────────────────────────────────────────────────
def fetch_swag_purchase_history(model_code, date_from, date_to):
    return fetch_purchase_history_for_system("SWAG", model_code, date_from, date_to)

def fetchswagsaleshistory(modelcode, datefrom, dateto):
    return fetch_sales_history_for_system("SWAG", modelcode, datefrom, dateto)

# ─────────────────────────────────────────────────────────────────────────────
# PRICE HISTORY
# ─────────────────────────────────────────────────────────────────────────────
def record_price_snapshot(df):
    if C_SALE_PRICE not in df.columns: return
    ok = df[df["_status"]=="OK"] if "_status" in df.columns else df
    if ok.empty: return
    ts = datetime.now().strftime("%H:%M:%S")
    for _, row in ok.iterrows():
        k = f"{row.get(C_SYSTEM,'?')}|{row.get(C_MODEL,'?')}"
        st.session_state.price_history.setdefault(k,[]).append({"time":ts,"price":float(row.get(C_SALE_PRICE,0))})

def build_price_history_df():
    hist = st.session_state.price_history
    if not hist: return pd.DataFrame()
    all_t = sorted({e["time"] for v in hist.values() for e in v})
    recs  = []
    for ts in all_t:
        row = {"time":ts}
        for k, entries in hist.items():
            px = [e["price"] for e in entries if e["time"]==ts]
            row[k] = px[-1] if px else None
        recs.append(row)
    return pd.DataFrame(recs).set_index("time")

# ─────────────────────────────────────────────────────────────────────────────
# QTY DISPLAY HELPER
# ─────────────────────────────────────────────────────────────────────────────
def get_qty_display(qty, lang="EN"):
    try:
        v = float(qty)
        if pd.isna(v) or v == 0:
            return "❌ " + ("غير متوفر" if lang == "AR" else "Not Available")
        return f"{int(v):,}"
    except Exception:
        return "❌ " + ("غير متوفر" if lang == "AR" else "Not Available")

# ─────────────────────────────────────────────────────────────────────────────
# PAGINATION HELPER
# ─────────────────────────────────────────────────────────────────────────────
def paginate_df(df, page_key, page_size=PAGE_SIZE):
    """
    Returns the current page slice of df.
    Renders Previous/Next buttons and page indicator.
    page_key: unique string for session state (e.g. "page_total")
    """
    total_rows = len(df)
    if total_rows == 0:
        return df
    total_pages = max(1, (total_rows + page_size - 1) // page_size)
    if page_key not in st.session_state:
        st.session_state[page_key] = 0
    # Clamp page index
    st.session_state[page_key] = min(st.session_state[page_key], total_pages - 1)
    st.session_state[page_key] = max(0, st.session_state[page_key])
    current_page = st.session_state[page_key]
    start = current_page * page_size
    end   = min(start + page_size, total_rows)
    page_df = df.iloc[start:end]
    # Pagination controls
    pc1, pc2, pc3, pc4, pc5 = st.columns([1, 1, 2, 1, 1])
    with pc1:
        if st.button(f"⏮ {t('First','الأول')}", key=f"{page_key}_first",
                     disabled=(current_page == 0)):
            st.session_state[page_key] = 0; st.rerun()
    with pc2:
        if st.button(f"◀ {t('Prev','السابق')}", key=f"{page_key}_prev",
                     disabled=(current_page == 0)):
            st.session_state[page_key] -= 1; st.rerun()
    with pc3:
        st.markdown(
            f"<div class='pagination-bar'>"
            f"<span class='page-info'>"
            f"{t('Page','صفحة')} {current_page+1} / {total_pages}"
            f"</span>"
            f"<span style='color:#8e9aaf;font-size:0.8rem;'>"
            f"({start+1}–{end} {t('of','من')} {total_rows:,} {t('rows','صف')})"
            f"</span>"
            f"</div>", unsafe_allow_html=True)
    with pc4:
        if st.button(f"▶ {t('Next','التالي')}", key=f"{page_key}_next",
                     disabled=(current_page >= total_pages - 1)):
            st.session_state[page_key] += 1; st.rerun()
    with pc5:
        if st.button(f"⏭ {t('Last','الأخير')}", key=f"{page_key}_last",
                     disabled=(current_page >= total_pages - 1)):
            st.session_state[page_key] = total_pages - 1; st.rerun()
    return page_df

# ─────────────────────────────────────────────────────────────────────────────
# PREMIUM TABLE RENDERER  ← now receives already-translated display df
# ─────────────────────────────────────────────────────────────────────────────
def render_premium_table(df_show, first_col_accent=True, page_key=None, page_size=PAGE_SIZE):
    """
    Render a styled HTML table.
    df_show: DataFrame with display-language column names (use df_for_display() before calling).
    page_key: if provided, pagination is applied.
    """
    if df_show is None or df_show.empty:
        st.info(t("No data available.", "لا توجد بيانات متاحة."))
        return
    # Apply pagination if requested
    if page_key:
        df_show = paginate_df(df_show, page_key, page_size)
    cols     = df_show.columns.tolist()
    dir_attr = f'dir="{get_dir()}"' if get_lang() == "AR" else ""
    th_html  = "".join(f"<th>{col}</th>" for col in cols)
    tbody_rows = []
    for _, row in df_show.iterrows():
        cells = []
        for ci, (col, val) in enumerate(row.items()):
            cell_class = "accent-cell" if first_col_accent and ci == 0 else ""
            cells.append(f"<td class='{cell_class}'>{val}</td>")
        tbody_rows.append(f"<tr>{''.join(cells)}</tr>")
    tbody_html = "".join(tbody_rows)
    st.markdown(f"""
    <div class="table-wrapper">
        <table class="premium-table" {dir_attr}>
            <thead><tr>{th_html}</tr></thead>
            <tbody>{tbody_html}</tbody>
        </table>
    </div>
    <style>
    .premium-table td.accent-cell {{ font-weight:700; color:#3498db; }}
    </style>
    """, unsafe_allow_html=True)
    if not page_key:
        st.caption(f"📊 {len(df_show)} {t('rows','صفوف')}")

# ─────────────────────────────────────────────────────────────────────────────
# DISPLAY DF  ← works on canonical-column DataFrames
# ─────────────────────────────────────────────────────────────────────────────
def display_df(df, thresh=0, table_key="tbl"):
    """
    Main table display with filters, sorting, search, and pagination.
    Input df must have canonical English column names.
    """
    if df is None or df.empty:
        st.info(t("No data.", "لا بيانات.")); return
    work = df.copy()
    fc = st.columns([2,2,2,1.5])
    has_sys = C_SYSTEM in work.columns
    has_br  = C_BRANCH in work.columns
    if has_sys:
        all_sys = sorted(work[C_SYSTEM].dropna().unique().tolist())
        # show display names
        sys_display = {k: get_system_name(k) for k in all_sys}
        sys_opts    = [sys_display.get(k, k) for k in all_sys]
        with fc[0]:
            sel_sys_disp = st.multiselect(f"🏢 {t('Company','الشركة')}", options=sys_opts,
                                          default=sys_opts, key=f"{table_key}_sys")
        # map back to keys
        disp_to_key = {v: k for k, v in sys_display.items()}
        sel_sys = [disp_to_key.get(d, d) for d in sel_sys_disp]
        if sel_sys: work = work[work[C_SYSTEM].isin(sel_sys)]
    if has_br:
        all_br = sorted(work[C_BRANCH].dropna().unique().tolist())
        with fc[1]:
            sel_br = st.multiselect(f"🏪 {t('Branch','الفرع')}", options=all_br,
                                    default=all_br, key=f"{table_key}_br")
        if sel_br: work = work[work[C_BRANCH].isin(sel_br)]
    with fc[2]:
        q = st.text_input(f"🔍 {t('Search model / product','بحث موديل / منتج')}",
                          value="", placeholder=t("e.g. XP6013","مثال: XP6013"),
                          key=f"{table_key}_q").strip()
    if q:
        ql   = q.lower()
        mask = pd.Series([False]*len(work), index=work.index)
        for col in [C_MODEL, C_PRODUCT, C_LOCATION]:
            if col in work.columns:
                mask = mask | work[col].fillna("").str.lower().str.contains(ql, regex=False)
        work = work[mask]
    with fc[3]:
        sortable = [c for c in work.columns if c != "_status"]
        sort_by  = st.selectbox(f"↕️ {t('Sort by','ترتيب')}", options=["—"]+sortable,
                                index=0, key=f"{table_key}_sort")
    if sort_by and sort_by != "—" and sort_by in work.columns:
        try:
            work = work.sort_values(
                by=sort_by,
                key=lambda s: pd.to_numeric(s,errors="coerce").fillna(0)
                              if pd.api.types.is_numeric_dtype(pd.to_numeric(s,errors="coerce"))
                              else s, ascending=True)
        except Exception:
            work = work.sort_values(by=sort_by)
    if work.empty:
        st.warning(t("⚠️ No rows match your filters.","لا توجد نتائج بعد الفلتر.")); return
    if C_ON_HAND in work.columns:
        raw_q = _to_num(work[C_ON_HAND])
        mn, mx = int(raw_q.min() or 0), int(raw_q.max() or 0)
        if mx > mn:
            qr = st.slider(f"📦 {t('Qty range','نطاق الكمية')}", min_value=mn, max_value=mx,
                           value=(mn,mx), key=f"{table_key}_qrange")
            raw_q2 = _to_num(work[C_ON_HAND])
            work   = work[(raw_q2>=qr[0])&(raw_q2<=qr[1])]
    ok_work = work[work["_status"]=="OK"] if "_status" in work.columns else work
    sm1,sm2,sm3,sm4 = st.columns(4)
    sm1.metric(t("Rows","الصفوف"), len(work))
    if C_ON_HAND in ok_work.columns:
        sm2.metric(t("Total Qty","إجمالي الكمية"), int(_to_num(ok_work[C_ON_HAND]).sum()))
    if C_SALE_PRICE in ok_work.columns:
        vp = _to_num(ok_work[C_SALE_PRICE])
        sm3.metric(t("Avg Price","متوسط السعر"),
                   f"{vp[vp>0].mean():.2f} SAR" if not vp[vp>0].empty else "—")
    if has_sys:
        sm4.metric(t("Companies","الشركات"), ok_work[C_SYSTEM].nunique() if C_SYSTEM in ok_work.columns else 0)
    # Prepare display copy (translate columns, format values)
    show = work.drop(columns=["_status"], errors="ignore").copy()
    if C_SALE_PRICE in show.columns:
        show[C_SALE_PRICE] = _to_num(show[C_SALE_PRICE]).map(lambda v: f"{v:.2f} SAR" if v else "—")
    if C_ON_HAND in show.columns:
        _lang = get_lang()
        show[C_ON_HAND] = _to_num(work[C_ON_HAND] if C_ON_HAND in work.columns else show[C_ON_HAND]).map(
            lambda v: get_qty_display(v, _lang))
    # Translate System key → display name
    if C_SYSTEM in show.columns:
        show[C_SYSTEM] = show[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)
    # Translate column headers
    show = df_for_display(show)
    render_premium_table(show, page_key=f"page_{table_key}")

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    _,_,lc = st.columns([2,1,0.5])
    with lc:
        lg = st.radio("",["EN","AR"],horizontal=True,
                      index=0 if get_lang()=="EN" else 1,
                      label_visibility="collapsed",key="llr")
        if lg!=get_lang(): st.session_state.lang=lg; st.rerun()
    _,col,_ = st.columns([1,1.1,1])
    with col:
        st.markdown("""
        <div style='display:flex;flex-direction:column;align-items:center;padding:20px 0 8px;'>
            <div class='login-orb'>📊</div>
            <div class='login-title'>SWAG Dashboard</div>
            <div class='login-subtitle'>Real-time Stock &amp; Price · 4 Odoo Systems</div>
        </div>""", unsafe_allow_html=True)
        wm = t("🌙 Welcome — Sign in to continue","🌙 مرحباً بك — سجّل دخولك للمتابعة")
        st.markdown(f"<div class='info-banner'>{wm}</div>", unsafe_allow_html=True)
        st.markdown("<div class='login-card'>", unsafe_allow_html=True)
        with st.form("lf", clear_on_submit=False):
            em  = st.text_input(t("📧 Email","📧 البريد الإلكتروني"), placeholder="you@swag.com.sa")
            pw  = st.text_input(t("🔑 Password","🔑 كلمة المرور"), type="password", placeholder="••••••••")
            st.markdown("<br>", unsafe_allow_html=True)
            sub = st.form_submit_button(t("🚀 Sign In","🚀 تسجيل الدخول"),
                                        use_container_width=True, type="primary")
        st.markdown("</div>", unsafe_allow_html=True)
        if sub:
            if not em or not pw:
                st.error(t("Fill in both fields.","يرجى ملء جميع الحقول.")); return
            if "LOGIN" not in st.secrets:
                st.error("❌ [LOGIN] section missing in secrets.toml"); return
            cfg = st.secrets["LOGIN"]
            if "url" not in cfg or "db" not in cfg:
                st.error("❌ LOGIN.url or LOGIN.db missing in secrets.toml"); return
            with st.spinner(t("⚡ Signing in…","⚡ جارٍ تسجيل الدخول…")):
                try:
                    proxy = xmlrpc.client.ServerProxy(f"{cfg['url']}/xmlrpc/2/common", allow_none=True)
                    uid   = proxy.authenticate(cfg["db"], em, pw, {})
                    if uid:
                        token = _make_token(em)
                        st.query_params["u"] = em
                        st.query_params["t"] = token
                        st.session_state.authenticated = True
                        st.session_state.user_email    = em
                        time.sleep(0.3); st.balloons(); st.rerun()
                    else:
                        st.error(t("❌ Wrong email or password.","❌ بريد إلكتروني أو كلمة مرور خاطئة."))
                except Exception as e:
                    st.error(f"❌ Connection error: {e}")
        st.markdown("""
        <div class='footer' style='margin-top:24px;'>
        © 2025 SWAG Fashion · Powered by Odoo · Built with ❤️
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────────────────────────────────────────
def do_logout():
    try: st.query_params.clear()
    except Exception: pass
    st.session_state.authenticated = False
    st.session_state.user_email    = ""
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PREMIUM KPI HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _premium_kpi_card(icon, value, label, trend=""):
    trend_html = f"<div class='kpi-trend'>{trend}</div>" if trend else ""
    return (f"<div class='kpi-card'>"
            f"<span class='kpi-icon'>{icon}</span>"
            f"<div class='kpi-value'>{value}</div>"
            f"<div class='kpi-label'>{label}</div>"
            f"{trend_html}</div>")

def _render_kpi_grid(cards):
    st.markdown(f"<div class='kpi-grid'>{''.join(cards)}</div>", unsafe_allow_html=True)

def _section_header(title, icon="📊"):
    st.markdown(f"<div class='section-header'>{icon} {title}</div>", unsafe_allow_html=True)

def _chart_card_open(title, icon="📈"):
    st.markdown(f"<div class='chart-card'><div class='chart-title'>"
                f"<span class='accent'></span>{icon} {title}</div>", unsafe_allow_html=True)

def _chart_card_close():
    st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# ALTAIR HELPERS
# ─────────────────────────────────────────────────────────────────────────────
_ALT_CONFIG = {
    "background": "transparent",
    "view": {"stroke": "transparent"},
    "axis": {
        "labelColor": "#8e9aaf", "titleColor": "#e8edf2",
        "gridColor": "rgba(255,255,255,0.1)",
        "domainColor": "rgba(52,152,219,0.3)",
        "tickColor": "rgba(52,152,219,0.3)",
        "labelFontSize": 11, "titleFontSize": 12,
    },
    "legend": {"labelColor": "#8e9aaf", "titleColor": "#e8edf2"},
    "title": {"color": "#e8edf2"},
}
_PALETTE = ["#3498db","#2ecc71","#e74c3c","#f1c40f","#9b59b6","#1abc9c","#e67e22","#34495e"]

def _alt_bar_chart(df, x_field, y_field, tooltip_fmt=",.0f", color="#3498db",
                   sort_order="-y", label_angle=-35, height=300):
    tooltip_label = f"{y_field}_fmt"
    plot_df = df.copy()
    plot_df[tooltip_label] = _to_num(plot_df[y_field]).map(lambda v: f"{v:{tooltip_fmt}}")
    chart = (
        alt.Chart(plot_df)
        .mark_bar(cornerRadiusTopLeft=8, cornerRadiusTopRight=8, opacity=0.9)
        .encode(
            x=alt.X(f"{x_field}:N", sort=sort_order,
                    axis=alt.Axis(labelAngle=label_angle, labelLimit=130), title=None),
            y=alt.Y(f"{y_field}:Q", title=y_field, axis=alt.Axis(grid=True)),
            tooltip=[alt.Tooltip(f"{x_field}:N", title=x_field),
                     alt.Tooltip(f"{tooltip_label}:N", title=y_field)],
            color=alt.condition(
                alt.datum[y_field] == plot_df[y_field].max(),
                alt.value("#2ecc71"), alt.value(color)),
        )
        .properties(height=height)
        .configure(**_ALT_CONFIG)
        .interactive()
    )
    return chart

def _alt_line_chart(df, x_field, y_field, height=260, color="#3498db"):
    line = (alt.Chart(df).mark_line(color=color, strokeWidth=2.5, interpolate="monotone")
            .encode(
                x=alt.X(f"{x_field}:T", title=None, axis=alt.Axis(format="%b %d", labelAngle=-30)),
                y=alt.Y(f"{y_field}:Q", title=y_field),
                tooltip=[alt.Tooltip(f"{x_field}:T", title=t("Date","التاريخ"), format="%Y-%m-%d"),
                         alt.Tooltip(f"{y_field}:Q", title=y_field, format=",.0f")]))
    area  = (alt.Chart(df).mark_area(color=color, opacity=0.12, interpolate="monotone")
             .encode(x=alt.X(f"{x_field}:T"), y=alt.Y(f"{y_field}:Q")))
    points = (alt.Chart(df).mark_circle(color="#2ecc71", size=55, opacity=0.9)
              .encode(
                  x=alt.X(f"{x_field}:T"), y=alt.Y(f"{y_field}:Q"),
                  tooltip=[alt.Tooltip(f"{x_field}:T", title=t("Date","التاريخ"), format="%Y-%m-%d"),
                           alt.Tooltip(f"{y_field}:Q", title=y_field, format=",.0f")]))
    return ((area + line + points).properties(height=height).configure(**_ALT_CONFIG).interactive())

def _top10_altair(title, group_col, value_col, df, color="#3498db", tooltip_fmt=",.0f"):
    """
    group_col and value_col are CANONICAL column names.
    Translates only for display labels.
    """
    if df is None or df.empty or group_col not in df.columns or value_col not in df.columns:
        st.info(t("No data available.","لا توجد بيانات.")); return
    grp = (
        df.copy()
        .assign(**{group_col: df[group_col].replace("", f"({group_col} N/A)").fillna(f"({group_col} N/A)")})
        .groupby(group_col, as_index=False)[value_col]
        .sum()
        .sort_values(value_col, ascending=False)
        .head(10).reset_index(drop=True)
    )
    grp[value_col] = _to_num(grp[value_col])
    if grp.empty:
        st.info(t("No data.","لا توجد بيانات.")); return
    display_label = col_label(value_col) + " (Total)"
    grp[display_label] = grp[value_col].map(lambda v: f"{v:{tooltip_fmt}}")
    _chart_card_open(title, "")
    ch_col, tbl_col = st.columns([1.6,1])
    with ch_col:
        st.altair_chart(_alt_bar_chart(grp, x_field=group_col, y_field=value_col,
                                       tooltip_fmt=tooltip_fmt, color=color),
                        use_container_width=True)
    with tbl_col:
        disp = grp[[group_col, display_label]].copy()
        disp = disp.rename(columns={group_col: col_label(group_col)})
        render_premium_table(disp)
    _chart_card_close()

# ─────────────────────────────────────────────────────────────────────────────
# PLOTLY DONUT
# ─────────────────────────────────────────────────────────────────────────────
def _plotly_donut(labels, values, title="", height=360):
    if not _HAS_PLOTLY:
        st.info(t("Install plotly for donut charts.","قم بتثبيت plotly للرسوم الدائرية.")); return
    clean_pairs = []
    for lbl, val in zip(labels, values):
        try:
            lbl_s = str(lbl) if lbl is not None else "(N/A)"
            val_f = float(val) if val is not None else 0.0
            if val_f > 0: clean_pairs.append((lbl_s, val_f))
        except (TypeError, ValueError):
            continue
    if not clean_pairs:
        st.info(t("No data for chart.","لا بيانات للرسم.")); return
    clean_labels = [p[0] for p in clean_pairs]
    clean_values = [p[1] for p in clean_pairs]
    _colors = ["#3498db","#2ecc71","#e74c3c","#f1c40f","#9b59b6","#1abc9c","#e67e22","#34495e"]
    used_colors = (_colors * ((len(clean_labels)//len(_colors))+1))[:len(clean_labels)]
    try:
        fig = go.Figure(data=[go.Pie(
            labels=clean_labels, values=clean_values, hole=0.55,
            marker=dict(colors=used_colors, line=dict(color="#0a0c10", width=2)),
            textinfo="percent+label",
            textfont=dict(color="#e8edf2", size=11),
            hovertemplate="<b>%{label}</b><br>" + t("Value","القيمة") +
                          ": %{value:,.0f}<br>" + t("Share","الحصة") + ": %{percent}<extra></extra>",
            sort=True, direction="clockwise")])
        fig.update_layout(
            title_text=title, title_x=0.5, title_font_color="#e8edf2", title_font_size=13,
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            height=height, margin=dict(t=52,b=20,l=10,r=10),
            legend=dict(font=dict(color="#8e9aaf",size=10),
                        bgcolor="rgba(20,24,30,0.75)",
                        bordercolor="rgba(52,152,219,0.3)", borderwidth=1, orientation="v"),
            showlegend=True)
        st.plotly_chart(fig, use_container_width=True)
    except Exception as exc:
        st.warning(f"Chart render error: {exc}")

# ─────────────────────────────────────────────────────────────────────────────
# PURCHASE KPI ROW  (uses canonical column names)
# ─────────────────────────────────────────────────────────────────────────────
def _po_kpi_row(df, prefix=""):
    if df is None or df.empty:
        st.info(t("No purchase data available.","لا توجد بيانات مشتريات متاحة.")); return
    total_qty  = float(_to_num(df[C_QTY]).sum())      if C_QTY      in df.columns else 0.0
    total_amt  = float(_to_num(df[C_SUBTOTAL]).sum())  if C_SUBTOTAL in df.columns else 0.0
    n_vendors  = int(df[C_VENDOR].nunique())            if C_VENDOR   in df.columns else 0
    n_products = int(df[C_MODEL].nunique())             if C_MODEL    in df.columns else 0
    n_systems  = int(df[C_SYSTEM].nunique())            if C_SYSTEM   in df.columns else 0
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric(t("Total Qty","إجمالي الكمية"),    f"{total_qty:,.0f}")
    c2.metric(t("Total Amount","إجمالي المبلغ"), f"{total_amt:,.2f}")
    c3.metric(t("Vendors","الموردون"),           int(n_vendors))
    c4.metric(t("Products","المنتجات"),          int(n_products))
    c5.metric(t("Systems","الأنظمة"),            int(n_systems))

def _po_full_table(df, page_key=None):
    show = df.copy()
    if C_UNIT_PRICE in show.columns:
        show[C_UNIT_PRICE] = _to_num(show[C_UNIT_PRICE]).map(lambda v: f"{v:.2f} SAR")
    if C_SUBTOTAL in show.columns:
        show[C_SUBTOTAL]   = _to_num(show[C_SUBTOTAL]).map(lambda v: f"{v:,.2f} SAR")
    if C_QTY in show.columns:
        show[C_QTY]        = _to_num(show[C_QTY]).map(lambda v: f"{v:,.0f}")
    if C_SYSTEM in show.columns:
        show[C_SYSTEM]     = show[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)
    show = df_for_display(show)
    render_premium_table(show, page_key=page_key)

def _po_download_row(df, tag_suffix=""):
    dl1,dl2,_ = st.columns([1,1,2])
    dl1.download_button(t("⬇️ CSV","⬇️ CSV"),
                        df.to_csv(index=False).encode("utf-8-sig"),
                        dl_name(f"purchase{tag_suffix}","csv"), "text/csv",
                        use_container_width=True, key=f"dl_csv_{tag_suffix}_{id(df)}")
    dl2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel_purchase(df),
                        dl_name(f"purchase{tag_suffix}","xlsx"),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"dl_xlsx_{tag_suffix}_{id(df)}")

# ─────────────────────────────────────────────────────────────────────────────
# SALES KPI ROW  (uses canonical column names)
# ─────────────────────────────────────────────────────────────────────────────
def _sales_kpi_row(df):
    total_qty   = float(_to_num(df[C_QTY]).sum())      if C_QTY       in df.columns else 0.0
    total_amt   = float(_to_num(df[C_SUBTOTAL]).sum())  if C_SUBTOTAL  in df.columns else 0.0
    n_customers = int(df[C_CUSTOMER].nunique())          if C_CUSTOMER  in df.columns else 0
    n_products  = int(df[C_MODEL].nunique())             if C_MODEL     in df.columns else 0
    n_orders    = int(df[C_SO].nunique())                if C_SO        in df.columns else 0
    n_systems   = int(df[C_SYSTEM].nunique())            if C_SYSTEM    in df.columns else 0
    if C_UNIT_PRICE in df.columns:
        pos_prices = _to_num(df[C_UNIT_PRICE])
        pos_prices = pos_prices[pos_prices > 0]
        avg_price  = float(pos_prices.mean()) if not pos_prices.empty else 0.0
    else:
        avg_price = 0.0
    cards = [
        _premium_kpi_card("📦", f"{total_qty:,.0f}",  t("Total Qty Sold","إجمالي الكمية المباعة")),
        _premium_kpi_card("💰", f"{total_amt:,.2f}",  t("Total Sales (SAR)","إجمالي المبيعات")),
        _premium_kpi_card("👤", str(n_customers),     t("Customers","العملاء")),
        _premium_kpi_card("🏷️", str(n_products),      t("Products","المنتجات")),
        _premium_kpi_card("🧾", str(n_orders),        t("Orders","الطلبات")),
        _premium_kpi_card("🏢", str(n_systems),       t("Systems","الأنظمة")),
    ]
    _render_kpi_grid(cards)

# ─────────────────────────────────────────────────────────────────────────────
# SALES FULL TABLE HELPER
# ─────────────────────────────────────────────────────────────────────────────
def _sales_full_table(df, page_key=None):
    show = df.copy()
    if C_UNIT_PRICE in show.columns:
        show[C_UNIT_PRICE] = _to_num(show[C_UNIT_PRICE]).map(lambda v: f"{v:.2f} SAR")
    if C_SUBTOTAL in show.columns:
        show[C_SUBTOTAL]   = _to_num(show[C_SUBTOTAL]).map(lambda v: f"{v:,.2f} SAR")
    if C_QTY in show.columns:
        show[C_QTY]        = _to_num(show[C_QTY]).map(lambda v: f"{v:,.0f}")
    if C_SYSTEM in show.columns:
        show[C_SYSTEM]     = show[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)
    show = df_for_display(show)
    render_premium_table(show, page_key=page_key)

# ─────────────────────────────────────────────────────────────────────────────
# SALES ANALYTICS VIEW  ← all canonical columns, all 4 systems
# ─────────────────────────────────────────────────────────────────────────────
def show_sales_analytics():
    _section_header(t("Sales Analytics — All Systems","تحليلات المبيعات — كل الأنظمة"), "💰")
    st.markdown("<div class='info-banner'>📌 " +
                t("Sales orders from <b>all configured systems</b> (state: sale / done). "
                  "Use the System filter to narrow down.",
                  "أوامر البيع من <b>جميع الأنظمة المُعدَّة</b> (الحالة: مبيع / منجز). "
                  "استخدم فلتر النظام للتضييق.") +
                "</div>", unsafe_allow_html=True)

    default_from = datetime.now().date() - timedelta(days=365)
    default_to   = datetime.now().date()

    sf1,sf2,sf3,sf4,sf5 = st.columns([1.2,1,1,1.4,1.4])
    with sf1:
        sa_model_input = st.text_input(
            f"🔖 {t('Model Code (optional)','رمز الموديل (اختياري)')}",
            placeholder=t("e.g. RVT196 — blank = all","مثال: RVT196 — فارغ = الكل"),
            key="sa_model_input").strip()
    with sf2:
        sa_date_from = st.date_input(f"📅 {t('From','من')}", value=default_from, key="sa_date_from")
    with sf3:
        sa_date_to   = st.date_input(f"📅 {t('To','إلى')}",  value=default_to,   key="sa_date_to")
    with sf4:
        # System filter
        sys_display_names = [get_system_name(k) for k in SYSTEM_KEYS]
        sa_system_sel = st.multiselect(
            f"🏢 {t('System','النظام')}", options=sys_display_names, default=sys_display_names,
            key="sa_system_sel")
        # map display → key
        _disp2key_s = {get_system_name(k): k for k in SYSTEM_KEYS}
        selected_system_keys_s = [_disp2key_s.get(d, d) for d in sa_system_sel] or SYSTEM_KEYS
    with sf5:
        cached_sa     = st.session_state.get("salesanalyticsdf")
        customer_opts = []
        if cached_sa is not None and not cached_sa.empty and C_CUSTOMER in cached_sa.columns:
            customer_opts = sorted(cached_sa[C_CUSTOMER].dropna().unique().tolist())
        sa_customer_sel = st.multiselect(
            f"👤 {t('Customer','العميل')}", options=customer_opts, default=[],
            placeholder=t("All Customers (default)","كل العملاء (افتراضي)"), key="sa_customer_sel")

    if st.button(f"🔍 {t('Fetch Sales Analytics','جلب تحليلات المبيعات')}",
                 type="primary", key="fetch_sa_btn"):
        with st.spinner(t("⚡ Fetching sales data from all systems…","⚡ جلب بيانات المبيعات من كل الأنظمة…")):
            fetched = fetch_all_systems_sales_history(
                model_code=None,
                date_from=sa_date_from.strftime("%Y-%m-%d"),
                date_to=sa_date_to.strftime("%Y-%m-%d"),
                system_keys=selected_system_keys_s)
        st.session_state.salesanalyticsdf = fetched
        st.session_state.page_sales = 0
        st.rerun()

    sa_full = st.session_state.get("salesanalyticsdf")
    if sa_full is None:
        st.info(t("👆 Set your date range and click **Fetch Sales Analytics** to load data.",
                  "👆 حدد نطاق التاريخ واضغط **جلب تحليلات المبيعات** لتحميل البيانات.")); return
    if sa_full.empty:
        st.info(t("No sales found for this period.","لا توجد مبيعات لهذه الفترة.")); return

    # Ensure numeric — always on canonical columns
    for nc in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
        if nc in sa_full.columns:
            sa_full[nc] = _to_num(sa_full[nc])

    # Apply system filter on canonical C_SYSTEM
    if C_SYSTEM in sa_full.columns:
        sa_df = sa_full[sa_full[C_SYSTEM].isin(selected_system_keys_s)].copy()
    else:
        sa_df = sa_full.copy()

    # Apply customer filter
    if sa_customer_sel and C_CUSTOMER in sa_df.columns:
        sa_df = sa_df[sa_df[C_CUSTOMER].isin(sa_customer_sel)].copy()

    # Apply model filter
    sa_df_model = None
    if sa_model_input and C_MODEL in sa_df.columns:
        sa_df_model = sa_df[sa_df[C_MODEL].str.upper() == sa_model_input.upper()].copy()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Sales KPIs","مؤشرات المبيعات"), "📊")
    _sales_kpi_row(sa_df)

    # Top 10 products by qty
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Top 10 Analytics","أفضل 10 تحليلات"), "🏆")

    prod_qty_grp = (
        sa_df.assign(**{C_MODEL: sa_df[C_MODEL].replace("","(No Code)").fillna("(No Code)")})
        .groupby([C_MODEL, C_PRODUCT], as_index=False)[C_QTY]
        .sum()
        .sort_values(C_QTY, ascending=False)
        .head(10).reset_index(drop=True)
    )
    prod_qty_grp[C_QTY] = _to_num(prod_qty_grp[C_QTY])
    prod_qty_grp["Total Qty Display"] = prod_qty_grp[C_QTY].map(lambda v: f"{v:,.0f}")

    _chart_card_open(t("Top 10 Products by Qty Sold","أعلى 10 منتجات حسب الكمية المباعة"), "🏆")
    pc1,pc2 = st.columns([1.6,1])
    with pc1:
        st.altair_chart(_alt_bar_chart(prod_qty_grp, x_field=C_MODEL, y_field=C_QTY,
                                       tooltip_fmt=",.0f", color="#2ecc71"),
                        use_container_width=True)
    with pc2:
        disp_pq = prod_qty_grp[[C_MODEL, C_PRODUCT, "Total Qty Display"]].rename(columns={
            C_MODEL: col_label(C_MODEL), C_PRODUCT: col_label(C_PRODUCT),
            "Total Qty Display": t("Total Qty","إجمالي الكمية")})
        render_premium_table(disp_pq)
    _chart_card_close()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Top 10 by amount
    prod_amt_grp = (
        sa_df.assign(**{C_MODEL: sa_df[C_MODEL].replace("","(No Code)").fillna("(No Code)")})
        .groupby([C_MODEL, C_PRODUCT], as_index=False)[C_SUBTOTAL]
        .sum()
        .sort_values(C_SUBTOTAL, ascending=False)
        .head(10).reset_index(drop=True)
    )
    prod_amt_grp[C_SUBTOTAL] = _to_num(prod_amt_grp[C_SUBTOTAL])
    prod_amt_grp["Total SAR Display"] = prod_amt_grp[C_SUBTOTAL].map(lambda v: f"{v:,.2f}")

    _chart_card_open(t("Top 10 Products by Sales Amount","أعلى 10 منتجات حسب المبلغ"), "💰")
    pa1,pa2 = st.columns([1.6,1])
    with pa1:
        st.altair_chart(_alt_bar_chart(prod_amt_grp, x_field=C_MODEL, y_field=C_SUBTOTAL,
                                       tooltip_fmt=",.2f", color="#3498db"),
                        use_container_width=True)
    with pa2:
        disp_pa = prod_amt_grp[[C_MODEL, C_PRODUCT, "Total SAR Display"]].rename(columns={
            C_MODEL: col_label(C_MODEL), C_PRODUCT: col_label(C_PRODUCT),
            "Total SAR Display": t("Total SAR","إجمالي المبلغ")})
        render_premium_table(disp_pa)
    _chart_card_close()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Top 10 by system
    sys_grp = (
        sa_df.assign(**{C_SYSTEM: sa_df[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)})
        .groupby(C_SYSTEM, as_index=False)[C_SUBTOTAL].sum()
        .sort_values(C_SUBTOTAL, ascending=False).reset_index(drop=True)
    )
    sys_grp[C_SUBTOTAL] = _to_num(sys_grp[C_SUBTOTAL])
    _chart_card_open(t("Sales by System","المبيعات حسب النظام"), "🏢")
    sg1,sg2 = st.columns([1.6,1])
    with sg1:
        st.altair_chart(_alt_bar_chart(sys_grp, x_field=C_SYSTEM, y_field=C_SUBTOTAL,
                                       tooltip_fmt=",.2f", color="#9b59b6"),
                        use_container_width=True)
    with sg2:
        sys_grp["Amount Display"] = sys_grp[C_SUBTOTAL].map(lambda v: f"{v:,.2f} SAR")
        disp_sg = sys_grp[[C_SYSTEM, "Amount Display"]].rename(columns={
            C_SYSTEM: col_label(C_SYSTEM), "Amount Display": t("Total SAR","إجمالي المبلغ")})
        render_premium_table(disp_sg)
    _chart_card_close()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _top10_altair(t("Top 10 Brand Categories by Qty","أعلى 10 فئات علامة تجارية حسب الكمية"),
                  C_BRAND_CAT, C_QTY, sa_df, color="#9b59b6")
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _top10_altair(t("Top 10 Categories by Qty","أعلى 10 فئات حسب الكمية"),
                  C_CATEGORY, C_QTY, sa_df, color="#e74c3c")
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Top customers
    cust_grp = (
        sa_df.assign(**{C_CUSTOMER: sa_df[C_CUSTOMER].replace("","(No Customer)").fillna("(No Customer)")})
        .groupby(C_CUSTOMER, as_index=False)[C_SUBTOTAL]
        .sum()
        .sort_values(C_SUBTOTAL, ascending=False)
        .head(10).reset_index(drop=True)
    )
    cust_grp[C_SUBTOTAL] = _to_num(cust_grp[C_SUBTOTAL])
    cust_grp["Total SAR Display"] = cust_grp[C_SUBTOTAL].map(lambda v: f"{v:,.2f}")

    _chart_card_open(t("Top 10 Customers by Sales Amount","أعلى 10 عملاء حسب المبلغ"), "👤")
    cc1,cc2 = st.columns([1.6,1])
    with cc1:
        st.altair_chart(_alt_bar_chart(cust_grp, x_field=C_CUSTOMER, y_field=C_SUBTOTAL,
                                       tooltip_fmt=",.2f", color="#f1c40f"),
                        use_container_width=True)
    with cc2:
        disp_cg = cust_grp[[C_CUSTOMER, "Total SAR Display"]].rename(columns={
            C_CUSTOMER: col_label(C_CUSTOMER),
            "Total SAR Display": t("Total SAR","إجمالي المبلغ")})
        render_premium_table(disp_cg)
    _chart_card_close()

    # Donut share analysis
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Sales Share Analysis","تحليل حصص المبيعات"), "🥧")
    pie1,pie2,pie3 = st.columns(3)

    with pie1:
        bc_share = (sa_df.assign(**{C_BRAND_CAT: sa_df[C_BRAND_CAT].replace("","(No Brand)").fillna("(No Brand)")})
                    .groupby(C_BRAND_CAT, as_index=False)[C_SUBTOTAL].sum()
                    .sort_values(C_SUBTOTAL, ascending=False))
        bc_share[C_SUBTOTAL] = _to_num(bc_share[C_SUBTOTAL])
        _plotly_donut(bc_share[C_BRAND_CAT].tolist(), bc_share[C_SUBTOTAL].tolist(),
                      title=t("Brand Category Share","حصة الفئة التجارية"))
    with pie2:
        cat_share = (sa_df.assign(**{C_CATEGORY: sa_df[C_CATEGORY].replace("","(No Category)").fillna("(No Category)")})
                     .groupby(C_CATEGORY, as_index=False)[C_SUBTOTAL].sum()
                     .sort_values(C_SUBTOTAL, ascending=False))
        cat_share[C_SUBTOTAL] = _to_num(cat_share[C_SUBTOTAL])
        _plotly_donut(cat_share[C_CATEGORY].tolist(), cat_share[C_SUBTOTAL].tolist(),
                      title=t("Category Share","حصة الفئة"))
    with pie3:
        cust_all = (sa_df.assign(**{C_CUSTOMER: sa_df[C_CUSTOMER].replace("","(No Customer)").fillna("(No Customer)")})
                    .groupby(C_CUSTOMER, as_index=False)[C_SUBTOTAL].sum()
                    .sort_values(C_SUBTOTAL, ascending=False))
        cust_all[C_SUBTOTAL] = _to_num(cust_all[C_SUBTOTAL])
        if not cust_all.empty:
            top10c   = cust_all.head(10)
            others_v = float(cust_all.iloc[10:][C_SUBTOTAL].sum()) if len(cust_all)>10 else 0
            p_labels = top10c[C_CUSTOMER].tolist()
            p_vals   = top10c[C_SUBTOTAL].tolist()
            if others_v > 0: p_labels.append("Others"); p_vals.append(others_v)
            _plotly_donut(p_labels, p_vals, title=t("Customer Share (Top 10)","حصة العملاء (أعلى 10)"))

    # Time series
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Sales Trend Over Time","اتجاه المبيعات عبر الزمن"), "📈")
    ts_col1,ts_col2 = st.columns(2)

    with ts_col1:
        _chart_card_open(t("Qty Sold Over Time","الكمية المباعة عبر الزمن"), "📦")
        ts_qty = (sa_df.assign(Date=pd.to_datetime(sa_df[C_DATE], errors="coerce"))
                  .dropna(subset=["Date"])
                  .groupby("Date", as_index=False)[C_QTY].sum()
                  .sort_values("Date"))
        ts_qty[C_QTY] = _to_num(ts_qty[C_QTY])
        if not ts_qty.empty:
            st.altair_chart(_alt_line_chart(ts_qty, "Date", C_QTY, height=240, color="#2ecc71"),
                            use_container_width=True)
        _chart_card_close()

    with ts_col2:
        _chart_card_open(t("Sales Amount Over Time","مبلغ المبيعات عبر الزمن"), "💰")
        ts_amt = (sa_df.assign(Date=pd.to_datetime(sa_df[C_DATE], errors="coerce"))
                  .dropna(subset=["Date"])
                  .groupby("Date", as_index=False)[C_SUBTOTAL].sum()
                  .sort_values("Date"))
        ts_amt[C_SUBTOTAL] = _to_num(ts_amt[C_SUBTOTAL])
        if not ts_amt.empty:
            st.altair_chart(_alt_line_chart(ts_amt, "Date", C_SUBTOTAL, height=240, color="#3498db"),
                            use_container_width=True)
        _chart_card_close()

    # Single model detail
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Single Model Sales Detail","تفاصيل مبيعات موديل واحد"), "🔍")

    if not sa_model_input:
        st.info(t("💡 Enter a **Model Code** in the filter above to see single-model sales analytics.",
                  "💡 أدخل **رمز الموديل** في الفلتر أعلاه لعرض تحليلات موديل واحد."))
    elif sa_df_model is not None and sa_df_model.empty:
        st.info(t(f"No sales records found for model **{sa_model_input}**.",
                  f"لا توجد سجلات مبيعات للموديل **{sa_model_input}**."))
    elif sa_df_model is not None:
        sm_qty  = float(_to_num(sa_df_model[C_QTY]).sum())
        sm_amt  = float(_to_num(sa_df_model[C_SUBTOTAL]).sum())
        sm_cust = int(sa_df_model[C_CUSTOMER].nunique()) if C_CUSTOMER in sa_df_model.columns else 0
        sm_cards = [
            _premium_kpi_card("📦", f"{sm_qty:,.0f}",  t("Total Qty (this model)","إجمالي الكمية (الموديل)")),
            _premium_kpi_card("💰", f"{sm_amt:,.2f}",  t("Total Sales (SAR)","إجمالي المبيعات")),
            _premium_kpi_card("👤", str(sm_cust),      t("Customers","العملاء")),
        ]
        _render_kpi_grid(sm_cards)

        st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
        _chart_card_open(f"{t('Sales Qty Over Time','كمية المبيعات عبر الزمن')} — {sa_model_input}", "📈")
        sm_ts = (sa_df_model.assign(Date=pd.to_datetime(sa_df_model[C_DATE], errors="coerce"))
                 .dropna(subset=["Date"])
                 .groupby("Date", as_index=False)[C_QTY].sum()
                 .sort_values("Date"))
        sm_ts[C_QTY] = _to_num(sm_ts[C_QTY])
        if not sm_ts.empty:
            st.altair_chart(_alt_line_chart(sm_ts, "Date", C_QTY, height=230, color="#2ecc71"),
                            use_container_width=True)
        _chart_card_close()

        st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
        sm_cust_grp = (
            sa_df_model.assign(**{C_CUSTOMER: sa_df_model[C_CUSTOMER].replace("","(No Customer)").fillna("(No Customer)")})
            .groupby(C_CUSTOMER, as_index=False)[C_QTY]
            .sum().sort_values(C_QTY, ascending=False).head(10).reset_index(drop=True))
        sm_cust_grp[C_QTY] = _to_num(sm_cust_grp[C_QTY])
        sm_cust_grp["Total Qty Display"] = sm_cust_grp[C_QTY].map(lambda v: f"{v:,.0f}")

        _chart_card_open(t("Top Customers for this Model","أعلى العملاء لهذا الموديل"), "👤")
        sc1,sc2 = st.columns([1.6,1])
        with sc1:
            st.altair_chart(_alt_bar_chart(sm_cust_grp, x_field=C_CUSTOMER, y_field=C_QTY,
                                           tooltip_fmt=",.0f", color="#9b59b6"),
                            use_container_width=True)
        with sc2:
            disp_smc = sm_cust_grp[[C_CUSTOMER, "Total Qty Display"]].rename(columns={
                C_CUSTOMER: col_label(C_CUSTOMER),
                "Total Qty Display": t("Total Qty","إجمالي الكمية")})
            render_premium_table(disp_smc)
        _chart_card_close()

        if not sm_cust_grp.empty:
            top_c    = sm_cust_grp.head(8)
            others_v = float(sm_cust_grp.iloc[8:][C_QTY].sum()) if len(sm_cust_grp)>8 else 0
            p_labels = top_c[C_CUSTOMER].tolist()
            p_vals   = top_c[C_QTY].tolist()
            if others_v > 0: p_labels.append("Others"); p_vals.append(others_v)
            _d1,_d2,_d3 = st.columns([1,1.2,1])
            with _d2:
                _plotly_donut(p_labels, p_vals, title=t("Customer Share","حصة العملاء"), height=320)

    # Full table + downloads
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Full Sales Detail","تفاصيل المبيعات الكاملة"), "📋")
    show_df = (sa_df_model if (sa_model_input and sa_df_model is not None and not sa_df_model.empty)
               else sa_df)
    _sales_full_table(show_df, page_key="page_sales_detail")

    st.markdown("<br>", unsafe_allow_html=True)
    sdl1,sdl2,_ = st.columns([1,1,2])
    tag_s = f"_{sa_model_input.upper()}" if sa_model_input else "_overall"
    export_df = (sa_df_model if (sa_model_input and sa_df_model is not None and not sa_df_model.empty)
                 else sa_df)
    sdl1.download_button(t("⬇️ CSV","⬇️ CSV"),
                         export_df.to_csv(index=False).encode("utf-8-sig"),
                         dl_name(f"sales{tag_s}","csv"), "text/csv",
                         use_container_width=True, key=f"sdl_csv{tag_s}")
    sdl2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel_sales(export_df),
                         dl_name(f"sales{tag_s}","xlsx"),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         use_container_width=True, key=f"sdl_xlsx{tag_s}")

# ─────────────────────────────────────────────────────────────────────────────
# PURCHASE ANALYTICS VIEW  ← all canonical columns, all 4 systems
# ─────────────────────────────────────────────────────────────────────────────
def show_purchase_analytics():
    _section_header(t("Purchase Analytics — All Systems","تحليلات المشتريات — كل الأنظمة"), "🛒")
    st.markdown("<div class='info-banner'>📌 " +
                t("Purchase orders from <b>all configured systems</b> (state: purchase / done). "
                  "Use the System filter to narrow down.",
                  "أوامر الشراء من <b>جميع الأنظمة المُعدَّة</b> (الحالة: مشترى / منجز). "
                  "استخدم فلتر النظام للتضييق.") +
                "</div>", unsafe_allow_html=True)

    default_from = datetime.now().date() - timedelta(days=365)
    default_to   = datetime.now().date()

    filt_col1,filt_col2,filt_col3,filt_col4,filt_col5 = st.columns([1.2,1,1,1.4,1.4])
    with filt_col1:
        po_model_input = st.text_input(
            f"🔖 {t('Model Code (optional)','رمز الموديل (اختياري)')}",
            placeholder=t("e.g. RVT196 — blank = all","مثال: RVT196 — فارغ = الكل"),
            key="po_model_input_v4").strip()
    with filt_col2:
        po_date_from = st.date_input(f"📅 {t('From','من')}", value=default_from, key="po_date_from_v4")
    with filt_col3:
        po_date_to   = st.date_input(f"📅 {t('To','إلى')}",  value=default_to,   key="po_date_to_v4")
    with filt_col4:
        sys_display_names_p = [get_system_name(k) for k in SYSTEM_KEYS]
        po_system_sel = st.multiselect(
            f"🏢 {t('System','النظام')}", options=sys_display_names_p, default=sys_display_names_p,
            key="po_system_sel_v4")
        _disp2key_p = {get_system_name(k): k for k in SYSTEM_KEYS}
        selected_system_keys_p = [_disp2key_p.get(d, d) for d in po_system_sel] or SYSTEM_KEYS
    with filt_col5:
        cached_po      = st.session_state.get("po_analytics_df")
        vendor_options = []
        if cached_po is not None and not cached_po.empty and C_VENDOR in cached_po.columns:
            vendor_options = sorted(cached_po[C_VENDOR].dropna().unique().tolist())
        po_vendor_sel = st.multiselect(f"🏭 {t('Vendor','المورد')}", options=vendor_options,
                                       default=[],
                                       placeholder=t("All Vendors (default)","كل الموردين (افتراضي)"),
                                       key="po_vendor_sel_v4")

    if st.button(f"🔍 {t('Fetch Purchase Analytics','جلب تحليلات المشتريات')}",
                 type="primary", key="fetch_po_btn_v4"):
        with st.spinner(t("⚡ Fetching all purchase data from all systems…",
                           "⚡ جلب بيانات المشتريات من كل الأنظمة…")):
            fetched = fetch_all_systems_purchase_history(
                model_code=None,
                date_from=po_date_from.strftime("%Y-%m-%d"),
                date_to=po_date_to.strftime("%Y-%m-%d"),
                system_keys=selected_system_keys_p)
        st.session_state.po_analytics_df = fetched
        st.session_state.page_po = 0
        st.rerun()

    po_full = st.session_state.get("po_analytics_df")
    if po_full is None:
        st.info(t("👆 Set your date range and click **Fetch Purchase Analytics** to load data.",
                  "👆 حدد نطاق التاريخ واضغط **جلب تحليلات المشتريات** لتحميل البيانات.")); return
    if po_full.empty:
        st.info(t("No purchases found for this period.","لا توجد مشتريات لهذه الفترة.")); return

    # Ensure numeric
    for nc in [C_QTY, C_UNIT_PRICE, C_SUBTOTAL]:
        if nc in po_full.columns:
            po_full[nc] = _to_num(po_full[nc])

    # Apply system filter on canonical key
    if C_SYSTEM in po_full.columns:
        pdf_sys = po_full[po_full[C_SYSTEM].isin(selected_system_keys_p)].copy()
    else:
        pdf_sys = po_full.copy()

    # Apply vendor filter
    pdf_vendor = (pdf_sys[pdf_sys[C_VENDOR].isin(po_vendor_sel)].copy()
                  if po_vendor_sel else pdf_sys.copy())

    # ── Single model detail ──
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Single Model Purchase Detail","تفاصيل شراء موديل واحد"), "🔍")

    if not po_model_input:
        st.info(t("💡 Enter a **Model Code** in the filter above to see single-model analytics.",
                  "💡 أدخل **رمز الموديل** في الفلتر أعلاه لعرض تحليلات موديل واحد."))
    else:
        mc_norm  = po_model_input.upper()
        model_df = pdf_sys[pdf_sys[C_MODEL].str.upper() == mc_norm].copy() if C_MODEL in pdf_sys.columns else pd.DataFrame()
        if po_vendor_sel and C_VENDOR in model_df.columns:
            model_df = model_df[model_df[C_VENDOR].isin(po_vendor_sel)]

        if model_df.empty:
            st.info(t(f"No purchase records found for model **{po_model_input}**.",
                      f"لا توجد سجلات شراء للموديل **{po_model_input}**."))
        else:
            pb_qty  = float(_to_num(model_df[C_QTY]).sum())
            pb_amt  = float(_to_num(model_df[C_SUBTOTAL]).sum())
            pb_vend = int(model_df[C_VENDOR].nunique()) if C_VENDOR in model_df.columns else 0
            pb_sys  = int(model_df[C_SYSTEM].nunique()) if C_SYSTEM in model_df.columns else 0
            pb_cards = [
                _premium_kpi_card("📦", f"{pb_qty:,.0f}",  t("Total Qty (this model)","إجمالي الكمية (الموديل)")),
                _premium_kpi_card("💰", f"{pb_amt:,.2f}",  t("Total Amount (SAR)","إجمالي المبلغ")),
                _premium_kpi_card("🏭", str(pb_vend),      t("Vendors","الموردون")),
                _premium_kpi_card("🏢", str(pb_sys),       t("Systems","الأنظمة")),
            ]
            _render_kpi_grid(pb_cards)

            model_vendors = sorted(model_df[C_VENDOR].dropna().unique().tolist()) if C_VENDOR in model_df.columns else []
            pb_vendor_sel = st.multiselect(
                f"🏭 {t('Filter vendors for this model','فلتر الموردين لهذا الموديل')}",
                options=model_vendors, default=[],
                placeholder=t("All vendors for this model","كل موردين هذا الموديل"),
                key="pb_vendor_sel_v4")
            model_vendor_df = (model_df[model_df[C_VENDOR].isin(pb_vendor_sel)].copy()
                               if pb_vendor_sel else model_df.copy())

            if model_vendor_df.empty:
                st.warning(t("No data for selected vendor(s).","لا بيانات للموردين المحددين."))
            else:
                st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
                _chart_card_open(t("Purchase Qty Over Time","كمية الشراء عبر الزمن"), "📈")
                ts_df = (model_vendor_df.groupby(C_DATE, as_index=False)[C_QTY]
                         .sum().sort_values(C_DATE))
                ts_df[C_QTY] = _to_num(ts_df[C_QTY])
                if not ts_df.empty:
                    ts_plot = ts_df.copy()
                    ts_plot[C_DATE] = pd.to_datetime(ts_plot[C_DATE], errors="coerce")
                    ts_plot = ts_plot.dropna(subset=[C_DATE])
                    if not ts_plot.empty:
                        st.altair_chart(_alt_line_chart(ts_plot, C_DATE, C_QTY, color="#e74c3c"),
                                        use_container_width=True)
                _chart_card_close()

                st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
                _chart_card_open(t("Vendor Share for this Model","حصة الموردين لهذا الموديل"), "🏭")
                vshare = (model_vendor_df
                          .assign(**{C_VENDOR: model_vendor_df[C_VENDOR].replace("","(No Vendor)").fillna("(No Vendor)")})
                          .groupby(C_VENDOR, as_index=False)[C_QTY]
                          .sum().sort_values(C_QTY, ascending=False).reset_index(drop=True))
                vshare[C_QTY] = _to_num(vshare[C_QTY])
                vshare["Total Qty Display"] = vshare[C_QTY].map(lambda v: f"{v:,.0f}")
                vs1,vs2 = st.columns([1.6,1])
                with vs1:
                    st.altair_chart(_alt_bar_chart(vshare, x_field=C_VENDOR, y_field=C_QTY,
                                                   tooltip_fmt=",.0f", color="#9b59b6"),
                                    use_container_width=True)
                with vs2:
                    disp_vs = vshare[[C_VENDOR, "Total Qty Display"]].rename(columns={
                        C_VENDOR: col_label(C_VENDOR),
                        "Total Qty Display": t("Total Qty","إجمالي الكمية")})
                    render_premium_table(disp_vs)
                _chart_card_close()

                if not vshare.empty:
                    _vd1,_vd2,_vd3 = st.columns([1,1.2,1])
                    with _vd2:
                        _plotly_donut(vshare[C_VENDOR].tolist(), vshare[C_QTY].tolist(),
                                      title=t("Vendor Share","حصة الموردين"), height=300)

                st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
                st.markdown(f"#### 📋 {t('Model Detail Table','جدول تفاصيل الموديل')} — {po_model_input}")
                _po_full_table(model_vendor_df, page_key="page_po_model")
                st.markdown("<br>", unsafe_allow_html=True)
                _po_download_row(model_vendor_df, tag_suffix=f"_{mc_norm}_v4")

    # ── Overall purchase analytics ──
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Overall Purchase Analytics","تحليلات المشتريات الإجمالية"), "📊")

    if pdf_vendor.empty:
        st.warning(t("No data for the selected filters.","لا توجد بيانات للفلاتر المحددة.")); return

    _po_kpi_row(pdf_vendor)
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Sales by system
    sys_grp_p = (
        pdf_vendor.assign(**{C_SYSTEM: pdf_vendor[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)})
        .groupby(C_SYSTEM, as_index=False)[C_SUBTOTAL].sum()
        .sort_values(C_SUBTOTAL, ascending=False).reset_index(drop=True)
    )
    sys_grp_p[C_SUBTOTAL] = _to_num(sys_grp_p[C_SUBTOTAL])
    _chart_card_open(t("Purchase Amount by System","مبلغ الشراء حسب النظام"), "🏢")
    spg1,spg2 = st.columns([1.6,1])
    with spg1:
        st.altair_chart(_alt_bar_chart(sys_grp_p, x_field=C_SYSTEM, y_field=C_SUBTOTAL,
                                       tooltip_fmt=",.2f", color="#1abc9c"),
                        use_container_width=True)
    with spg2:
        sys_grp_p["Amount Display"] = sys_grp_p[C_SUBTOTAL].map(lambda v: f"{v:,.2f} SAR")
        disp_spg = sys_grp_p[[C_SYSTEM, "Amount Display"]].rename(columns={
            C_SYSTEM: col_label(C_SYSTEM), "Amount Display": t("Total SAR","إجمالي المبلغ")})
        render_premium_table(disp_spg)
    _chart_card_close()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Top 10 vendors by amount
    _chart_card_open(t("Top 10 Vendors by Purchase Amount","أعلى 10 موردين حسب مبلغ الشراء"), "🏭")
    vendor_grp = (
        pdf_vendor.assign(**{C_VENDOR: pdf_vendor[C_VENDOR].replace("","(No Vendor)").fillna("(No Vendor)")})
        .groupby(C_VENDOR, as_index=False)[C_SUBTOTAL]
        .sum().sort_values(C_SUBTOTAL, ascending=False).head(10).reset_index(drop=True))
    vendor_grp[C_SUBTOTAL] = _to_num(vendor_grp[C_SUBTOTAL])
    vendor_grp["Amount Display"] = vendor_grp[C_SUBTOTAL].map(lambda v: f"{v:,.2f}")
    vc1,vc2 = st.columns([1.6,1])
    with vc1:
        st.altair_chart(_alt_bar_chart(vendor_grp, x_field=C_VENDOR, y_field=C_SUBTOTAL,
                                       tooltip_fmt=",.2f", color="#3498db"),
                        use_container_width=True)
    with vc2:
        disp_vg = vendor_grp[[C_VENDOR, "Amount Display"]].rename(columns={
            C_VENDOR: col_label(C_VENDOR),
            "Amount Display": t("Total Amount (SAR)","إجمالي المبلغ")})
        render_premium_table(disp_vg)
    _chart_card_close()

    if not vendor_grp.empty:
        _plotly_donut(vendor_grp[C_VENDOR].tolist(), vendor_grp[C_SUBTOTAL].tolist(),
                      title=t("Vendor Share (Top 10)","حصة الموردين (أعلى 10)"))

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)

    # Top 10 products by qty
    _chart_card_open(t("Top 10 Products by Qty","أعلى 10 منتجات حسب الكمية"), "🏆")
    prod_grp_a = (
        pdf_vendor.assign(**{
            C_MODEL:   pdf_vendor[C_MODEL].replace("","(No Code)").fillna("(No Code)"),
            C_PRODUCT: pdf_vendor[C_PRODUCT].replace("","").fillna("") if C_PRODUCT in pdf_vendor.columns else "",
        })
        .groupby([C_MODEL, C_PRODUCT] if C_PRODUCT in pdf_vendor.columns else [C_MODEL], as_index=False)[C_QTY]
        .sum().sort_values(C_QTY, ascending=False).head(10).reset_index(drop=True))
    prod_grp_a[C_QTY] = _to_num(prod_grp_a[C_QTY])
    prod_grp_a["Total Qty Display"] = prod_grp_a[C_QTY].map(lambda v: f"{v:,.0f}")
    pc1,pc2 = st.columns([1.6,1])
    with pc1:
        st.altair_chart(_alt_bar_chart(prod_grp_a, x_field=C_MODEL, y_field=C_QTY,
                                       tooltip_fmt=",.0f", color="#2ecc71"),
                        use_container_width=True)
    with pc2:
        disp_pg = prod_grp_a[[C_MODEL, "Total Qty Display"]].rename(columns={
            C_MODEL: col_label(C_MODEL),
            "Total Qty Display": t("Total Qty","إجمالي الكمية")})
        render_premium_table(disp_pg)
    _chart_card_close()

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _top10_altair(t("Top 10 Categories by Qty","أعلى 10 فئات حسب الكمية"),
                  C_CATEGORY, C_QTY, pdf_vendor, color="#e74c3c")
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _top10_altair(t("Top 10 Brand Categories by Qty","أعلى 10 فئات علامة تجارية حسب الكمية"),
                  C_BRAND_CAT, C_QTY, pdf_vendor, color="#9b59b6")

    # Share donuts
    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    _section_header(t("Purchase Share Analysis","تحليل حصص المشتريات"), "🥧")
    d1,d2,d3 = st.columns(3)
    with d1:
        cat_share = (pdf_vendor.assign(**{C_CATEGORY: pdf_vendor[C_CATEGORY].replace("","(No Category)").fillna("(No Category)")})
                     .groupby(C_CATEGORY, as_index=False)[C_QTY].sum()
                     .sort_values(C_QTY, ascending=False))
        cat_share[C_QTY] = _to_num(cat_share[C_QTY])
        _plotly_donut(cat_share[C_CATEGORY].tolist(), cat_share[C_QTY].tolist(),
                      title=t("Category Share","حصة الفئة"))
    with d2:
        bc_share_p = (pdf_vendor.assign(**{C_BRAND_CAT: pdf_vendor[C_BRAND_CAT].replace("","(No Brand)").fillna("(No Brand)")})
                      .groupby(C_BRAND_CAT, as_index=False)[C_QTY].sum()
                      .sort_values(C_QTY, ascending=False))
        bc_share_p[C_QTY] = _to_num(bc_share_p[C_QTY])
        _plotly_donut(bc_share_p[C_BRAND_CAT].tolist(), bc_share_p[C_QTY].tolist(),
                      title=t("Brand Category Share","حصة الفئة التجارية"))
    with d3:
        sys_share_p = (
            pdf_vendor.assign(**{C_SYSTEM: pdf_vendor[C_SYSTEM].map(lambda k: get_system_name(k) if k in SYSTEM_KEYS else k)})
            .groupby(C_SYSTEM, as_index=False)[C_SUBTOTAL].sum()
            .sort_values(C_SUBTOTAL, ascending=False))
        sys_share_p[C_SUBTOTAL] = _to_num(sys_share_p[C_SUBTOTAL])
        _plotly_donut(sys_share_p[C_SYSTEM].tolist(), sys_share_p[C_SUBTOTAL].tolist(),
                      title=t("System Share","حصة النظام"))

    st.markdown("<div class='divider'></div>", unsafe_allow_html=True)
    st.markdown(f"#### 📋 {t('Full Purchase Detail','تفاصيل المشتريات الكاملة')}")
    _po_full_table(pdf_vendor, page_key="page_po_full")
    st.markdown("<br>", unsafe_allow_html=True)
    _po_download_row(pdf_vendor, tag_suffix="_overall_v4")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def show_dashboard():
    with st.sidebar:
        st.markdown(f"### ⚙️ {t('Settings','الإعدادات')}")
        lc2 = st.radio(t("🌐 Language","🌐 اللغة"),["EN","AR"],
                       index=0 if get_lang()=="EN" else 1, horizontal=True)
        if lc2!=get_lang(): st.session_state.lang=lc2; st.rerun()
        st.divider()
        st.markdown(f"##### 📊 {t('Analytics','التحليلات')}")
        current_view = st.session_state.get("analytics_view","purchase")
        col_s,col_p = st.columns(2)
        with col_s:
            if st.button(f"💰 {t('Sales','المبيعات')}",
                         type="primary" if current_view=="sales" else "secondary",
                         use_container_width=True, key="nav_sales_btn"):
                st.session_state.analytics_view = "sales"; st.rerun()
        with col_p:
            if st.button(f"🛒 {t('Purchase','المشتريات')}",
                         type="primary" if current_view=="purchase" else "secondary",
                         use_container_width=True, key="nav_purchase_btn"):
                st.session_state.analytics_view = "purchase"; st.rerun()
        if current_view=="sales":
            st.markdown("<div style='background:rgba(46,204,113,0.1);border-left:3px solid #2ecc71;"
                        "border-radius:8px;padding:8px 12px;font-size:0.75rem;color:#2ecc71;margin-top:8px;'>"
                        f"✅ {t('Viewing: All Systems Sales','عرض: مبيعات كل الأنظمة')}</div>",
                        unsafe_allow_html=True)
        else:
            st.markdown("<div style='background:rgba(52,152,219,0.1);border-left:3px solid #3498db;"
                        "border-radius:8px;padding:8px 12px;font-size:0.75rem;color:#3498db;margin-top:8px;'>"
                        f"✅ {t('Viewing: All Systems Purchase','عرض: مشتريات كل الأنظمة')}</div>",
                        unsafe_allow_html=True)
        st.divider()
        st.markdown(f"👤 **{st.session_state.user_email}**")
        if st.button(f"🚪 {t('Logout','تسجيل الخروج')}", use_container_width=True):
            do_logout()
        st.divider()
        st.markdown(f"##### 🔬 {t('Search Mode','وضع البحث')}")
        et = st.toggle(t("Exact match only","تطابق تام فقط"), value=st.session_state.search_exact)
        if et!=st.session_state.search_exact:
            st.session_state.search_exact = et
            st.session_state.total_df = st.session_state.branch_df = st.session_state.transfers_df = None
            st.rerun()
        st.caption(t("🎯 Exact","🎯 تطابق تام") if st.session_state.search_exact
                   else t("🔍 Variant wildcard","🔍 كل المتغيرات"))
        st.divider()
        st.markdown(f"##### 🔴 {t('Low Stock Alert','تنبيه المخزون')}")
        thr = st.number_input(t("Threshold (qty ≤)","الحد (كمية ≤)"),
                              min_value=0, max_value=1000,
                              value=st.session_state.low_stock_thresh, step=1)
        if thr!=st.session_state.low_stock_thresh:
            st.session_state.low_stock_thresh = int(thr)
        st.divider()
        if st.session_state.last_run:
            st.markdown(f"🕒 **{t('Last Run','آخر تشغيل')}**")
            st.caption(st.session_state.last_run.get("time",""))

    current_view = st.session_state.get("analytics_view","purchase")

    if current_view == "sales":
        st.markdown(f"""<div class='dash-header'>
            <div class='dash-title'>💰 {t('Sales Dashboard — All Systems','لوحة المبيعات — كل الأنظمة')}</div>
            <div class='dash-subtitle'>{t('Sales analytics from all configured Odoo systems','تحليلات المبيعات من كل أنظمة أودو المُعدَّة')}</div>
        </div>""", unsafe_allow_html=True)
        st.divider(); show_sales_analytics(); return

    if current_view == "purchase":
        st.markdown(f"""<div class='dash-header'>
            <div class='dash-title'>🛒 {t('Purchase Dashboard — All Systems','لوحة المشتريات — كل الأنظمة')}</div>
            <div class='dash-subtitle'>{t('Purchase analytics from all configured Odoo systems','تحليلات المشتريات من كل أنظمة أودو المُعدَّة')}</div>
        </div>""", unsafe_allow_html=True)
        st.divider(); show_purchase_analytics(); return

    # ── Stock comparison view ────────────────────────────────────────────────
    st.markdown(f"""<div class='dash-header'>
        <div class='dash-title'>📊 {t('SWAG Product Comparison','مقارنة منتجات سواغ')}</div>
        <div class='dash-subtitle'>{t('Real-time stock & price across 4 Odoo systems','المخزون والسعر الآني عبر 4 أنظمة أودو')}</div>
    </div>""", unsafe_allow_html=True)
    st.divider()

    # PDF Upload
    _section_header(t("Upload Invoice PDF","رفع فاتورة PDF"), "📄")
    p1,p2 = st.columns([2.5,1.5])
    with p1:
        updf = st.file_uploader(t("Upload PDF","رفع PDF"), type=["pdf"], label_visibility="collapsed")
    with p2:
        emode = None
        if updf:
            emode = st.radio(t("Extract mode","وضع الاستخراج"),
                             [t("Main models","موديلات رئيسية"), t("With sizes","مع المقاسات")],
                             horizontal=True)
    if updf:
        fbytes = updf.read(); fhash = hashlib.md5(fbytes).hexdigest(); ck = f"pdf_{fhash}"
        if ck not in st.session_state:
            with st.spinner(t("⚡ Parsing PDF...","⚡ جاري قراءة الفاتورة...")):
                st.session_state[ck] = parse_invoice_pdf_cached(fbytes)
        raw = st.session_state[ck]
        if raw:
            is_main = emode is None or "Main" in emode or "رئيسية" in emode
            if is_main:
                unique = get_unique_base_models(raw)
            else:
                seen_ws, unique = set(), []
                for item in raw:
                    if item["code"] not in seen_ws:
                        seen_ws.add(item["code"]); unique.append(item)
            unique_sorted = sorted(unique, key=lambda x: x["sequence"])
            unique_codes  = [item["code"] for item in unique_sorted]
            c1,c2,c3 = st.columns(3)
            c1.metric(t("Raw codes","رموز مستخرجة"), len(raw))
            c2.metric(t("Unique models","موديلات فريدة"), len(unique_codes))
            c3.info(f"📌 {t('Main','رئيسية') if is_main else t('With sizes','مع المقاسات')}")
            with st.expander(t(f"📋 {len(unique_codes)} codes","📋 الرموز"), expanded=False):
                st.code("\n".join(f"{item['sequence']:>3}. {item['code']}" for item in unique_sorted))
            ca,cb = st.columns(2)
            with ca:
                if st.button(f"🚀 {t('Total Stock','مخزون إجمالي')}", type="primary",
                             use_container_width=True, key="pt"):
                    st.session_state.pdf_codes = unique_codes; st.session_state.pdf_mode = "total"; st.rerun()
            with cb:
                if st.button(f"🗺️ {t('Branch-wise','حسب الفرع')}", type="secondary",
                             use_container_width=True, key="pb"):
                    st.session_state.pdf_codes = unique_codes; st.session_state.pdf_mode = "branch"; st.rerun()
        else:
            st.warning(t("No codes found in PDF.","لم يتم العثور على رموز."))

    st.divider()

    # Manual Search
    _section_header(t("Manual Search","بحث يدوي"), "✍️")
    L,R = st.columns([1.5,1])
    with L:
        if not st.session_state.search_exact:
            st.markdown("<div class='info-banner'>🔍 <b>Variant mode</b> — XP6013 → XP6013-S/M/L</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div class='warn-banner'>🎯 <b>Exact match mode</b> — identical codes only.</div>", unsafe_allow_html=True)
        ms   = t("Single Model","موديل واحد"); mm = t("Multiple Models","موديلات متعددة")
        mode = st.radio(t("Mode","الوضع"),[ms,mm], horizontal=True, label_visibility="collapsed")
        if mode==mm:
            rt    = st.text_area(t("Codes","الرموز"), height=130, placeholder="ABC123\nDEF456")
            codes = [c.strip() for c in rt.replace(",","\n").splitlines() if c.strip()]
        else:
            sg    = st.text_input(t("Model Code","رمز الموديل"), placeholder="e.g. XP6013")
            codes = [sg.strip()] if sg.strip() else []
        t1,t2,t3,t4,t5 = st.columns(5)
        sz  = t1.toggle(t("Zero","الصفري"),     value=False)
        sb  = t2.toggle(t("Branch","فروع"),      value=False)
        ss  = t3.toggle(t("Sort","ترتيب"),       value=False)
        st_ = t4.toggle(t("Transfers","نقليات"), value=False)
        sr  = t5.toggle(t("Reorder","طلب"),      value=False)
        if sr:
            with st.expander(f"⚙️ {t('Reorder Settings','إعدادات')}", expanded=True):
                rx,ry = st.columns(2)
                with rx:
                    rm = st.radio(t("Mode","الوضع"),
                                  [t("Days cover","تغطية أيام"),t("Max level","مستوى أقصى")],
                                  horizontal=True,
                                  index=0 if st.session_state.reorder_mode=="days_cover" else 1)
                    st.session_state.reorder_mode = ("days_cover" if "Days" in rm or "تغطية" in rm else "max_level")
                with ry:
                    st.session_state.reorder_point = st.number_input(
                        t("Reorder point","نقطة الطلب"), min_value=0, max_value=9999,
                        value=st.session_state.reorder_point, step=1)
                if st.session_state.reorder_mode=="days_cover":
                    st.session_state.reorder_target_days = st.slider(
                        t("Target days","أيام"), 7, 180, st.session_state.reorder_target_days)
                else:
                    st.session_state.reorder_max_level = st.number_input(
                        t("Max level","الحد"), min_value=1, max_value=99999,
                        value=st.session_state.reorder_max_level, step=1)
        cbtn = st.button(f"🔍 {t('Compare','مقارنة')}", use_container_width=True, type="primary")

    with R:
        st.markdown(f"#### 📋 {t('Last Run','آخر تشغيل')}")
        snap  = st.session_state.last_run; stats = st.session_state.sys_stats
        if not snap:
            st.info(t("Run a comparison first.","قم بتشغيل مقارنة أولاً."))
        else:
            on = sum(1 for v in stats.values() if v=="OK")
            st.markdown(f"<div class='stats-card'>"
                        f"🕒 <b>{t('Time','الوقت')}:</b> {snap.get('time','—')}<br>"
                        f"📦 <b>{t('Models','الموديلات')}:</b> {snap.get('models','—')}<br>"
                        f"🌐 <b>{t('Online','متصل')}:</b> {on}/4<br>"
                        f"📊 <b>{t('Rows','الصفوف')}:</b> {snap.get('rows','—')}"
                        f"</div>", unsafe_allow_html=True)
            st.markdown("")
            for key in SYSTEM_KEYS:
                s  = stats.get(key,"—")
                bc = "badge-ok" if s=="OK" else "badge-off" if s=="NOT_FOUND" else "badge-err"
                bt = "✅ OK" if s=="OK" else "🔴 OFF" if s=="NOT_FOUND" else "⚠️ ERR"
                st.markdown(f"<div class='sys-row'>"
                            f"<span style='font-size:.85rem;color:#e8edf2'><b>{get_system_name(key)}</b></span>"
                            f"<span class='{bc}'>{bt}</span></div>", unsafe_allow_html=True)

    # Run comparison
    run_codes = None; force_branch = False
    if st.session_state.get("pdf_codes"):
        run_codes = st.session_state.pdf_codes
        force_branch = st.session_state.get("pdf_mode","total")=="branch"
        sb = True; st.session_state.pdf_codes = None; st.session_state.pdf_mode = "total"
    elif cbtn:
        run_codes = codes

    if run_codes is not None:
        if not run_codes:
            st.warning(t("Enter at least one model code.","أدخل رمزاً واحداً.")); st.stop()
        run_codes = list(dict.fromkeys([c.strip() for c in run_codes if c.strip()]))
        ct = tuple(run_codes)
        with st.spinner(t("⚡ Fetching from 4 systems…","⚡ جلب البيانات من 4 أنظمة…")):
            data = fetch_all_data(ct, exact=st.session_state.search_exact,
                                  need_branch=sb or force_branch,
                                  need_transfers=st_, need_reorder=sr,
                                  reorder_mode=st.session_state.reorder_mode,
                                  target_days=st.session_state.reorder_target_days,
                                  max_level=st.session_state.reorder_max_level,
                                  reorder_point=st.session_state.reorder_point)
        tdf  = data["total"];  bdf  = data["branch"]
        trdf = data["transfers"]; rdf = data["reorder"]
        ns   = {k:"NOT_FOUND" for k in SYSTEM_KEYS}
        if "_status" in tdf.columns and C_SYSTEM in tdf.columns:
            for key in SYSTEM_KEYS:
                mask = tdf[C_SYSTEM] == key
                if mask.any():
                    sv = tdf.loc[mask,"_status"]
                    if   "OK"    in sv.values: ns[key]="OK"
                    elif "ERROR" in sv.values: ns[key]="ERROR"
        if C_ON_HAND in tdf.columns:
            zero_mask = _to_num(tdf[C_ON_HAND]) == 0
            tdf.loc[zero_mask, "_status"] = "not_available"
        if ss and C_SYSTEM in tdf.columns: tdf = tdf.sort_values(C_SYSTEM).reset_index(drop=True)
        if not bdf.empty and ss and C_SYSTEM in bdf.columns: bdf = bdf.sort_values(C_SYSTEM).reset_index(drop=True)
        if sz and C_ON_HAND in tdf.columns:
            zero_count = int((_to_num(tdf[C_ON_HAND]) == 0).sum())
            if zero_count:
                st.sidebar.info(t(f"ℹ️ {zero_count} rows have zero qty",f"ℹ️ {zero_count} صف بكمية صفر"))
        st.session_state.total_df     = tdf; st.session_state.branch_df    = bdf
        st.session_state.transfers_df = trdf; st.session_state.reorder_df  = rdf
        st.session_state.show_transfers = st_; st.session_state.show_reorder = sr
        st.session_state.sys_stats    = ns
        st.session_state.last_run     = {"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                         "models": len(run_codes), "rows": len(tdf)}
        record_price_snapshot(tdf); st.rerun()

    # Results
    tdf  = st.session_state.total_df;   bdf  = st.session_state.branch_df
    trdf = st.session_state.transfers_df; rdf = st.session_state.reorder_df
    if tdf is None or tdf.empty: return

    st.divider()
    thr   = st.session_state.low_stock_thresh
    stats = st.session_state.sys_stats
    ok    = tdf[tdf["_status"]=="OK"] if "_status" in tdf.columns else tdf
    on    = sum(1 for v in stats.values() if v=="OK")

    if thr>0 and C_ON_HAND in ok.columns:
        num_q = _to_num(ok[C_ON_HAND])
        low   = ok[(num_q>0)&(num_q<=thr)]
        if not low.empty:
            det = ", ".join(f"{r.get(C_MODEL,'?')}@{get_system_name(r.get(C_SYSTEM,'?'))}({r.get(C_ON_HAND,0)})"
                            for _,r in low.head(8).iterrows())
            if len(low)>8: det+=f" +{len(low)-8}"
            st.markdown(f"<div class='alert-banner'>🔴 <b>{t('Low Stock','مخزون منخفض')}:</b> "
                        f"{len(low)} ≤{thr} — {det}</div>", unsafe_allow_html=True)

    m1,m2,m3,m4 = st.columns(4)
    m1.metric(t("Total Rows","إجمالي الصفوف"), len(tdf))
    m2.metric(t("Systems Online","الأنظمة"), f"{on}/4")
    if C_ON_HAND in ok.columns:
        m3.metric(t("Total Qty","إجمالي الكمية"), int(_to_num(ok[C_ON_HAND]).sum()))
    if C_SALE_PRICE in ok.columns:
        vp = _to_num(ok[C_SALE_PRICE]); vp_pos = vp[vp>0]
        m4.metric(t("Avg Price","متوسط السعر"), f"{vp_pos.mean():.2f} SAR" if not vp_pos.empty else "—")

    hb = bdf  is not None and not bdf.empty
    ht = st.session_state.show_transfers and trdf is not None and not trdf.empty
    hr = st.session_state.show_reorder   and rdf  is not None and not rdf.empty

    tlabels = [f"📦 {t('Total Stock','المخزون الإجمالي')}", f"📊 {t('Price History','تاريخ الأسعار')}"]
    if hb: tlabels.append(f"🗺️ {t('Branch Stock','مخزون الفروع')}")
    if ht: tlabels.append(f"🚚 {t('Transfers','النقليات')}")
    if hr: tlabels.append(f"📦 {t('Reorder','إعادة الطلب')}")

    tabs = st.tabs(tlabels); ti = 0

    with tabs[ti]:
        ti+=1
        _section_header(t("Total Stock","المخزون الإجمالي"), "📦")
        display_df(tdf, thr, table_key="total")
        st.markdown("<br>", unsafe_allow_html=True)
        d1,d2,d3,_ = st.columns([1,1,1,1])
        d1.download_button(t("⬇️ CSV","⬇️ CSV"), to_csv(tdf), dl_name("total","csv"),
                           "text/csv", use_container_width=True)
        d2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel(tdf), dl_name("total","xlsx"),
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        d3.download_button(t("📥 All Systems","📥 كل الأنظمة"), to_excel_bulk(tdf),
                           dl_name("bulk","xlsx"),
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    with tabs[ti]:
        ti+=1
        _section_header(t("Price History","تاريخ الأسعار"), "📈")
        hdf = build_price_history_df()
        if hdf.empty:
            st.info(t("Run multiple comparisons to track prices.","قم بتشغيل مقارنات متعددة لتتبع الأسعار."))
        else:
            st.line_chart(hdf, use_container_width=True)
            if st.button(t("🗑️ Clear History","🗑️ مسح السجل")):
                st.session_state.price_history={}; st.rerun()

    if hb:
        with tabs[ti]:
            ti+=1
            _section_header(t("Branch-wise Stock","مخزون حسب الفرع"), "🗺️")
            display_df(bdf, thr, table_key="branch")
            okb = bdf[bdf["_status"]=="OK"] if "_status" in bdf.columns else bdf
            if not okb.empty and C_BRANCH in okb.columns and C_ON_HAND in okb.columns:
                chart = okb.groupby([C_SYSTEM, C_BRANCH])[C_ON_HAND].apply(lambda s: _to_num(s).sum()).reset_index()
                if not chart.empty:
                    st.markdown(f"#### 📊 {t('Qty by Branch','الكميات حسب الفرع')}")
                    st.bar_chart(chart.set_index(C_BRANCH)[C_ON_HAND], use_container_width=True)
            b1,b2,_ = st.columns([1,1,2])
            b1.download_button(t("⬇️ CSV","⬇️ CSV"), to_csv(bdf), dl_name("branch","csv"),
                               "text/csv", use_container_width=True)
            b2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel(bdf), dl_name("branch","xlsx"),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    if ht:
        with tabs[ti]:
            ti+=1
            _section_header(t("Pending Transfers","النقليات المعلقة"), "🚚")
            okt = trdf[trdf["_status"]=="OK"] if "_status" in trdf.columns else trdf
            if not okt.empty:
                k1,k2,k3 = st.columns(3)
                k1.metric(t("Total","إجمالي"), len(okt))
                if C_QTY in okt.columns: k2.metric(t("Total Qty","إجمالي الكمية"), int(_to_num(okt[C_QTY]).sum()))
                if C_SYSTEM in okt.columns: k3.metric(t("Systems","الأنظمة"), okt[C_SYSTEM].nunique())
            display_df(trdf, thresh=0, table_key="transfers")
            x1,x2,_ = st.columns([1,1,2])
            x1.download_button(t("⬇️ CSV","⬇️ CSV"), to_csv(trdf), dl_name("transfers","csv"),
                               "text/csv", use_container_width=True)
            x2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel(trdf), dl_name("transfers","xlsx"),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    if hr:
        with tabs[ti]:
            ti+=1
            _section_header(t("Reorder Suggestions","اقتراحات إعادة الطلب"), "📦")
            okr = rdf[rdf["_status"]=="OK"] if "_status" in rdf.columns else rdf
            if not okr.empty:
                crit = okr[okr[C_PRIORITY].str.startswith("🔴")].shape[0] if C_PRIORITY in okr.columns else 0
                lo   = okr[okr[C_PRIORITY].str.startswith("🟡")].shape[0] if C_PRIORITY in okr.columns else 0
                okn  = okr[okr[C_PRIORITY].str.startswith("🟢")].shape[0] if C_PRIORITY in okr.columns else 0
                sg   = int(_to_num(okr[C_SUGGEST]).sum())                   if C_SUGGEST  in okr.columns else 0
                r1,r2,r3,r4 = st.columns(4)
                r1.metric(t("🔴 Critical","🔴 حرج"), crit)
                r2.metric(t("🟡 Low","🟡 منخفض"), lo)
                r3.metric(t("🟢 OK","🟢 كافٍ"), okn)
                r4.metric(t("To Order","للطلب"), sg)
                if crit+lo>0:
                    st.markdown(f"<div class='alert-banner'>🔴 {crit+lo} "
                                f"{t('products need reordering','منتجات تحتاج إعادة طلب')}</div>",
                                unsafe_allow_html=True)
                sa2 = st.toggle(t("Show all","عرض الكل"), value=False)
                dr = (okr if sa2 else
                      okr[okr[C_PRIORITY].str.startswith(("🔴","🟡"))] if C_PRIORITY in okr.columns else okr)
                display_df(dr.reset_index(drop=True), table_key="reorder")
            else:
                st.info(t("No reorder data.","لا بيانات إعادة طلب."))
            o1,o2,_ = st.columns([1,1,2])
            o1.download_button(t("⬇️ CSV","⬇️ CSV"), to_csv(rdf), dl_name("reorder","csv"),
                               "text/csv", use_container_width=True)
            o2.download_button(t("⬇️ Excel","⬇️ إكسل"), to_excel(rdf), dl_name("reorder","xlsx"),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    st.markdown("<div class='footer'>© 2025 SWAG Fashion · Powered by Odoo · Built with ❤️</div>",
                unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
restore_session()

if not st.session_state.authenticated:
    show_login()
else:
    show_dashboard()
