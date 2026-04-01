"""
SWAG Product Comparison Dashboard
Version 27.0 — Premium BI Design + Plotly Donut Fix

CHANGES FROM v26:
  1. Fixed _plotly_donut() — safe data cleaning, title_text/title_x, crash-proof.
  2. Premium KPI cards via _premium_kpi_card() and _premium_kpi_row() helpers.
  3. Upgraded _sales_kpi_row() and _po_kpi_row() with styled HTML metric cards.
  4. Executive BI dashboard CSS added: .kpi-card, .kpi-value, .kpi-label, etc.
  5. Refined chart color palettes, tooltips, and section headers.
  6. All existing features preserved unchanged.
"""

import io
import re
import hashlib
import time
import xmlrpc.client
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import altair as alt
import pandas as pd
import streamlit as st

try:
    import plotly.express as px
    import plotly.graph_objects as go
    _HAS_PLOTLY = True
except ImportError:
    _HAS_PLOTLY = False

st.set_page_config(
    page_title="SWAG Product Comparison",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS  — base theme + premium BI additions
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
*,html,body,[class*="css"]{font-family:'IBM Plex Sans Arabic',sans-serif;box-sizing:border-box;}
.stApp{background:linear-gradient(135deg,#0f0c29,#302b63,#24243e);min-height:100vh;}
section[data-testid="stSidebar"]{background:linear-gradient(180deg,#1a1a2e 0%,#16213e 100%)!important;border-right:1px solid #ffffff15;}
section[data-testid="stSidebar"] *,section[data-testid="stSidebar"] label,section[data-testid="stSidebar"] span,section[data-testid="stSidebar"] p,section[data-testid="stSidebar"] div{color:#e8e8ff!important;}
section[data-testid="stSidebar"] input{color:#1a1a2e!important;}
@keyframes fadeInUp{from{opacity:0;transform:translateY(40px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeInDown{from{opacity:0;transform:translateY(-30px)}to{opacity:1;transform:translateY(0)}}
@keyframes bounceIn{0%{transform:scale(0.2) rotate(-10deg);opacity:0}60%{transform:scale(1.2) rotate(5deg);opacity:1}80%{transform:scale(0.9)}100%{transform:scale(1);opacity:1}}
@keyframes shimmer{0%{background-position:-400% center}100%{background-position:400% center}}
@keyframes pulse{0%,100%{box-shadow:0 0 0 0 #7c3aed44}50%{box-shadow:0 0 20px 8px #7c3aed22}}
@keyframes glow{0%,100%{text-shadow:0 0 10px #667eea88}50%{text-shadow:0 0 30px #f093fbcc,0 0 60px #667eea88}}
@keyframes slideInLeft{from{opacity:0;transform:translateX(-40px)}to{opacity:1;transform:translateX(0)}}
@keyframes slideInRight{from{opacity:0;transform:translateX(40px)}to{opacity:1;transform:translateX(0)}}
@keyframes float{0%,100%{transform:translateY(0)}50%{transform:translateY(-8px)}}
@keyframes btnShine{0%{background-position:-200% center}100%{background-position:200% center}}
@keyframes borderGlow{0%,100%{border-color:#667eea;box-shadow:0 0 5px #667eea44}50%{border-color:#f093fb;box-shadow:0 0 15px #f093fb66}}
@keyframes countUp{from{opacity:0;transform:scale(0.5)}to{opacity:1;transform:scale(1)}}
@keyframes cardPop{from{opacity:0;transform:translateY(12px) scale(0.97)}to{opacity:1;transform:translateY(0) scale(1)}}
.login-orb{width:120px;height:120px;border-radius:50%;background:linear-gradient(135deg,#667eea,#764ba2,#f093fb);display:flex;align-items:center;justify-content:center;font-size:3rem;margin:0 auto 20px;animation:float 3s ease-in-out infinite,bounceIn 1s ease forwards;box-shadow:0 8px 40px #667eea66,0 0 60px #f093fb33;}
.login-title{font-size:2.4rem;font-weight:700;background:linear-gradient(90deg,#667eea,#f093fb,#667eea);background-size:200% auto;-webkit-background-clip:text;-webkit-text-fill-color:transparent;animation:shimmer 3s linear infinite,fadeInDown 0.8s ease forwards;text-align:center;margin-bottom:6px;}
.login-subtitle{color:#c4b5fd!important;font-size:0.95rem;text-align:center;animation:fadeInUp 1s ease forwards;margin-bottom:28px;}
.login-card{background:linear-gradient(145deg,#1e1e3f,#2d2b55);border:1px solid #ffffff18;border-radius:20px;padding:32px 36px;width:100%;animation:fadeInUp 0.9s ease forwards,pulse 3s infinite;}
.welcome-banner{background:linear-gradient(135deg,#667eea22,#f093fb22);border:1px solid #667eea44;border-radius:12px;padding:14px 20px;text-align:center;margin-bottom:20px;font-size:0.95rem;color:#c4b5fd!important;animation:fadeInDown 0.7s ease forwards,borderGlow 3s infinite;}
.stTextInput input,.stNumberInput input,.stTextArea textarea{background:#1e1e3f!important;border:1px solid #667eea66!important;border-radius:10px!important;color:#e8e8ff!important;caret-color:#c4b5fd!important;transition:all 0.3s ease!important;}
.stTextInput input::placeholder,.stNumberInput input::placeholder,.stTextArea textarea::placeholder{color:#7070aa!important;}
.stTextInput input:focus,.stNumberInput input:focus,.stTextArea textarea:focus{border-color:#667eea!important;box-shadow:0 0 0 3px #667eea33!important;background:#252550!important;}
.stTextInput label,.stNumberInput label,.stTextArea label{color:#c4b5fd!important;font-weight:600!important;}
.stFormSubmitButton button,.stButton button[kind="primary"]{background:linear-gradient(90deg,#667eea,#764ba2,#f093fb,#667eea)!important;background-size:300% auto!important;border:none!important;border-radius:12px!important;color:white!important;font-weight:700!important;font-size:1rem!important;padding:12px!important;animation:btnShine 3s linear infinite!important;transition:transform 0.2s,box-shadow 0.2s!important;box-shadow:0 4px 20px #667eea55!important;}
.stFormSubmitButton button:hover,.stButton button[kind="primary"]:hover{transform:translateY(-2px) scale(1.02)!important;box-shadow:0 8px 30px #764ba299!important;}
.stButton button[kind="secondary"]{background:#1e1e3f!important;border:1px solid #667eea66!important;color:#c4b5fd!important;border-radius:10px!important;}
.stButton button[kind="secondary"]:hover{background:linear-gradient(135deg,#667eea,#764ba2)!important;color:white!important;}
.stButton button{color:#c4b5fd!important;}
.stDownloadButton button{background:linear-gradient(135deg,#1e1e3f,#2d2b55)!important;border:1px solid #667eea66!important;border-radius:10px!important;color:#c4b5fd!important;font-size:0.78rem!important;font-weight:600!important;padding:6px 14px!important;transition:all 0.25s ease!important;box-shadow:0 2px 8px #00000044!important;}
.stDownloadButton button:hover{background:linear-gradient(135deg,#667eea,#764ba2)!important;color:white!important;border-color:transparent!important;transform:translateY(-2px) scale(1.04)!important;box-shadow:0 6px 20px #667eea55!important;}
.dash-header{text-align:center;padding:16px 0 24px;animation:fadeInDown 0.6s ease forwards;}
.dash-title{font-size:2.4rem;font-weight:700;background:linear-gradient(90deg,#667eea,#f093fb,#43e97b,#667eea);background-size:300% auto;-webkit-background-clip:text;-webkit-text-fill-color:transparent;animation:shimmer 4s linear infinite,glow 3s ease-in-out infinite;}
.dash-subtitle{color:#a0aec0;font-size:0.95rem;margin-top:-4px;}
[data-testid="stMetric"]{background:linear-gradient(135deg,#1e1e3f,#2d2b55)!important;border:1px solid #ffffff15!important;border-radius:16px!important;padding:16px 20px!important;animation:countUp 0.6s ease forwards;transition:transform 0.2s,box-shadow 0.2s;}
[data-testid="stMetric"]:hover{transform:translateY(-4px);box-shadow:0 8px 30px #667eea44;}
[data-testid="stMetricLabel"]{color:#a0aec0!important;font-size:0.82rem!important;}
[data-testid="stMetricValue"]{font-size:1.7rem!important;font-weight:700!important;background:linear-gradient(90deg,#667eea,#f093fb);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.stTabs [data-baseweb="tab-list"]{background:linear-gradient(90deg,#1e1e3f,#2d2b55);border-radius:12px;padding:4px;gap:4px;}
.stTabs [data-baseweb="tab"]{color:#a0aec0!important;border-radius:10px!important;font-size:0.83rem!important;font-weight:600!important;padding:8px 16px!important;transition:all 0.2s ease!important;}
.stTabs [aria-selected="true"]{background:linear-gradient(90deg,#667eea,#764ba2)!important;color:white!important;box-shadow:0 4px 12px #667eea55!important;}
.info-banner{background:linear-gradient(135deg,#1e3a5f,#1e3a5f99);border-left:4px solid #3b82f6;border-radius:10px;padding:11px 16px;margin:8px 0 16px;font-size:0.85rem;color:#93c5fd!important;animation:slideInLeft 0.4s ease;}
.warn-banner{background:linear-gradient(135deg,#3b2a0a,#3b2a0a99);border-left:4px solid #f59e0b;border-radius:10px;padding:11px 16px;margin:8px 0 16px;font-size:0.85rem;color:#fcd34d!important;}
.alert-banner{background:linear-gradient(135deg,#3b0a1e,#3b0a1e99);border-left:4px solid #f43f5e;border-radius:10px;padding:11px 16px;margin:8px 0 16px;font-size:0.85rem;color:#fca5a5!important;animation:pulse 2s infinite;}
.ok-banner{background:linear-gradient(135deg,#0a3b1e,#0a3b1e99);border-left:4px solid #22c55e;border-radius:10px;padding:11px 16px;margin:8px 0 16px;font-size:0.85rem;color:#86efac!important;}
.snap-card{background:linear-gradient(145deg,#1e1e3f,#2d2b55);border:1px solid #ffffff18;border-radius:14px;padding:16px 20px;font-size:0.87rem;color:#e8e8ff!important;line-height:2;animation:slideInRight 0.5s ease;box-shadow:0 4px 20px #00000055;}
.snap-card b{color:#c4b5fd!important;}
.sys-row{display:flex;align-items:center;gap:8px;margin-bottom:6px;}
.sys-row span{color:#e8e8ff!important;}
.badge-ok{background:linear-gradient(90deg,#065f46,#047857);color:#d1fae5!important;border-radius:20px;padding:3px 12px;font-size:0.76rem;font-weight:700;}
.badge-off{background:linear-gradient(90deg,#991b1b,#b91c1c);color:#fee2e2!important;border-radius:20px;padding:3px 12px;font-size:0.76rem;font-weight:700;}
.badge-err{background:linear-gradient(90deg,#78350f,#92400e);color:#fef3c7!important;border-radius:20px;padding:3px 12px;font-size:0.76rem;font-weight:700;}
.stRadio label,.stRadio div[role="radiogroup"] label span,[data-testid="stToggle"] label,.stCheckbox label{color:#e8e8ff!important;}
div[data-testid="stRadio"] p{color:#e8e8ff!important;}
h1,h2,h3,h4,h5,h6{color:#e8e8ff!important;}
.stMarkdown p,.stMarkdown li{color:#c4b5fd!important;}
.stCaption,[data-testid="stCaptionContainer"] p{color:#8888bb!important;}
.stAlert p{color:#1a1a2e!important;font-weight:600;}
[data-testid="stExpander"]{background:linear-gradient(135deg,#1e1e3f,#2d2b55)!important;border:1px solid #ffffff18!important;border-radius:12px!important;}
[data-testid="stExpander"] summary,[data-testid="stExpander"] summary p{color:#c4b5fd!important;}
[data-testid="stFileUploader"]{background:linear-gradient(135deg,#1e1e3f,#2d2b55)!important;border:2px dashed #667eea66!important;border-radius:14px!important;}
[data-testid="stFileUploader"] p,[data-testid="stFileUploader"] span{color:#c4b5fd!important;}
hr{border:none!important;height:1px!important;background:linear-gradient(90deg,transparent,#667eea66,transparent)!important;margin:16px 0!important;}
[data-testid="stProgressBar"]>div{background:linear-gradient(90deg,#667eea,#f093fb)!important;border-radius:10px!important;}
::-webkit-scrollbar{width:6px;height:6px;}
::-webkit-scrollbar-track{background:#1a1a2e;}
::-webkit-scrollbar-thumb{background:linear-gradient(#667eea,#764ba2);border-radius:10px;}
::-webkit-scrollbar-thumb:hover{background:#f093fb;}
.stNumberInput button{color:#c4b5fd!important;background:#2d2b55!important;}
.mono{font-family:'IBM Plex Mono',monospace;font-size:0.82rem;color:#c4b5fd;}
footer{visibility:hidden;}
[data-baseweb="tag"]{background:#667eea33!important;color:#c4b5fd!important;}
[data-baseweb="select"] div{background:#1e1e3f!important;color:#e8e8ff!important;border-color:#667eea55!important;}
.panel-header{background:linear-gradient(135deg,#1e1e3f,#2d2b55);border:1px solid #667eea44;border-radius:12px;padding:12px 20px;margin:16px 0 12px;font-size:1.05rem;font-weight:700;color:#c4b5fd!important;}
.nav-btn-active{background:linear-gradient(90deg,#667eea,#764ba2)!important;border:none!important;border-radius:10px!important;color:white!important;font-weight:700!important;width:100%!important;padding:10px!important;margin-bottom:4px!important;box-shadow:0 4px 12px #667eea55!important;}
.nav-btn-inactive{background:linear-gradient(135deg,#1e1e3f,#2d2b55)!important;border:1px solid #667eea44!important;border-radius:10px!important;color:#c4b5fd!important;font-weight:600!important;width:100%!important;padding:10px!important;margin-bottom:4px!important;}

/* ── Premium KPI Cards ───────────────────────────────────────────── */
.kpi-row{display:flex;gap:14px;flex-wrap:wrap;margin:0 0 8px;}
.kpi-card{
  flex:1 1 140px;
  min-width:130px;
  background:linear-gradient(145deg,#1c1c42,#282858);
  border:1px solid #ffffff14;
  border-radius:16px;
  padding:18px 20px 16px;
  position:relative;
  overflow:hidden;
  animation:cardPop 0.45s ease forwards;
  transition:transform 0.2s ease,box-shadow 0.2s ease;
  box-shadow:0 4px 24px #00000055;
}
.kpi-card:hover{transform:translateY(-4px) scale(1.02);box-shadow:0 12px 36px #667eea33;}
.kpi-card::before{
  content:'';
  position:absolute;
  top:0;left:0;right:0;
  height:3px;
  border-radius:16px 16px 0 0;
}
.kpi-card.c-blue::before{background:linear-gradient(90deg,#667eea,#4facfe);}
.kpi-card.c-purple::before{background:linear-gradient(90deg,#764ba2,#f093fb);}
.kpi-card.c-green::before{background:linear-gradient(90deg,#43e97b,#38f9d7);}
.kpi-card.c-orange::before{background:linear-gradient(90deg,#fa8231,#f7b731);}
.kpi-card.c-pink::before{background:linear-gradient(90deg,#f093fb,#c471ed);}
.kpi-card.c-cyan::before{background:linear-gradient(90deg,#00f2fe,#4facfe);}
.kpi-icon{font-size:1.5rem;margin-bottom:8px;display:block;line-height:1;}
.kpi-value{
  font-size:1.65rem;
  font-weight:700;
  line-height:1.1;
  margin-bottom:4px;
  background:linear-gradient(90deg,#e8e8ff,#c4b5fd);
  -webkit-background-clip:text;
  -webkit-text-fill-color:transparent;
}
.kpi-card.c-blue .kpi-value{background:linear-gradient(90deg,#a5b4fc,#60a5fa);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-card.c-purple .kpi-value{background:linear-gradient(90deg,#e879f9,#c084fc);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-card.c-green .kpi-value{background:linear-gradient(90deg,#4ade80,#34d399);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-card.c-orange .kpi-value{background:linear-gradient(90deg,#fb923c,#fbbf24);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-card.c-pink .kpi-value{background:linear-gradient(90deg,#f9a8d4,#f472b6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-card.c-cyan .kpi-value{background:linear-gradient(90deg,#67e8f9,#38bdf8);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.kpi-label{font-size:0.76rem;color:#94a3b8;font-weight:500;text-transform:uppercase;letter-spacing:0.6px;}
.kpi-sub{font-size:0.71rem;color:#667eea;margin-top:3px;font-weight:500;}

/* ── Chart section cards ─────────────────────────────────────────── */
.chart-card{
  background:linear-gradient(145deg,#181830,#222248);
  border:1px solid #ffffff0d;
  border-radius:18px;
  padding:20px 22px 16px;
  margin-bottom:14px;
  box-shadow:0 6px 32px #0000004a;
}
.chart-title{
  font-size:0.92rem;
  font-weight:700;
  color:#c4b5fd;
  margin-bottom:14px;
  display:flex;
  align-items:center;
  gap:8px;
  letter-spacing:0.2px;
}
.chart-title span.ct-accent{
  width:4px;height:18px;
  border-radius:3px;
  background:linear-gradient(180deg,#667eea,#f093fb);
  display:inline-block;
  flex-shrink:0;
}

/* ── Section divider line ────────────────────────────────────────── */
.section-sep{height:1px;background:linear-gradient(90deg,transparent,#667eea44,transparent);margin:22px 0;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_KEYS = ["SWAG", "LAROUCHE", "DIFFC", "FASHION_LIMITS"]

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE
# ─────────────────────────────────────────────────────────────────────────────
def get_lang():
    return st.session_state.get("lang", "EN")

def t(en, ar):
    return ar if get_lang() == "AR" else en

def get_system_name(key):
    cfg = st.secrets.get(key, {})
    return cfg.get("name_ar", cfg.get("name", key)) if get_lang() == "AR" else cfg.get("name", key)

# ─────────────────────────────────────────────────────────────────────────────
# TRANSLATE SYSTEM NAMES
# ─────────────────────────────────────────────────────────────────────────────
def translate_system_names(df):
    if df is None or df.empty:
        return df
    sys_col = t("System", "النظام")
    if sys_col not in df.columns:
        return df
    key_to_name = {k: get_system_name(k) for k in SYSTEM_KEYS}
    out = df.copy()
    out[sys_col] = out[sys_col].map(lambda v: key_to_name.get(v, v))
    return out

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────
_DEF = {
    "authenticated"      : False,
    "user_email"         : "",
    "lang"               : "EN",
    "last_run"           : None,
    "total_df"           : None,
    "branch_df"          : None,
    "transfers_df"       : None,
    "reorder_df"         : None,
    "sys_stats"          : {},
    "search_exact"       : False,
    "low_stock_thresh"   : 5,
    "price_history"      : {},
    "show_transfers"     : False,
    "show_reorder"       : False,
    "reorder_mode"       : "days_cover",
    "reorder_target_days": 30,
    "reorder_max_level"  : 100,
    "reorder_point"      : 10,
    "pdf_codes"          : None,
    "pdf_mode"           : "total",
    "po_analytics_df"    : None,
    "pc_purch_df"        : None,
    "pc_stock_df"        : None,
    "pc_last_code"       : "",
    "salesanalyticsdf"   : None,
    "analytics_view"     : "purchase",
}
for k, v in _DEF.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# SESSION LOGIN RESTORE (query_params)
# ─────────────────────────────────────────────────────────────────────────────
_COOKIE_SECRET = "swag_2025_secure"

def _make_token(email):
    return hashlib.sha256(f"{_COOKIE_SECRET}_{email}".encode()).hexdigest()[:32]

def _verify_token(email, token):
    return bool(email and token and token == _make_token(email))

def restore_session():
    if st.session_state.get("authenticated"):
        return
    try:
        params = st.query_params
        email  = params.get("u", "")
        token  = params.get("t", "")
        if email and token and _verify_token(email, token):
            st.session_state.authenticated = True
            st.session_state.user_email    = email
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# XML-RPC
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def _proxy(url, ep):
    return xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/{ep}", allow_none=True)

@st.cache_data(ttl=28800, show_spinner=False)
def _auth(url, db, user, key):
    try:
        uid = _proxy(url, "common").authenticate(db, user, key, {})
        return uid or None
    except Exception:
        return None

def _x(url, db, uid, key, model, method, domain, kw):
    return _proxy(url, "object").execute_kw(db, uid, key, model, method, domain, kw)

# ─────────────────────────────────────────────────────────────────────────────
# DOMAIN BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def _domain(codes, exact):
    if exact:
        return [["default_code", "in", codes]]
    if len(codes) == 1:
        return [["default_code", "=like", f"{codes[0]}%"]]
    parts = [["default_code", "=like", f"{c}%"] for c in codes]
    return ["|"] * (len(parts) - 1) + parts

# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSING
# ─────────────────────────────────────────────────────────────────────────────
_RE_BRACKET = re.compile(r'\[([A-Za-z0-9\-_()]{3,30})\]')
_RE_SR_LINE = re.compile(
    r'(?:^|\s)([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9()]{1,10})?)\s+.{0,80}?\d+\.?\d*\s+SR',
    re.MULTILINE)
_RE_GENERAL = re.compile(
    r'\b([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9]{1,4})?(?:\([^)]{1,15}\))?)\b')
_EXCLUDE = frozenset([
    'SR','VAT','TAX','PCS','QTY','NO','REF','INV','PO','SO',
    'DO','ID','EN','AR','PDF','AED','SAR','USD','KWD','OMR',
    'BHD','JOD','EGP','TRY'
])

def _valid(code):
    c = code.strip().upper()
    return (bool(re.search(r'[A-Z]', c)) and bool(re.search(r'\d', c))
            and 4 <= len(c) <= 25 and c not in _EXCLUDE)

def extract_base_model(code):
    code = re.sub(r'\([^)]*\)', '', code)
    for s in ['-2XL','-3XL','-4XL','-XXL','-XL','-L','-M','-S','-XS','-2X','-3X']:
        if code.upper().endswith(s.upper()):
            code = code[:-len(s)]; break
    return re.sub(r'-\d{2,3}$', '', code).strip()

def get_unique_base_models(raw):
    seen, out = set(), []
    for item in raw:
        b = extract_base_model(item["code"])
        if b and b not in seen:
            seen.add(b)
            out.append({"sequence": item["sequence"], "code": b})
    return out


@st.cache_data(show_spinner=False)
def parse_invoice_pdf_cached(file_bytes):
    try:
        from pypdf import PdfReader
    except ImportError:
        return []

    text = ""
    for page in PdfReader(io.BytesIO(file_bytes)).pages:
        text += (page.extract_text() or "") + "\n"
    if not text.strip():
        return []

    raw = (_RE_BRACKET.findall(text)
           + [m.group(1) for m in _RE_SR_LINE.finditer(text)]
           + _RE_GENERAL.findall(text))

    seen, out = set(), []
    seq = 1
    for c in raw:
        u = c.strip().upper()
        if _valid(u) and u not in seen:
            seen.add(u)
            out.append({"sequence": seq, "code": u})
            seq += 1

    return out


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _style_worksheet(ws, df_clean, lang="EN"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
    from openpyxl.chart import BarChart, Reference
    if lang == "AR":
        ws.sheet_view.rightToLeft = True
    hdr_fill     = PatternFill("solid", fgColor="4B0082")
    hdr_font     = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_align    = Alignment(horizontal="center", vertical="center")
    thin         = Side(border_style="thin", color="D0D0D0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill("solid", fgColor="F3EFFF")
    zero_fill    = PatternFill("solid", fgColor="FFE0E0")
    zero_font    = Font(color="CC0000", bold=True, name="Calibri")
    normal_font  = Font(name="Calibri", size=10)
    num_align    = Alignment(horizontal="right",  vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    total_fill   = PatternFill("solid", fgColor="2E2E2E")
    total_font   = Font(bold=True, name="Calibri", color="FFFFFF")
    max_row = ws.max_row
    max_col = ws.max_column
    ws.row_dimensions[1].height = 28
    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    col_names = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    on_hand_col    = None
    sale_price_col = None
    loc_col        = None
    branch_col     = None
    model_col      = None
    for i, name in enumerate(col_names, 1):
        if name in ("On Hand", "متوفر"):
            on_hand_col = i
        if name in ("Sale Price", "سعر البيع"):
            sale_price_col = i
        if name in ("Location", "الموقع"):
            loc_col = i
        if name in ("Branch", "الفرع"):
            branch_col = i
        if name in ("Model Code", "رمز الموديل"):
            model_col = i
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        is_zero = False
        if on_hand_col:
            val = ws.cell(row=row[0].row, column=on_hand_col).value
            is_zero = (
                val is None or
                str(val).strip() in ['0', 'Not Available', 'غير متوفر', '—', '-', ''] or
                val == 0
            )
        for cell in row:
            cell.border = border
            cell.font = zero_font if is_zero else normal_font
            if is_zero:
                cell.fill = zero_fill
            elif cell.row % 2 == 0:
                cell.fill = alt_fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = num_align
            else:
                cell.alignment = center_align
        ws.row_dimensions[row[0].row].height = 18
    for col_num in range(1, max_col + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for r in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if on_hand_col and max_row > 1:
        col_letter = get_column_letter(on_hand_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            DataBarRule(start_type="min", end_type="max", color="4472C4"),
        )
    if sale_price_col and max_row > 1:
        col_letter = get_column_letter(sale_price_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            ColorScaleRule(
                start_type="min",        start_color="63BE7B",
                mid_type="percentile",   mid_value=50, mid_color="FFEB84",
                end_type="max",          end_color="F8696B",
            ),
        )
    if on_hand_col and max_row > 1:
        col_letter     = get_column_letter(on_hand_col)
        low_stock_fill = PatternFill("solid", fgColor="FFF2CC")
        low_stock_font = Font(color="7F6000", bold=True, name="Calibri")
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            CellIsRule(operator="lessThanOrEqual", formula=["3"],
                       fill=low_stock_fill, font=low_stock_font),
        )
    total_row = max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font      = total_font
    ws.cell(row=total_row, column=1).fill      = total_fill
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
    if on_hand_col:
        col = get_column_letter(on_hand_col)
        ws.cell(row=total_row, column=on_hand_col,
                value=f"=SUM({col}2:{col}{max_row})")
        ws.cell(row=total_row, column=on_hand_col).font      = total_font
        ws.cell(row=total_row, column=on_hand_col).fill      = total_fill
        ws.cell(row=total_row, column=on_hand_col).alignment = Alignment(horizontal="center")
    ws.row_dimensions[total_row].height = 20
    ws.sheet_properties.tabColor = "667EEA"
    footer_row = total_row + 2
    ws.cell(row=footer_row, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SWAG Dashboard")
    ws.cell(row=footer_row, column=1).font = Font(
        italic=True, color="888888", size=9, name="Calibri")
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.print_title_rows        = "1:1"
    ws.print_area              = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.oddHeader.center.text   = "SWAG Product Report"
    ws.oddHeader.center.font   = "Calibri,Bold"
    ws.oddFooter.center.text   = "Page &P of &N  |  Generated: &D"
    ws.sheet_view.zoomScale = 85
    if loc_col:
        ws.column_dimensions[get_column_letter(loc_col)].width = 35
        for row_num in range(2, max_row + 1):
            ws.cell(row=row_num, column=loc_col).alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="left")
            ws.row_dimensions[row_num].height = 28
    if on_hand_col and model_col and max_row > 2:
        chart = BarChart()
        chart.type              = "bar"
        chart.shape             = 4
        chart.title             = "Stock by Branch"
        chart.style             = 10
        chart.y_axis.title      = "On Hand"
        chart.x_axis.title      = "Branch"
        chart.width             = 20
        chart.height            = 12
        data_ref = Reference(ws, min_col=on_hand_col, min_row=1, max_row=max_row)
        cats_ref = Reference(ws, min_col=model_col,   min_row=2, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{max_row + 5}")

def to_csv(df):
    return df.drop(columns=["_status"], errors="ignore").to_csv(index=False).encode("utf-8-sig")

def to_excel(df):
    lang = st.session_state.get('lang', 'EN')
    buf = io.BytesIO()
    clean = df.drop(columns=['_status'], errors='ignore').copy()
    on_hand_col = 'On Hand' if 'On Hand' in clean.columns else (
        'متوفر' if 'متوفر' in clean.columns else None)
    if on_hand_col:
        na_text = 'غير متوفر' if lang == 'AR' else 'Not Available'
        clean[on_hand_col] = clean[on_hand_col].apply(
            lambda x: na_text if (pd.isna(x) or str(x).strip() in ['0', '']) or x == 0 else x
        )
    desired_order = [
        t("Model Code","رمز الموديل"), t("System","النظام"),
        t("Branch","الفرع"), t("Location","الموقع"),
        t("Sale Price","سعر البيع"), t("On Hand","متوفر"),
    ]
    ordered_cols = [c for c in desired_order if c in clean.columns]
    remaining = [c for c in clean.columns if c not in ordered_cols]
    clean = clean[ordered_cols + remaining]
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        clean.to_excel(w, index=False, sheet_name='Data')
        _style_worksheet(w.sheets['Data'], clean, lang=lang)
    return buf.getvalue()

def to_excel_bulk(df):
    lang    = st.session_state.get("lang", "EN")
    buf     = io.BytesIO()
    sys_col = t("System", "النظام")
    _desired = [
        t("Model Code","رمز الموديل"), t("System","النظام"),
        t("Branch","الفرع"), t("Location","الموقع"),
        t("Sale Price","سعر البيع"), t("On Hand","متوفر"),
    ]
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        def _ws(data, name):
            c = data.drop(columns=["_status"], errors="ignore").copy()
            on_hand_col = t("On Hand", "متوفر")
            if on_hand_col in c.columns:
                na_text = 'غير متوفر' if lang == 'AR' else 'Not Available'
                c[on_hand_col] = c[on_hand_col].apply(
                    lambda x: na_text if (pd.isna(x) or str(x).strip() in ['0', '']) or x == 0 else x
                )
            _ordered   = [col for col in _desired if col in c.columns]
            _remaining = [col for col in c.columns if col not in _ordered]
            c = c[_ordered + _remaining]
            c.to_excel(w, index=False, sheet_name=name[:31])
            _style_worksheet(w.sheets[name[:31]], c, lang=lang)
        _ws(df, t("All Systems", "كل الأنظمة"))
        if sys_col in df.columns:
            for key in SYSTEM_KEYS:
                nm  = get_system_name(key)
                sub = df[df[sys_col] == nm]
                if not sub.empty:
                    _ws(sub, nm)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PURCHASE EXCEL HELPER
# ─────────────────────────────────────────────────────────────────────────────
def to_excel_purchase(df):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    clean = df.copy()

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        clean.to_excel(w, index=False, sheet_name="SWAG Purchase")
        ws = w.sheets["SWAG Purchase"]

        hdr_fill    = PatternFill("solid", fgColor="4B0082")
        hdr_font    = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        hdr_align   = Alignment(horizontal="center", vertical="center")
        thin        = Side(border_style="thin", color="D0D0D0")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill("solid", fgColor="F3EFFF")
        normal_font = Font(name="Calibri", size=10)
        num_align   = Alignment(horizontal="right", vertical="center")
        ctr_align   = Alignment(horizontal="center", vertical="center")
        total_fill  = PatternFill("solid", fgColor="2E2E2E")
        total_font  = Font(bold=True, name="Calibri", color="FFFFFF")

        max_row = ws.max_row
        max_col = ws.max_column

        ws.row_dimensions[1].height = 28
        for col_num in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = border

        for row in ws.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                cell.border = border
                cell.font   = normal_font
                if cell.row % 2 == 0:
                    cell.fill = alt_fill
                if isinstance(cell.value, (int, float)):
                    cell.alignment = num_align
                else:
                    cell.alignment = ctr_align
            ws.row_dimensions[row[0].row].height = 18

        for col_num in range(1, max_col + 1):
            col_letter = get_column_letter(col_num)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_num).value or ""))
                 for r in range(1, max_row + 1)),
                default=8
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        total_row = max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font      = total_font
        ws.cell(row=total_row, column=1).fill      = total_fill
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

        col_names = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        for summary_col_name in ("Qty", "Subtotal", "Qty Purchased", "On Hand"):
            if summary_col_name in col_names:
                ci = col_names.index(summary_col_name) + 1
                cl = get_column_letter(ci)
                ws.cell(row=total_row, column=ci, value=f"=SUM({cl}2:{cl}{max_row})")
                ws.cell(row=total_row, column=ci).font      = total_font
                ws.cell(row=total_row, column=ci).fill      = total_fill
                ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="center")

        ws.row_dimensions[total_row].height = 20
        ws.sheet_properties.tabColor = "667EEA"

        footer_row = total_row + 2
        ws.cell(row=footer_row, column=1,
                value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SWAG Purchase History")
        ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9, name="Calibri")

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.print_title_rows       = "1:1"
        ws.sheet_view.zoomScale   = 85

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# SALES EXCEL HELPER
# ─────────────────────────────────────────────────────────────────────────────
def to_excel_sales(df):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    clean = df.copy()

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        clean.to_excel(w, index=False, sheet_name="SWAG Sales")
        ws = w.sheets["SWAG Sales"]

        hdr_fill    = PatternFill("solid", fgColor="1a5276")
        hdr_font    = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        hdr_align   = Alignment(horizontal="center", vertical="center")
        thin        = Side(border_style="thin", color="D0D0D0")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill    = PatternFill("solid", fgColor="EAF2FF")
        normal_font = Font(name="Calibri", size=10)
        num_align   = Alignment(horizontal="right", vertical="center")
        ctr_align   = Alignment(horizontal="center", vertical="center")
        total_fill  = PatternFill("solid", fgColor="1a3a5c")
        total_font  = Font(bold=True, name="Calibri", color="FFFFFF")

        max_row = ws.max_row
        max_col = ws.max_column

        ws.row_dimensions[1].height = 28
        for col_num in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = border

        for row in ws.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                cell.border = border
                cell.font   = normal_font
                if cell.row % 2 == 0:
                    cell.fill = alt_fill
                if isinstance(cell.value, (int, float)):
                    cell.alignment = num_align
                else:
                    cell.alignment = ctr_align
            ws.row_dimensions[row[0].row].height = 18

        for col_num in range(1, max_col + 1):
            col_letter = get_column_letter(col_num)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_num).value or ""))
                 for r in range(1, max_row + 1)),
                default=8
            )
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        total_row = max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font      = total_font
        ws.cell(row=total_row, column=1).fill      = total_fill
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

        col_names = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        for summary_col_name in ("Qty", "Subtotal"):
            if summary_col_name in col_names:
                ci = col_names.index(summary_col_name) + 1
                cl = get_column_letter(ci)
                ws.cell(row=total_row, column=ci, value=f"=SUM({cl}2:{cl}{max_row})")
                ws.cell(row=total_row, column=ci).font      = total_font
                ws.cell(row=total_row, column=ci).fill      = total_fill
                ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="center")

        ws.row_dimensions[total_row].height = 20
        ws.sheet_properties.tabColor = "4facfe"

        footer_row = total_row + 2
        ws.cell(row=footer_row, column=1,
                value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SWAG Sales History")
        ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9, name="Calibri")

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.print_title_rows       = "1:1"
        ws.sheet_view.zoomScale   = 85

    return buf.getvalue()


def dl_name(tag, ext):
    return f"swag_{tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"

# ─────────────────────────────────────────────────────────────────────────────
# FETCH ALL DATA
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=180, show_spinner=False)
def fetch_all_data(
    codes_tuple, exact=False,
    need_branch=False, need_transfers=False, need_reorder=False,
    reorder_mode="days_cover", target_days=30,
    max_level=100, reorder_point=10,
):
    DAYS  = 30
    dfrom = (datetime.now() - timedelta(days=DAYS)).strftime("%Y-%m-%d 00:00:00")
    codes = list(codes_tuple)
    dom   = _domain(codes, exact)

    CS="System"; CM="Model Code"
    CPR="Product"; CP="Sale Price"
    CQ="On Hand"; CB="Branch"
    CL="Location"; CR="Reference"
    CT="Type"; CST="State"
    CF="From"; CTO="To"; CQT="Qty"
    CD="Scheduled"; CSOLD="Sold(30d)"
    CVEL="Daily Vel"; CDAY="Days Left"
    CSUGG="Suggest"; CPRI="Priority"
    SM={
        "draft"    : "Draft",
        "waiting"  : "Waiting",
        "confirmed": "Confirmed",
        "assigned" : "Ready",
    }

    def _one(key):
        cfg = st.secrets.get(key)
        sn  = key
        R   = {"key":key,"total":[],"branch":[],"transfers":[],"reorder":[]}
        if not cfg:
            R["total"].append({CS:sn,CM:"—",CPR:"No config",CP:0.0,CQ:0,"_status":"ERROR"})
            return R
        uid = _auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
        if not uid:
            R["total"].append({CS:sn,CM:"—",CPR:"⚠️ Auth failed",CP:0.0,CQ:0,"_status":"ERROR"})
            return R
        u=cfg["url"]; db=cfg["db"]; ak=cfg["api_key"]
        try:
            prods = _x(u,db,uid,ak,"product.product","search_read",[dom],
                       {"fields":["id","display_name","default_code","qty_available","list_price"],
                        "limit":2000,"order":"default_code asc"})
            if not prods:
                R["total"].append({CS:sn,CM:"—",CPR:"Not found",
                                   CP:0.0,CQ:0,"_status":"NOT_FOUND"})
                return R
            pids = [p["id"] for p in prods]
            pmap = {p["id"]:p for p in prods}

            for p in prods:
                R["total"].append({
                    CS:sn, CM:p.get("default_code") or "—",
                    CPR:p.get("display_name") or "",
                    CP:float(p.get("list_price") or 0),
                    CQ:int(p.get("qty_available") or 0),
                    "_status":"OK"
                })

            if need_branch:
                internal_locs = _x(u,db,uid,ak,"stock.location","search_read",
                                   [[["usage","=","internal"],["active","=",True]]],
                                   {"fields":["id"],"limit":10000})
                internal_ids  = {l["id"] for l in internal_locs}

                qs = _x(u,db,uid,ak,"stock.quant","search_read",
                        [[["product_id","in",pids],
                          ["location_id","in",list(internal_ids)],
                          ["quantity",">",0]]],
                        {"fields":["product_id","location_id","quantity"],"limit":5000})
                for q in qs:
                    pid = q["product_id"][0] if isinstance(q.get("product_id"),list) else None
                    loc = q.get("location_id") or [None,"—"]
                    ln  = loc[1] if isinstance(loc,list) else str(loc)
                    pm  = pmap.get(pid,{})
                    R["branch"].append({
                        CS:sn, CB:ln.split("/")[0].strip(),
                        CM:pm.get("default_code") or "—", CL:ln,
                        CP:float(pm.get("list_price") or 0),
                        CQ:int(q.get("quantity") or 0), "_status":"OK"
                    })

            if need_transfers:
                mvs = _x(u,db,uid,ak,"stock.move","search_read",
                         [[["product_id","in",pids],
                           ["state","in",["draft","waiting","confirmed","assigned"]]]],
                         {"fields":["picking_id","product_id","product_uom_qty"],"limit":2000})
                if mvs:
                    pkids = list({m["picking_id"][0] for m in mvs
                                  if isinstance(m.get("picking_id"),list)})
                    if pkids:
                        pks   = _x(u,db,uid,ak,"stock.picking","search_read",
                                   [[["id","in",pkids]]],
                                   {"fields":["id","name","picking_type_id","state",
                                              "location_id","location_dest_id","scheduled_date"]})
                        pkmap = {p["id"]:p for p in pks}
                        for mv in mvs:
                            pr = mv.get("picking_id")
                            if not isinstance(pr,list): continue
                            pk = pkmap.get(pr[0],{})
                            def _n(f,_p=pk):
                                v=_p.get(f); return v[1] if isinstance(v,list) else (v or "—")
                            sd = pk.get("scheduled_date") or "—"
                            if sd != "—":
                                try: sd=datetime.strptime(sd,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
                                except: pass
                            pid2 = mv["product_id"][0] if isinstance(mv.get("product_id"),list) else None
                            pm2  = pmap.get(pid2,{})
                            R["transfers"].append({
                                CS:sn, CR:pk.get("name") or "—",
                                CT:_n("picking_type_id"),
                                CST:SM.get(pk.get("state",""),pk.get("state","")),
                                CF:_n("location_id"), CTO:_n("location_dest_id"),
                                CM:pm2.get("default_code") or "—",
                                CQT:int(mv.get("product_uom_qty") or 0),
                                CD:sd, "_status":"OK"
                            })

            if need_reorder:
                sl = _x(u,db,uid,ak,"sale.order.line","search_read",
                        [[["product_id","in",pids],
                          ["order_id.state","in",["sale","done"]],
                          ["order_id.date_order",">=",dfrom]]],
                        {"fields":["product_id","product_uom_qty"],"limit":10000})
                sm2 = {}
                for l in sl:
                    pid = l["product_id"][0] if isinstance(l.get("product_id"),list) else None
                    if pid: sm2[pid] = sm2.get(pid,0)+float(l.get("product_uom_qty") or 0)
                for p in prods:
                    pid  = p["id"]; cq=int(p.get("qty_available") or 0)
                    sold = sm2.get(pid,0); vel=round(sold/DAYS,2)
                    dl   = str(round(cq/vel,1)) if vel>0 else "∞"
                    sg   = max(0,round(target_days*vel-cq)) if reorder_mode=="days_cover" else max(0,max_level-cq)
                    pr2  = ("🔴 Critical" if cq<=0
                            else "🟡 Low" if cq<=reorder_point
                            else "🟢 OK")
                    R["reorder"].append({
                        CS:sn, CM:p.get("default_code") or "—",
                        CPR:p.get("display_name") or "",
                        CQ:cq, CSOLD:int(sold), CVEL:vel,
                        CDAY:dl, CSUGG:sg, CPRI:pr2, "_status":"OK"
                    })
        except Exception as e:
            R["total"].append({CS:sn,CM:"—",CPR:f"❌ {e}",CP:0.0,CQ:0,"_status":"ERROR"})
        return R

    at=[]; ab=[]; atr=[]; ar=[]
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_one,k):k for k in SYSTEM_KEYS}
        for f in as_completed(futs):
            r = f.result()
            at.extend(r["total"]); ab.extend(r["branch"])
            atr.extend(r["transfers"]); ar.extend(r["reorder"])

    def _df(rows,cols):
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)
    return {
        "total"    : _df(at,  ["System","Model Code","Product","Sale Price","On Hand","_status"]),
        "branch"   : _df(ab,  ["System","Branch","Model Code","Location","Sale Price","On Hand","_status"]),
        "transfers": _df(atr, ["System","Reference","Type","State","From","To","Model Code","Qty","Scheduled","_status"]),
        "reorder"  : _df(ar,  ["System","Model Code","Product","On Hand","Sold(30d)","Daily Vel","Days Left","Suggest","Priority","_status"]),
    }


# ─────────────────────────────────────────────────────────────────────────────
# FETCH SWAG PURCHASE HISTORY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_swag_purchase_history(model_code, date_from, date_to):
    empty_cols = ["Date", "PO", "Vendor", "Brand Category", "Category",
                  "Model Code", "Product", "Qty", "Unit Price", "Subtotal"]
    empty_df = pd.DataFrame(columns=empty_cols)

    cfg = st.secrets.get("SWAG")
    if not cfg:
        return empty_df

    uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not uid:
        return empty_df

    u  = cfg["url"]
    db = cfg["db"]
    ak = cfg["api_key"]

    try:
        date_from_dt = f"{date_from} 00:00:00"
        date_to_dt   = f"{date_to} 23:59:59"

        line_domain = [
            ["order_id.state", "in", ["purchase", "done"]],
            ["order_id.date_order", ">=", date_from_dt],
            ["order_id.date_order", "<=", date_to_dt],
        ]
        if model_code and str(model_code).strip():
            line_domain.append(["product_id.default_code", "=", str(model_code).strip()])

        lines = _x(u, db, uid, ak, "purchase.order.line", "search_read",
                   [line_domain],
                   {"fields": ["order_id", "product_id", "product_qty", "price_unit"],
                    "limit": 10000,
                    "order": "order_id desc"})

        if not lines:
            return empty_df

        order_ids   = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"), list)})
        product_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})

        orders = _x(u, db, uid, ak, "purchase.order", "search_read",
                    [[["id", "in", order_ids]]],
                    {"fields": ["id", "name", "partner_id", "date_order"],
                     "limit": len(order_ids) + 10})
        order_map = {o["id"]: o for o in orders}

        products = _x(u, db, uid, ak, "product.product", "search_read",
                      [[["id", "in", product_ids]]],
                      {"fields": ["id", "default_code", "display_name", "categ_id", "product_tmpl_id"],
                       "limit": len(product_ids) + 10})
        prod_map = {p["id"]: p for p in products}

        tmpl_ids = list({p["product_tmpl_id"][0] for p in products
                         if isinstance(p.get("product_tmpl_id"), list)})
        tmpl_map = {}
        if tmpl_ids:
            try:
                tmpls = _x(u, db, uid, ak, "product.template", "search_read",
                           [[["id", "in", tmpl_ids]]],
                           {"fields": ["id", "x_brand_category_id"],
                            "limit": len(tmpl_ids) + 10})
                tmpl_map = {t_["id"]: t_ for t_ in tmpls}
            except Exception:
                tmpl_map = {}

        rows = []
        for line in lines:
            oid = line["order_id"][0] if isinstance(line.get("order_id"), list) else None
            pid = line["product_id"][0] if isinstance(line.get("product_id"), list) else None

            order = order_map.get(oid, {})
            prod  = prod_map.get(pid, {})

            raw_date = order.get("date_order") or ""
            try:
                date_str = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except Exception:
                date_str = raw_date[:10] if raw_date else ""

            partner = order.get("partner_id")
            if isinstance(partner, list):
                vendor = str(partner[1]) if len(partner) > 1 else ""
            else:
                vendor = str(partner) if partner else ""

            categ = prod.get("categ_id")
            if isinstance(categ, list):
                category = str(categ[1]) if len(categ) > 1 else ""
            else:
                category = str(categ) if categ else ""

            brand_category = ""
            tmpl_ref = prod.get("product_tmpl_id")
            if isinstance(tmpl_ref, list) and tmpl_ref:
                tmpl = tmpl_map.get(tmpl_ref[0], {})
                bc   = tmpl.get("x_brand_category_id")
                if isinstance(bc, list):
                    brand_category = str(bc[1]) if len(bc) > 1 else ""
                elif bc:
                    brand_category = str(bc)

            model_code_val = str(prod.get("default_code") or "")
            product_name   = str(prod.get("display_name") or "")

            qty        = float(line.get("product_qty") or 0)
            unit_price = float(line.get("price_unit") or 0)
            subtotal   = round(qty * unit_price, 2)

            rows.append({
                "Date"          : date_str,
                "PO"            : str(order.get("name") or ""),
                "Vendor"        : vendor,
                "Brand Category": brand_category,
                "Category"      : category,
                "Model Code"    : model_code_val,
                "Product"       : product_name,
                "Qty"           : qty,
                "Unit Price"    : unit_price,
                "Subtotal"      : subtotal,
            })

        if not rows:
            return empty_df

        df = pd.DataFrame(rows)
        for col in ["Vendor", "Brand Category", "Category", "Model Code", "Product", "PO", "Date"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str)

        df = df.sort_values(by="Date", ascending=False).reset_index(drop=True)
        return df

    except Exception:
        return empty_df


# ─────────────────────────────────────────────────────────────────────────────
# FETCH SWAG SALES HISTORY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetchswagsaleshistory(modelcode, datefrom, dateto):
    empty_cols = ["Date", "SO", "Customer", "Brand Category", "Category",
                  "Model Code", "Product", "Qty", "Unit Price", "Subtotal"]
    empty_df = pd.DataFrame(columns=empty_cols)

    cfg = st.secrets.get("SWAG")
    if not cfg:
        return empty_df

    uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not uid:
        return empty_df

    u  = cfg["url"]
    db = cfg["db"]
    ak = cfg["api_key"]

    try:
        date_from_dt = f"{datefrom} 00:00:00"
        date_to_dt   = f"{dateto} 23:59:59"

        line_domain = [
            ["order_id.state", "in", ["sale", "done"]],
            ["order_id.date_order", ">=", date_from_dt],
            ["order_id.date_order", "<=", date_to_dt],
        ]
        if modelcode and str(modelcode).strip():
            line_domain.append(["product_id.default_code", "=", str(modelcode).strip()])

        lines = _x(u, db, uid, ak, "sale.order.line", "search_read",
                   [line_domain],
                   {"fields": ["order_id", "product_id", "product_uom_qty", "price_unit"],
                    "limit": 20000,
                    "order": "order_id desc"})

        if not lines:
            return empty_df

        order_ids   = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"), list)})
        product_ids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})

        orders = _x(u, db, uid, ak, "sale.order", "search_read",
                    [[["id", "in", order_ids]]],
                    {"fields": ["id", "name", "partner_id", "date_order"],
                     "limit": len(order_ids) + 10})
        order_map = {o["id"]: o for o in orders}

        products = _x(u, db, uid, ak, "product.product", "search_read",
                      [[["id", "in", product_ids]]],
                      {"fields": ["id", "default_code", "display_name", "categ_id", "product_tmpl_id"],
                       "limit": len(product_ids) + 10})
        prod_map = {p["id"]: p for p in products}

        tmpl_ids = list({p["product_tmpl_id"][0] for p in products
                         if isinstance(p.get("product_tmpl_id"), list)})
        tmpl_map = {}
        if tmpl_ids:
            try:
                tmpls = _x(u, db, uid, ak, "product.template", "search_read",
                           [[["id", "in", tmpl_ids]]],
                           {"fields": ["id", "x_brand_category_id"],
                            "limit": len(tmpl_ids) + 10})
                tmpl_map = {t_["id"]: t_ for t_ in tmpls}
            except Exception:
                tmpl_map = {}

        rows = []
        for line in lines:
            oid = line["order_id"][0] if isinstance(line.get("order_id"), list) else None
            pid = line["product_id"][0] if isinstance(line.get("product_id"), list) else None

            order = order_map.get(oid, {})
            prod  = prod_map.get(pid, {})

            raw_date = order.get("date_order") or ""
            try:
                date_str = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except Exception:
                date_str = raw_date[:10] if raw_date else ""

            partner = order.get("partner_id")
            if isinstance(partner, list):
                customer = str(partner[1]) if len(partner) > 1 else ""
            else:
                customer = str(partner) if partner else ""

            categ = prod.get("categ_id")
            if isinstance(categ, list):
                category = str(categ[1]) if len(categ) > 1 else ""
            else:
                category = str(categ) if categ else ""

            brand_category = ""
            tmpl_ref = prod.get("product_tmpl_id")
            if isinstance(tmpl_ref, list) and tmpl_ref:
                tmpl = tmpl_map.get(tmpl_ref[0], {})
                bc   = tmpl.get("x_brand_category_id")
                if isinstance(bc, list):
                    brand_category = str(bc[1]) if len(bc) > 1 else ""
                elif bc:
                    brand_category = str(bc)

            model_code_val = str(prod.get("default_code") or "")
            product_name   = str(prod.get("display_name") or "")

            qty        = float(line.get("product_uom_qty") or 0)
            unit_price = float(line.get("price_unit") or 0)
            subtotal   = round(qty * unit_price, 2)

            rows.append({
                "Date"          : date_str,
                "SO"            : str(order.get("name") or ""),
                "Customer"      : customer,
                "Brand Category": brand_category,
                "Category"      : category,
                "Model Code"    : model_code_val,
                "Product"       : product_name,
                "Qty"           : qty,
                "Unit Price"    : unit_price,
                "Subtotal"      : subtotal,
            })

        if not rows:
            return empty_df

        df = pd.DataFrame(rows)
        for col in ["Customer", "Brand Category", "Category", "Model Code", "Product", "SO", "Date"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str)

        df = df.sort_values(by="Date", ascending=False).reset_index(drop=True)
        return df

    except Exception:
        return empty_df


# ─────────────────────────────────────────────────────────────────────────────
# FETCH SWAG MODEL PURCHASES AND STOCK
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_swag_model_purchases_and_stock(
    model_code: str,
    date_from: str,
    date_to: str,
) -> tuple:
    purch_empty = pd.DataFrame(columns=["Branch", "Vendor", "Date", "Qty Purchased"])
    stock_empty = pd.DataFrame(columns=["Branch", "On Hand"])

    cfg = st.secrets.get("SWAG")
    if not cfg:
        return purch_empty, stock_empty

    uid = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not uid:
        return purch_empty, stock_empty

    u  = cfg["url"]
    db = cfg["db"]
    ak = cfg["api_key"]

    code = str(model_code).strip().upper()

    def _loc_to_branch(loc_name: str) -> str:
        if not loc_name:
            return "Main"
        return loc_name.split("/")[0].strip() or "Main"

    try:
        date_from_dt = f"{date_from} 00:00:00"
        date_to_dt   = f"{date_to} 23:59:59"

        line_domain = [
            ["order_id.state", "in", ["purchase", "done"]],
            ["order_id.date_order", ">=", date_from_dt],
            ["order_id.date_order", "<=", date_to_dt],
            ["product_id.default_code", "=ilike", code],
        ]

        lines = _x(u, db, uid, ak, "purchase.order.line", "search_read",
                   [line_domain],
                   {"fields": ["order_id", "product_id", "product_qty"],
                    "limit": 20000})

        purch_rows = []

        if lines:
            order_ids = list({l["order_id"][0] for l in lines
                              if isinstance(l.get("order_id"), list)})

            orders = _x(u, db, uid, ak, "purchase.order", "search_read",
                        [[["id", "in", order_ids]]],
                        {"fields": ["id", "name", "partner_id", "date_order",
                                    "picking_type_id"],
                         "limit": len(order_ids) + 10})
            order_map = {o["id"]: o for o in orders}

            pt_ids = list({
                o["picking_type_id"][0]
                for o in orders
                if isinstance(o.get("picking_type_id"), list)
            })
            pt_map = {}
            if pt_ids:
                try:
                    pts = _x(u, db, uid, ak, "stock.picking.type", "search_read",
                             [[["id", "in", pt_ids]]],
                             {"fields": ["id", "name", "warehouse_id"], "limit": len(pt_ids) + 10})
                    wh_ids = list({
                        p["warehouse_id"][0]
                        for p in pts
                        if isinstance(p.get("warehouse_id"), list)
                    })
                    wh_map = {}
                    if wh_ids:
                        whs = _x(u, db, uid, ak, "stock.warehouse", "search_read",
                                 [[["id", "in", wh_ids]]],
                                 {"fields": ["id", "name"], "limit": len(wh_ids) + 10})
                        wh_map = {w["id"]: w["name"] for w in whs}
                    for p in pts:
                        wh_ref = p.get("warehouse_id")
                        if isinstance(wh_ref, list) and wh_ref:
                            pt_map[p["id"]] = wh_map.get(wh_ref[0], str(wh_ref[1]) if len(wh_ref) > 1 else "Main")
                        else:
                            pt_map[p["id"]] = "Main"
                except Exception:
                    pass

            for line in lines:
                oid = line["order_id"][0] if isinstance(line.get("order_id"), list) else None
                order = order_map.get(oid, {})

                raw_date = order.get("date_order") or ""
                try:
                    date_str = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
                except Exception:
                    date_str = raw_date[:10] if raw_date else ""

                partner = order.get("partner_id")
                if isinstance(partner, list):
                    vendor = str(partner[1]) if len(partner) > 1 else "Unknown"
                else:
                    vendor = str(partner) if partner else "Unknown"

                pt_ref = order.get("picking_type_id")
                if isinstance(pt_ref, list) and pt_ref:
                    branch = pt_map.get(pt_ref[0], "Main")
                else:
                    branch = "Main"

                qty = float(line.get("product_qty") or 0)
                if qty <= 0:
                    continue

                purch_rows.append({
                    "Branch"       : branch,
                    "Vendor"       : vendor,
                    "Date"         : date_str,
                    "Qty Purchased": qty,
                })

        prod_ids_res = _x(u, db, uid, ak, "product.product", "search_read",
                          [[["default_code", "=ilike", code]]],
                          {"fields": ["id"], "limit": 500})
        prod_ids = [p["id"] for p in prod_ids_res]

        stock_rows = []

        if prod_ids:
            internal_locs = _x(u, db, uid, ak, "stock.location", "search_read",
                                [[["usage", "=", "internal"], ["active", "=", True]]],
                                {"fields": ["id", "complete_name"], "limit": 10000})
            internal_id_map = {l["id"]: l.get("complete_name", "") for l in internal_locs}
            internal_ids    = list(internal_id_map.keys())

            quants = _x(u, db, uid, ak, "stock.quant", "search_read",
                        [[["product_id", "in", prod_ids],
                          ["location_id", "in", internal_ids]]],
                        {"fields": ["location_id", "quantity"], "limit": 5000})

            branch_qty: dict = {}
            for q in quants:
                qty = float(q.get("quantity") or 0)
                if qty <= 0:
                    continue
                loc_ref = q.get("location_id")
                if isinstance(loc_ref, list):
                    loc_id   = loc_ref[0]
                    loc_name = loc_ref[1] if len(loc_ref) > 1 else internal_id_map.get(loc_id, "")
                else:
                    loc_name = internal_id_map.get(loc_ref, "") if loc_ref else ""
                branch = _loc_to_branch(loc_name)
                branch_qty[branch] = branch_qty.get(branch, 0.0) + qty

            for branch, qty in branch_qty.items():
                stock_rows.append({"Branch": branch, "On Hand": qty})

        purch_df = pd.DataFrame(purch_rows) if purch_rows else purch_empty.copy()
        stock_df = pd.DataFrame(stock_rows) if stock_rows else stock_empty.copy()

        if not purch_df.empty:
            purch_df["Qty Purchased"] = pd.to_numeric(purch_df["Qty Purchased"], errors="coerce").fillna(0)
            for col in ["Branch", "Vendor", "Date"]:
                purch_df[col] = purch_df[col].fillna("").astype(str)
        if not stock_df.empty:
            stock_df["On Hand"] = pd.to_numeric(stock_df["On Hand"], errors="coerce").fillna(0)
            stock_df["Branch"]  = stock_df["Branch"].fillna("").astype(str)

        return purch_df, stock_df

    except Exception:
        return purch_empty, stock_empty


# ─────────────────────────────────────────────────────────────────────────────
# RENAME CACHED COLUMNS TO CURRENT LANGUAGE
# ─────────────────────────────────────────────────────────────────────────────
_COL_MAP_EN = {
    "System":"System","Model Code":"Model Code","Product":"Product",
    "Sale Price":"Sale Price","On Hand":"On Hand","Branch":"Branch",
    "Location":"Location","Reference":"Reference","Type":"Type",
    "State":"State","From":"From","To":"To","Qty":"Qty",
    "Scheduled":"Scheduled","Sold(30d)":"Sold(30d)","Daily Vel":"Daily Vel",
    "Days Left":"Days Left","Suggest":"Suggest","Priority":"Priority",
}
_COL_MAP_AR = {
    "System":"النظام","Model Code":"رمز الموديل","Product":"المنتج",
    "Sale Price":"سعر البيع","On Hand":"متوفر","Branch":"الفرع",
    "Location":"الموقع","Reference":"المرجع","Type":"النوع",
    "State":"الحالة","From":"من","To":"إلى","Qty":"الكمية",
    "Scheduled":"المجدول","Sold(30d)":"مباع(30ي)","Daily Vel":"معدل/يوم",
    "Days Left":"أيام متبقية","Suggest":"المقترح","Priority":"الأولوية",
}

def localize_columns(df):
    if df is None or df.empty:
        return df
    col_map = _COL_MAP_AR if get_lang() == "AR" else _COL_MAP_EN
    return df.rename(columns=col_map)

def prepare_df(df):
    df = localize_columns(df)
    df = translate_system_names(df)
    return df

# ─────────────────────────────────────────────────────────────────────────────
# PRICE HISTORY
# ─────────────────────────────────────────────────────────────────────────────
def record_price_snapshot(df):
    pc=t("Sale Price","سعر البيع"); sc=t("System","النظام"); mc=t("Model Code","رمز الموديل")
    if pc not in df.columns: return
    ok = df[df["_status"]=="OK"] if "_status" in df.columns else df
    if ok.empty: return
    ts = datetime.now().strftime("%H:%M:%S")
    for _, row in ok.iterrows():
        k = f"{row.get(sc,'?')}|{row.get(mc,'?')}"
        st.session_state.price_history.setdefault(k,[]).append(
            {"time":ts,"price":float(row.get(pc,0))})

def build_price_history_df():
    hist = st.session_state.price_history
    if not hist: return pd.DataFrame()
    all_t = sorted({e["time"] for v in hist.values() for e in v})
    recs  = []
    for ts in all_t:
        row = {"time":ts}
        for k, entries in hist.items():
            px = [e["price"] for e in entries if e["time"]==ts]
            row[k] = px[-1] if px else None
        recs.append(row)
    return pd.DataFrame(recs).set_index("time")

# ─────────────────────────────────────────────────────────────────────────────
# QTY DISPLAY HELPER
# ─────────────────────────────────────────────────────────────────────────────
def get_qty_display(qty, lang="EN"):
    try:
        v = float(qty)
        if pd.isna(v) or v == 0:
            return "❌ لا يوجد" if lang == "AR" else "❌ Not Available"
        return int(v)
    except Exception:
        return "❌ لا يوجد" if lang == "AR" else "❌ Not Available"


# ─────────────────────────────────────────────────────────────────────────────
# HTML TABLE
# ─────────────────────────────────────────────────────────────────────────────
_TABLE_CSS = """<style>
.swag-wrap{width:100%;overflow-x:auto;border-radius:16px;box-shadow:0 4px 32px rgba(0,0,0,.5);margin-bottom:4px;}
.swag-tbl{width:100%;border-collapse:collapse;font-family:'IBM Plex Sans Arabic',sans-serif;font-size:.84rem;}
.swag-tbl thead tr{background:linear-gradient(90deg,#667eea,#764ba2,#9b59b6);}
.swag-tbl thead th{color:#fff;font-weight:700;padding:14px 16px;text-align:center;white-space:nowrap;letter-spacing:.4px;border:none;position:sticky;top:0;z-index:2;}
.swag-tbl thead th:first-child{border-radius:16px 0 0 0;}
.swag-tbl thead th:last-child{border-radius:0 16px 0 0;}
.swag-tbl tbody tr:nth-child(odd){background:#1a1a3e;}
.swag-tbl tbody tr:nth-child(odd) td{color:#e8e8ff;}
.swag-tbl tbody tr:nth-child(even){background:#22224a;}
.swag-tbl tbody tr:nth-child(even) td{color:#c4b5fd;}
.swag-tbl tbody td{padding:10px 16px;text-align:center;border-bottom:1px solid #ffffff08;transition:background .15s,color .15s;}
.swag-tbl tbody td.cf{font-weight:700;color:#a78bfa!important;border-right:2px solid #667eea33;}
.swag-tbl tbody tr:hover td{background:#3b2f7a!important;color:#fff!important;}
.swag-tbl tbody tr:hover td.cf{color:#f093fb!important;}
.swag-tbl tbody tr.rl td{background:#3b0a1e!important;color:#fca5a5!important;font-weight:600;}
.swag-tbl tbody tr.rl:hover td{background:#5b1030!important;color:#ffd5d5!important;}
.swag-tbl tbody tr.hi td{background:#1a3b1a!important;color:#86efac!important;font-weight:600;}
.swag-tbl tbody tr.na-row td{background:#2a1a1a!important;opacity:.82;}
.swag-tbl tbody td.na-cell{color:#f97316!important;font-weight:700;letter-spacing:.3px;}
</style>"""

def _render_html_table(df_show, first_col_class="cf"):
    if df_show is None or df_show.empty:
        st.info(t("No data.", "لا بيانات."))
        return
    cols  = df_show.columns.tolist()
    th_   = "".join(f"<th>{c}</th>" for c in cols)
    def _row(idx_row):
        _, row = idx_row
        cells = "".join(
            f'<td class="{first_col_class}">{v}</td>' if ci == 0 else f"<td>{v}</td>"
            for ci, v in enumerate(row)
        )
        return f"<tr>{cells}</tr>"
    tbody = "".join(_row(x) for x in df_show.iterrows())
    st.markdown(
        f'{_TABLE_CSS}<div class="swag-wrap">'
        f'<table class="swag-tbl"><thead><tr>{th_}</tr></thead>'
        f'<tbody>{tbody}</tbody></table></div>',
        unsafe_allow_html=True
    )
    st.caption(f"📊 {len(df_show)} {t('rows', 'صفوف')}")


def display_df(df, thresh=0, table_key="tbl"):
    if df is None or df.empty:
        st.info(t("No data.","لا بيانات."))
        return

    work = df.copy()
    sys_col = t("System","النظام")
    mc_col  = t("Model Code","رمز الموديل")
    pr_col  = t("Product","المنتج")
    br_col  = t("Branch","الفرع")
    loc_col = t("Location","الموقع")
    qc      = t("On Hand","متوفر")
    pc      = t("Sale Price","سعر البيع")

    has_sys = sys_col in work.columns
    has_br  = br_col  in work.columns

    fc = st.columns([2, 2, 2, 1.5])

    if has_sys:
        all_sys = sorted(work[sys_col].dropna().unique().tolist())
        with fc[0]:
            sel_sys = st.multiselect(
                f"🏢 {t('Company','الشركة')}",
                options=all_sys,
                default=all_sys,
                key=f"{table_key}_sys"
            )
        if sel_sys:
            work = work[work[sys_col].isin(sel_sys)]

    if has_br:
        all_br = sorted(work[br_col].dropna().unique().tolist())
        with fc[1]:
            sel_br = st.multiselect(
                f"🏪 {t('Branch','الفرع')}",
                options=all_br,
                default=all_br,
                key=f"{table_key}_br"
            )
        if sel_br:
            work = work[work[br_col].isin(sel_br)]

    with fc[2]:
        q = st.text_input(
            f"🔍 {t('Search model / product','بحث موديل / منتج')}",
            value="",
            placeholder=t("e.g. XP6013 or Shirt","مثال: XP6013"),
            key=f"{table_key}_q"
        ).strip()
    if q:
        ql   = q.lower()
        mask = pd.Series([False] * len(work), index=work.index)
        for col in [mc_col, pr_col, loc_col]:
            if col in work.columns:
                mask = mask | work[col].fillna("").str.lower().str.contains(ql, regex=False)
        work = work[mask]

    with fc[3]:
        sortable = [c for c in work.columns if c != "_status"]
        sort_by  = st.selectbox(
            f"↕️ {t('Sort by','ترتيب')}",
            options=["—"] + sortable,
            index=0,
            key=f"{table_key}_sort"
        )
    if sort_by and sort_by != "—" and sort_by in work.columns:
        try:
            work = work.sort_values(
                by=sort_by,
                key=lambda s: pd.to_numeric(s, errors="coerce").fillna(0)
                              if pd.api.types.is_numeric_dtype(pd.to_numeric(s, errors="coerce"))
                              else s,
                ascending=True
            )
        except Exception:
            work = work.sort_values(by=sort_by)

    if work.empty:
        st.warning(t("⚠️ No rows match your filters.","لا توجد نتائج بعد الفلتر."))
        return

    if qc in work.columns:
        raw_q = pd.to_numeric(work[qc], errors="coerce")
        mn, mx = int(raw_q.min() or 0), int(raw_q.max() or 0)
        if mx > mn:
            qr = st.slider(
                f"📦 {t('Qty range','نطاق الكمية')}",
                min_value=mn, max_value=mx,
                value=(mn, mx),
                key=f"{table_key}_qrange"
            )
            raw_q2 = pd.to_numeric(work[qc], errors="coerce")
            work   = work[(raw_q2 >= qr[0]) & (raw_q2 <= qr[1])]

    ok_work = work[work["_status"]=="OK"] if "_status" in work.columns else work
    sm1, sm2, sm3, sm4 = st.columns(4)
    sm1.metric(t("Rows","الصفوف"), len(work))
    if qc in ok_work.columns:
        sm2.metric(t("Total Qty","إجمالي الكمية"),
                   int(pd.to_numeric(ok_work[qc], errors="coerce").fillna(0).sum()))
    if pc in ok_work.columns:
        vp = pd.to_numeric(ok_work[pc], errors="coerce")
        sm3.metric(t("Avg Price","متوسط السعر"),
                   f"{vp[vp>0].mean():.2f} SAR" if not vp[vp>0].empty else "—")
    if has_sys and sys_col in ok_work.columns:
        sm4.metric(t("Companies","الشركات"), ok_work[sys_col].nunique())

    show = work.drop(columns=["_status"], errors="ignore").copy()

    _raw_qty = (
        pd.to_numeric(work[qc], errors="coerce").fillna(0)
        if qc in work.columns else pd.Series(dtype=float, index=work.index)
    )

    if pc in show.columns:
        show[pc] = pd.to_numeric(show[pc], errors="coerce").map(
            lambda v: f"{v:.2f} SAR" if pd.notna(v) else "—")

    if qc in show.columns:
        _lang = get_lang()
        show[qc] = pd.to_numeric(show[qc], errors="coerce").map(
            lambda v: get_qty_display(v, _lang))

    low_idx = set()
    if thresh > 0 and qc in work.columns:
        raw_q3  = pd.to_numeric(work[qc], errors="coerce")
        low_idx = set(work.index[(raw_q3 > 0) & (raw_q3 <= thresh)])

    _zero_set     = set(_raw_qty.index[_raw_qty == 0]) if not _raw_qty.empty else set()
    _na_label_en  = "❌ Not Available"
    _na_label_ar  = "❌ لا يوجد"

    cols  = show.columns.tolist()
    th_   = "".join(f"<th>{c}</th>" for c in cols)

    def _row(idx_row):
        i, row = idx_row
        is_zero = i in _zero_set
        if is_zero:
            cls = " na-row"
        elif i in low_idx:
            cls = " rl"
        else:
            cls = ""

        cells = "".join(
            f'<td class="cf">{v}</td>'
            if ci == 0
            else (
                f'<td class="na-cell">{v}</td>'
                if is_zero
                   and isinstance(v, str)
                   and v in (_na_label_en, _na_label_ar)
                else f"<td>{v}</td>"
            )
            for ci, v in enumerate(row)
        )
        return f'<tr class="{cls}">{cells}</tr>'

    tbody = "".join(_row(x) for x in show.iterrows())
    st.markdown(
        f'{_TABLE_CSS}<div class="swag-wrap">'
        f'<table class="swag-tbl"><thead><tr>{th_}</tr></thead>'
        f'<tbody>{tbody}</tbody></table></div>',
        unsafe_allow_html=True
    )
    st.caption(f"📊 {len(show)} {t('rows shown','صفوف معروضة')} "
               f"/ {len(df)} {t('total','إجمالي')}")

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    _,_,lc = st.columns([2,1,0.5])
    with lc:
        lg = st.radio("",["EN","AR"],horizontal=True,
                      index=0 if get_lang()=="EN" else 1,
                      label_visibility="collapsed",key="llr")
        if lg!=get_lang(): st.session_state.lang=lg; st.rerun()

    _,col,_ = st.columns([1,1.1,1])
    with col:
        st.markdown("""
        <div style='display:flex;flex-direction:column;align-items:center;padding:20px 0 8px;'>
            <div class='login-orb'>📊</div>
            <div class='login-title'>SWAG Dashboard</div>
            <div class='login-subtitle'>Real-time Stock &amp; Price · 4 Odoo Systems</div>
        </div>""", unsafe_allow_html=True)
        wm = ("🌙 مرحباً بك — سجّل دخولك للمتابعة" if get_lang()=="AR"
              else "👋 Welcome back! Sign in to continue.")
        st.markdown(f"<div class='welcome-banner'>{wm}</div>", unsafe_allow_html=True)
        st.markdown("<div class='login-card'>", unsafe_allow_html=True)

        with st.form("lf", clear_on_submit=False):
            em = st.text_input(
                "📧 Email" if get_lang()=="EN" else "📧 البريد الإلكتروني",
                placeholder="you@swag.com.sa")
            pw = st.text_input(
                "🔑 Password" if get_lang()=="EN" else "🔑 كلمة المرور",
                type="password", placeholder="••••••••")
            st.markdown("<br>", unsafe_allow_html=True)
            sub = st.form_submit_button(
                "🚀 Sign In" if get_lang()=="EN" else "🚀 تسجيل الدخول",
                use_container_width=True, type="primary")
        st.markdown("</div>", unsafe_allow_html=True)

        if sub:
            if not em or not pw:
                st.error(t("Fill in both fields.","يرجى ملء جميع الحقول."))
                return
            if "LOGIN" not in st.secrets:
                st.error("❌ [LOGIN] section missing in secrets.toml")
                return
            cfg = st.secrets["LOGIN"]
            if "url" not in cfg or "db" not in cfg:
                st.error("❌ LOGIN.url or LOGIN.db missing in secrets.toml")
                return
            with st.spinner(t("⚡ Signing in…","⚡ جارٍ تسجيل الدخول…")):
                try:
                    proxy = xmlrpc.client.ServerProxy(
                        f"{cfg['url']}/xmlrpc/2/common", allow_none=True)
                    uid = proxy.authenticate(cfg["db"], em, pw, {})
                    if uid:
                        token = _make_token(em)
                        st.query_params["u"] = em
                        st.query_params["t"] = token
                        st.session_state.authenticated = True
                        st.session_state.user_email    = em
                        time.sleep(0.3)
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(t(
                            "❌ Wrong email or password.",
                            "❌ بريد إلكتروني أو كلمة مرور خاطئة."))
                except Exception as e:
                    st.error(f"❌ Connection error: {e}")

        st.markdown("""<p style='text-align:center;color:#4a4a6a;font-size:.75rem;margin-top:24px;'>
        © 2025 SWAG Fashion · Powered by Odoo · Built with ❤️</p>""",
                    unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────────────────────────────────────────
def do_logout():
    try: st.query_params.clear()
    except Exception: pass
    st.session_state.authenticated = False
    st.session_state.user_email    = ""
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PREMIUM KPI CARD HELPERS  (NEW)
# ─────────────────────────────────────────────────────────────────────────────

def _premium_kpi_card(icon: str, value: str, label: str, color: str = "c-blue", sub: str = "") -> str:
    """Return HTML for a single premium KPI card."""
    sub_html = f"<div class='kpi-sub'>{sub}</div>" if sub else ""
    return (
        f"<div class='kpi-card {color}'>"
        f"<span class='kpi-icon'>{icon}</span>"
        f"<div class='kpi-value'>{value}</div>"
        f"<div class='kpi-label'>{label}</div>"
        f"{sub_html}"
        f"</div>"
    )


def _render_kpi_row(cards: list):
    """Render a flex row of premium KPI cards. Each card is an HTML string."""
    inner = "".join(cards)
    st.markdown(f"<div class='kpi-row'>{inner}</div>", unsafe_allow_html=True)


def _chart_card_open(title: str, icon: str = "📊") -> None:
    """Open a styled chart card section."""
    st.markdown(
        f"<div class='chart-card'>"
        f"<div class='chart-title'>"
        f"<span class='ct-accent'></span>{icon} {title}"
        f"</div>",
        unsafe_allow_html=True
    )


def _chart_card_close() -> None:
    st.markdown("</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# ALTAIR CHART HELPERS  (dark-theme, premium look)
# ─────────────────────────────────────────────────────────────────────────────

_ALT_CONFIG = {
    "background": "transparent",
    "view": {"stroke": "transparent"},
    "axis": {
        "labelColor": "#c4b5fd",
        "titleColor": "#a0aec0",
        "gridColor": "#ffffff12",
        "domainColor": "#667eea33",
        "tickColor": "#667eea33",
        "labelFontSize": 11,
        "titleFontSize": 12,
    },
    "legend": {
        "labelColor": "#c4b5fd",
        "titleColor": "#a0aec0",
    },
    "title": {"color": "#e8e8ff"},
}

# Refined palette — smooth gradients
_PALETTE = [
    "#667eea", "#764ba2", "#f093fb", "#43e97b",
    "#4facfe", "#38f9d7", "#fa8231", "#fd79a8",
    "#a29bfe", "#00cec9", "#fdcb6e", "#e17055",
]


def _alt_bar_chart(
    df: pd.DataFrame,
    x_field: str,
    y_field: str,
    tooltip_fmt: str = ",.0f",
    color: str = "#667eea",
    sort_order: str = "-y",
    label_angle: int = -35,
    height: int = 300,
) -> alt.Chart:
    tooltip_label = f"{y_field} formatted"
    plot_df = df.copy()
    plot_df[tooltip_label] = plot_df[y_field].map(lambda v: f"{v:{tooltip_fmt}}")

    chart = (
        alt.Chart(plot_df)
        .mark_bar(
            cornerRadiusTopLeft=6,
            cornerRadiusTopRight=6,
            opacity=0.9,
        )
        .encode(
            x=alt.X(
                f"{x_field}:N",
                sort=sort_order,
                axis=alt.Axis(labelAngle=label_angle, labelLimit=130),
                title=None,
            ),
            y=alt.Y(f"{y_field}:Q", title=y_field, axis=alt.Axis(grid=True)),
            tooltip=[
                alt.Tooltip(f"{x_field}:N", title=x_field),
                alt.Tooltip(f"{tooltip_label}:N", title=y_field),
            ],
            color=alt.condition(
                alt.datum[y_field] == plot_df[y_field].max(),
                alt.value("#f093fb"),
                alt.value(color),
            ),
        )
        .properties(height=height)
        .configure(**_ALT_CONFIG)
        .interactive()
    )
    return chart


def _alt_line_chart(
    df: pd.DataFrame,
    x_field: str,
    y_field: str,
    height: int = 260,
    color: str = "#667eea",
) -> alt.Chart:
    line = (
        alt.Chart(df)
        .mark_line(
            color=color,
            strokeWidth=2.5,
            interpolate="monotone",
        )
        .encode(
            x=alt.X(f"{x_field}:T", title=None, axis=alt.Axis(format="%b %d", labelAngle=-30)),
            y=alt.Y(f"{y_field}:Q", title=y_field),
            tooltip=[
                alt.Tooltip(f"{x_field}:T", title="Date", format="%Y-%m-%d"),
                alt.Tooltip(f"{y_field}:Q", title=y_field, format=",.0f"),
            ],
        )
    )
    area = (
        alt.Chart(df)
        .mark_area(
            color=color,
            opacity=0.12,
            interpolate="monotone",
        )
        .encode(
            x=alt.X(f"{x_field}:T"),
            y=alt.Y(f"{y_field}:Q"),
        )
    )
    points = (
        alt.Chart(df)
        .mark_circle(color="#f093fb", size=55, opacity=0.9)
        .encode(
            x=alt.X(f"{x_field}:T"),
            y=alt.Y(f"{y_field}:Q"),
            tooltip=[
                alt.Tooltip(f"{x_field}:T", title="Date", format="%Y-%m-%d"),
                alt.Tooltip(f"{y_field}:Q", title=y_field, format=",.0f"),
            ],
        )
    )
    chart = (
        (area + line + points)
        .properties(height=height)
        .configure(**_ALT_CONFIG)
        .interactive()
    )
    return chart


def _po_top10_altair(
    title: str,
    group_col: str,
    value_col: str,
    df: pd.DataFrame,
    color: str = "#764ba2",
    tooltip_fmt: str = ",.0f",
):
    if df is None or df.empty:
        st.info(t("No data available.", "لا توجد بيانات."))
        return

    grp = (
        df.copy()
        .assign(**{group_col: df[group_col].replace("", f"({group_col} N/A)").fillna(f"({group_col} N/A)")})
        .groupby(group_col, as_index=False)[value_col]
        .sum()
        .sort_values(value_col, ascending=False)
        .head(10)
        .reset_index(drop=True)
    )

    if grp.empty:
        st.info(t("No data.", "لا توجد بيانات."))
        return

    display_label = "Total Qty" if value_col == "Qty" else "Total Amount (SAR)"
    grp[display_label] = grp[value_col].map(lambda v: f"{v:{tooltip_fmt}}")

    _chart_card_open(title, "")
    ch_col, tbl_col = st.columns([1.6, 1])
    with ch_col:
        chart = _alt_bar_chart(
            grp, x_field=group_col, y_field=value_col,
            tooltip_fmt=tooltip_fmt, color=color,
        )
        st.altair_chart(chart, use_container_width=True)
    with tbl_col:
        _render_html_table(grp[[group_col, display_label]])
    _chart_card_close()


def _po_kpi_row(df, prefix=""):
    """Premium KPI row for purchase analytics."""
    total_qty    = float(df["Qty"].sum())
    total_amt    = float(df["Subtotal"].sum())
    n_vendors    = int(df["Vendor"].nunique())
    n_products   = int(df["Model Code"].nunique())

    cards = [
        _premium_kpi_card("📦", f"{total_qty:,.0f}",    t("Total Qty Purchased","إجمالي الكمية المشتراة"), "c-blue"),
        _premium_kpi_card("💵", f"{total_amt:,.2f}",    t("Total Amount (SAR)","إجمالي المبلغ"),         "c-purple"),
        _premium_kpi_card("🏭", str(n_vendors),         t("Vendors","الموردون"),                         "c-green"),
        _premium_kpi_card("🏷️", str(n_products),        t("Products","المنتجات"),                        "c-orange"),
    ]
    _render_kpi_row(cards)


def _po_full_table(df):
    show = df.copy()
    show["Unit Price"] = show["Unit Price"].map(lambda v: f"{v:.2f} SAR")
    show["Subtotal"]   = show["Subtotal"].map(lambda v: f"{v:,.2f} SAR")
    show["Qty"]        = show["Qty"].map(lambda v: f"{v:,.0f}")
    _render_html_table(show)


def _po_download_row(df, tag_suffix=""):
    dl1, dl2, _ = st.columns([1, 1, 2])
    dl1.download_button(
        "⬇️ CSV",
        df.to_csv(index=False).encode("utf-8-sig"),
        dl_name(f"purchase{tag_suffix}", "csv"),
        "text/csv",
        use_container_width=True,
        key=f"dl_csv_{tag_suffix}_{id(df)}"
    )
    dl2.download_button(
        "⬇️ Excel",
        to_excel_purchase(df),
        dl_name(f"purchase{tag_suffix}", "xlsx"),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"dl_xlsx_{tag_suffix}_{id(df)}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# PLOTLY DONUT CHART HELPER  — FIXED & CRASH-PROOF  (v27)
# ─────────────────────────────────────────────────────────────────────────────
def _plotly_donut(labels, values, title="", height=360):
    """
    Render a dark-theme Plotly donut chart.
    Fully crash-proof: cleans labels/values, handles empty/null/zero-sum data,
    uses only flat title_text / title_x kwargs (no nested dict).
    """
    if not _HAS_PLOTLY:
        st.info("Install plotly for donut charts: pip install plotly")
        return

    # ── Safe data cleaning ────────────────────────────────────────────────────
    clean_pairs = []
    for lbl, val in zip(labels, values):
        try:
            lbl_s = str(lbl) if lbl is not None else "(N/A)"
            val_f = float(val) if val is not None else 0.0
            if val_f > 0:
                clean_pairs.append((lbl_s, val_f))
        except (TypeError, ValueError):
            continue

    if not clean_pairs:
        st.info(t("No data for chart.", "لا بيانات للرسم."))
        return

    clean_labels = [p[0] for p in clean_pairs]
    clean_values = [p[1] for p in clean_pairs]

    _colors = [
        "#667eea", "#764ba2", "#f093fb", "#43e97b",
        "#4facfe", "#38f9d7", "#fa8231", "#fd79a8",
        "#a29bfe", "#00cec9", "#fdcb6e", "#e17055",
    ]
    used_colors = (_colors * ((len(clean_labels) // len(_colors)) + 1))[:len(clean_labels)]

    try:
        fig = go.Figure(data=[go.Pie(
            labels=clean_labels,
            values=clean_values,
            hole=0.54,
            marker=dict(
                colors=used_colors,
                line=dict(color="#1a1a2e", width=2),
            ),
            textinfo="percent+label",
            textfont=dict(color="#e8e8ff", size=11),
            hovertemplate="<b>%{label}</b><br>Value: %{value:,.0f}<br>Share: %{percent}<extra></extra>",
            sort=True,
            direction="clockwise",
        )])

        fig.update_layout(
            title_text=title,
            title_x=0.5,
            title_font_color="#c4b5fd",
            title_font_size=13,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            height=height,
            margin=dict(t=52, b=20, l=10, r=10),
            legend=dict(
                font=dict(color="#c4b5fd", size=10),
                bgcolor="rgba(24,24,48,0.75)",
                bordercolor="#667eea33",
                borderwidth=1,
                orientation="v",
            ),
            showlegend=True,
        )

        st.plotly_chart(fig, use_container_width=True)

    except Exception as exc:
        st.warning(f"Chart render error: {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# SALES KPI ROW  — Premium version
# ─────────────────────────────────────────────────────────────────────────────
def _sales_kpi_row(df):
    total_qty   = float(df["Qty"].sum())
    total_amt   = float(df["Subtotal"].sum())
    n_customers = int(df["Customer"].nunique())
    n_products  = int(df["Model Code"].nunique())
    n_orders    = int(df["SO"].nunique())
    pos_prices  = df["Unit Price"][df["Unit Price"] > 0]
    avg_price   = float(pos_prices.mean()) if not pos_prices.empty else 0.0

    cards = [
        _premium_kpi_card("📦", f"{total_qty:,.0f}",   t("Total Qty Sold","إجمالي الكمية المباعة"), "c-green"),
        _premium_kpi_card("💰", f"{total_amt:,.2f}",   t("Total Sales (SAR)","إجمالي المبيعات"),    "c-blue"),
        _premium_kpi_card("👤", str(n_customers),       t("Customers","العملاء"),                    "c-purple"),
        _premium_kpi_card("🏷️", str(n_products),        t("Products","المنتجات"),                    "c-orange"),
        _premium_kpi_card("🧾", str(n_orders),          t("Orders","الطلبات"),                       "c-pink"),
        _premium_kpi_card("💲", f"{avg_price:,.2f}",   t("Avg Unit Price","متوسط سعر الوحدة"),      "c-cyan"),
    ]
    _render_kpi_row(cards)


# ─────────────────────────────────────────────────────────────────────────────
# SWAG SALES ANALYTICS VIEW
# ─────────────────────────────────────────────────────────────────────────────
def show_sales_analytics():
    st.markdown(f"### 💰 {t('SWAG Sales Analytics','تحليلات مبيعات سواغ')}")
    st.markdown(
        "<div class='info-banner'>📌 "
        + t("Sales orders from the <b>SWAG</b> system only (state: sale / done).",
            "أوامر البيع من نظام <b>سواغ</b> فقط (الحالة: مبيع / منجز).")
        + "</div>",
        unsafe_allow_html=True
    )

    # ── Filters row ───────────────────────────────────────────────────────────
    default_from = datetime.now().date() - timedelta(days=365)
    default_to   = datetime.now().date()

    sf1, sf2, sf3, sf4 = st.columns([1.2, 1, 1, 1.4])

    with sf1:
        sa_model_input = st.text_input(
            f"🔖 {t('Model Code (optional)','رمز الموديل (اختياري)')}",
            placeholder=t("e.g. RVT196 — blank = all", "مثال: RVT196 — فارغ = الكل"),
            key="sa_model_input"
        ).strip()

    with sf2:
        sa_date_from = st.date_input(
            f"📅 {t('From','من')}",
            value=default_from,
            key="sa_date_from"
        )

    with sf3:
        sa_date_to = st.date_input(
            f"📅 {t('To','إلى')}",
            value=default_to,
            key="sa_date_to"
        )

    with sf4:
        cached_sa      = st.session_state.get("salesanalyticsdf")
        customer_opts  = []
        if cached_sa is not None and not cached_sa.empty and "Customer" in cached_sa.columns:
            customer_opts = sorted(cached_sa["Customer"].dropna().unique().tolist())

        sa_customer_sel = st.multiselect(
            f"👤 {t('Customer','العميل')}",
            options=customer_opts,
            default=[],
            placeholder=t("All Customers (default)", "كل العملاء (افتراضي)"),
            key="sa_customer_sel"
        )

    fetch_sa_btn = st.button(
        f"🔍 {t('Fetch Sales Analytics','جلب تحليلات المبيعات')}",
        type="primary",
        use_container_width=False,
        key="fetch_sa_btn"
    )

    if fetch_sa_btn:
        with st.spinner(t("⚡ Fetching sales data from SWAG…","⚡ جلب بيانات المبيعات من نظام سواغ…")):
            fetched = fetchswagsaleshistory(
                modelcode=None,
                datefrom=sa_date_from.strftime("%Y-%m-%d"),
                dateto=sa_date_to.strftime("%Y-%m-%d"),
            )
        st.session_state.salesanalyticsdf = fetched
        st.rerun()

    # ── Work with cached sales data ───────────────────────────────────────────
    sa_full = st.session_state.get("salesanalyticsdf")

    if sa_full is None:
        st.info(t(
            "👆 Set your date range and click **Fetch Sales Analytics** to load data.",
            "👆 حدد نطاق التاريخ واضغط **جلب تحليلات المبيعات** لتحميل البيانات."
        ))
        return

    if sa_full.empty:
        st.info(t("No sales found for this period.", "لا توجد مبيعات لهذه الفترة."))
        return

    # Apply customer filter
    if sa_customer_sel:
        sa_df = sa_full[sa_full["Customer"].isin(sa_customer_sel)].copy()
    else:
        sa_df = sa_full.copy()

    # Apply model filter (local)
    if sa_model_input:
        mc_norm = sa_model_input.upper()
        sa_df_model = sa_df[sa_df["Model Code"].str.upper() == mc_norm].copy()
    else:
        sa_df_model = None

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>📊 {t('Sales KPIs','مؤشرات المبيعات')}</div>",
        unsafe_allow_html=True
    )
    _sales_kpi_row(sa_df)

    # ── Top 10 Analytics ──────────────────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>🏆 {t('Top 10 Analytics','أفضل 10 تحليلات')}</div>",
        unsafe_allow_html=True
    )

    # Top 10 Products by Qty
    prod_qty_grp = (
        sa_df.copy()
        .assign(**{"Model Code": sa_df["Model Code"].replace("", "(No Code)").fillna("(No Code)")})
        .groupby(["Model Code", "Product"], as_index=False)["Qty"]
        .sum()
        .sort_values("Qty", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    prod_qty_grp["Total Qty"] = prod_qty_grp["Qty"].map(lambda v: f"{v:,.0f}")

    _chart_card_open(t("Top 10 Products by Qty Sold","أعلى 10 منتجات حسب الكمية المباعة"), "🏆")
    pc1, pc2 = st.columns([1.6, 1])
    with pc1:
        st.altair_chart(
            _alt_bar_chart(prod_qty_grp, x_field="Model Code", y_field="Qty",
                           tooltip_fmt=",.0f", color="#43e97b"),
            use_container_width=True
        )
    with pc2:
        _render_html_table(prod_qty_grp[["Model Code", "Product", "Total Qty"]])
    _chart_card_close()

    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Products by Sales Amount
    prod_amt_grp = (
        sa_df.copy()
        .assign(**{"Model Code": sa_df["Model Code"].replace("", "(No Code)").fillna("(No Code)")})
        .groupby(["Model Code", "Product"], as_index=False)["Subtotal"]
        .sum()
        .sort_values("Subtotal", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    prod_amt_grp["Total SAR"] = prod_amt_grp["Subtotal"].map(lambda v: f"{v:,.2f}")

    _chart_card_open(t("Top 10 Products by Sales Amount","أعلى 10 منتجات حسب المبلغ"), "💰")
    pa1, pa2 = st.columns([1.6, 1])
    with pa1:
        st.altair_chart(
            _alt_bar_chart(prod_amt_grp, x_field="Model Code", y_field="Subtotal",
                           tooltip_fmt=",.2f", color="#4facfe"),
            use_container_width=True
        )
    with pa2:
        _render_html_table(prod_amt_grp[["Model Code", "Product", "Total SAR"]])
    _chart_card_close()

    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Brand Categories by Qty
    _po_top10_altair(
        t("Top 10 Brand Categories by Qty","أعلى 10 فئات علامة تجارية حسب الكمية"),
        "Brand Category", "Qty", sa_df,
        color="#f093fb", tooltip_fmt=",.0f"
    )
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Categories by Qty
    _po_top10_altair(
        t("Top 10 Categories by Qty","أعلى 10 فئات حسب الكمية"),
        "Category", "Qty", sa_df,
        color="#764ba2", tooltip_fmt=",.0f"
    )
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Customers by Sales Amount
    cust_grp = (
        sa_df.copy()
        .assign(Customer=sa_df["Customer"].replace("", "(No Customer)").fillna("(No Customer)"))
        .groupby("Customer", as_index=False)["Subtotal"]
        .sum()
        .sort_values("Subtotal", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    cust_grp["Total SAR"] = cust_grp["Subtotal"].map(lambda v: f"{v:,.2f}")

    _chart_card_open(t("Top 10 Customers by Sales Amount","أعلى 10 عملاء حسب المبلغ"), "👤")
    cc1, cc2 = st.columns([1.6, 1])
    with cc1:
        st.altair_chart(
            _alt_bar_chart(cust_grp, x_field="Customer", y_field="Subtotal",
                           tooltip_fmt=",.2f", color="#667eea"),
            use_container_width=True
        )
    with cc2:
        _render_html_table(cust_grp[["Customer", "Total SAR"]])
    _chart_card_close()

    # ── Share Analysis — Donut Charts ─────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>🥧 {t('Sales Share Analysis','تحليل حصص المبيعات')}</div>",
        unsafe_allow_html=True
    )

    pie1, pie2, pie3 = st.columns(3)

    with pie1:
        bc_share = (
            sa_df.copy()
            .assign(**{"Brand Category": sa_df["Brand Category"].replace("", "(No Brand)").fillna("(No Brand)")})
            .groupby("Brand Category", as_index=False)["Subtotal"]
            .sum()
            .sort_values("Subtotal", ascending=False)
        )
        _plotly_donut(
            bc_share["Brand Category"].tolist(),
            bc_share["Subtotal"].tolist(),
            title=t("Brand Category Share", "حصة الفئة التجارية")
        )

    with pie2:
        cat_share = (
            sa_df.copy()
            .assign(Category=sa_df["Category"].replace("", "(No Category)").fillna("(No Category)"))
            .groupby("Category", as_index=False)["Subtotal"]
            .sum()
            .sort_values("Subtotal", ascending=False)
        )
        _plotly_donut(
            cat_share["Category"].tolist(),
            cat_share["Subtotal"].tolist(),
            title=t("Category Share", "حصة الفئة")
        )

    with pie3:
        cust_all = (
            sa_df.copy()
            .assign(Customer=sa_df["Customer"].replace("", "(No Customer)").fillna("(No Customer)"))
            .groupby("Customer", as_index=False)["Subtotal"]
            .sum()
            .sort_values("Subtotal", ascending=False)
        )
        if not cust_all.empty:
            top10c   = cust_all.head(10)
            others_v = float(cust_all.iloc[10:]["Subtotal"].sum()) if len(cust_all) > 10 else 0
            pie_labels = top10c["Customer"].tolist()
            pie_vals   = top10c["Subtotal"].tolist()
            if others_v > 0:
                pie_labels.append("Others")
                pie_vals.append(others_v)
            _plotly_donut(
                pie_labels, pie_vals,
                title=t("Customer Share (Top 10)", "حصة العملاء (أعلى 10)")
            )

    # ── Time-series Sales Trend ───────────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>📈 {t('Sales Trend Over Time','اتجاه المبيعات عبر الزمن')}</div>",
        unsafe_allow_html=True
    )

    ts_col1, ts_col2 = st.columns(2)

    with ts_col1:
        _chart_card_open(t("Qty Sold Over Time","الكمية المباعة عبر الزمن"), "📦")
        ts_qty = (
            sa_df.copy()
            .assign(Date=pd.to_datetime(sa_df["Date"], errors="coerce"))
            .dropna(subset=["Date"])
            .groupby("Date", as_index=False)["Qty"]
            .sum()
            .sort_values("Date")
        )
        if not ts_qty.empty:
            st.altair_chart(
                _alt_line_chart(ts_qty, "Date", "Qty", height=240, color="#43e97b"),
                use_container_width=True
            )
        _chart_card_close()

    with ts_col2:
        _chart_card_open(t("Sales Amount Over Time","مبلغ المبيعات عبر الزمن"), "💰")
        ts_amt = (
            sa_df.copy()
            .assign(Date=pd.to_datetime(sa_df["Date"], errors="coerce"))
            .dropna(subset=["Date"])
            .groupby("Date", as_index=False)["Subtotal"]
            .sum()
            .sort_values("Date")
        )
        if not ts_amt.empty:
            st.altair_chart(
                _alt_line_chart(ts_amt, "Date", "Subtotal", height=240, color="#4facfe"),
                use_container_width=True
            )
        _chart_card_close()

    # ── Single Model Sales Detail ──────────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>🔍 {t('Single Model Sales Detail','تفاصيل مبيعات موديل واحد')}</div>",
        unsafe_allow_html=True
    )

    if not sa_model_input:
        st.info(t(
            "💡 Enter a **Model Code** in the filter above to see single-model sales analytics.",
            "💡 أدخل **رمز الموديل** في الفلتر أعلاه لعرض تحليلات موديل واحد."
        ))
    elif sa_df_model is not None and sa_df_model.empty:
        st.info(t(
            f"No sales records found for model **{sa_model_input}**.",
            f"لا توجد سجلات مبيعات للموديل **{sa_model_input}**."
        ))
    elif sa_df_model is not None:
        sm_qty  = float(sa_df_model["Qty"].sum())
        sm_amt  = float(sa_df_model["Subtotal"].sum())
        sm_cust = int(sa_df_model["Customer"].nunique())

        sm_cards = [
            _premium_kpi_card("📦", f"{sm_qty:,.0f}",  t("Total Qty (this model)","إجمالي الكمية (الموديل)"), "c-green"),
            _premium_kpi_card("💰", f"{sm_amt:,.2f}",  t("Total Sales (SAR)","إجمالي المبيعات"),             "c-blue"),
            _premium_kpi_card("👤", str(sm_cust),       t("Customers","العملاء"),                              "c-purple"),
        ]
        _render_kpi_row(sm_cards)

        st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

        # Time series for this model
        _chart_card_open(f"{t('Sales Qty Over Time','كمية المبيعات عبر الزمن')} — {sa_model_input}", "📈")
        sm_ts = (
            sa_df_model.copy()
            .assign(Date=pd.to_datetime(sa_df_model["Date"], errors="coerce"))
            .dropna(subset=["Date"])
            .groupby("Date", as_index=False)["Qty"]
            .sum()
            .sort_values("Date")
        )
        if not sm_ts.empty:
            st.altair_chart(
                _alt_line_chart(sm_ts, "Date", "Qty", height=230, color="#43e97b"),
                use_container_width=True
            )
        _chart_card_close()

        st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

        # Top customers for this model
        sm_cust_grp = (
            sa_df_model.copy()
            .assign(Customer=sa_df_model["Customer"].replace("", "(No Customer)").fillna("(No Customer)"))
            .groupby("Customer", as_index=False)["Qty"]
            .sum()
            .sort_values("Qty", ascending=False)
            .head(10)
            .reset_index(drop=True)
        )
        sm_cust_grp["Total Qty"] = sm_cust_grp["Qty"].map(lambda v: f"{v:,.0f}")

        _chart_card_open(t("Top Customers for this Model","أعلى العملاء لهذا الموديل"), "👤")
        sc1, sc2 = st.columns([1.6, 1])
        with sc1:
            st.altair_chart(
                _alt_bar_chart(sm_cust_grp, x_field="Customer", y_field="Qty",
                               tooltip_fmt=",.0f", color="#9b59b6"),
                use_container_width=True
            )
        with sc2:
            _render_html_table(sm_cust_grp[["Customer", "Total Qty"]])
        _chart_card_close()

        # Customer share donut
        if not sm_cust_grp.empty:
            top_c    = sm_cust_grp.head(8)
            others_v = float(sm_cust_grp.iloc[8:]["Qty"].sum()) if len(sm_cust_grp) > 8 else 0
            p_labels = top_c["Customer"].tolist()
            p_vals   = top_c["Qty"].tolist()
            if others_v > 0:
                p_labels.append("Others"); p_vals.append(others_v)
            _d1, _d2, _d3 = st.columns([1, 1.2, 1])
            with _d2:
                _plotly_donut(p_labels, p_vals,
                              title=t("Customer Share", "حصة العملاء"), height=320)

    # ── Full Table + Downloads ─────────────────────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>📋 {t('Full Sales Detail','تفاصيل المبيعات الكاملة')}</div>",
        unsafe_allow_html=True
    )

    show_df = (
        sa_df_model
        if (sa_model_input and sa_df_model is not None and not sa_df_model.empty)
        else sa_df
    )
    full_show = show_df.copy()
    full_show["Unit Price"] = full_show["Unit Price"].map(lambda v: f"{v:.2f} SAR")
    full_show["Subtotal"]   = full_show["Subtotal"].map(lambda v: f"{v:,.2f} SAR")
    full_show["Qty"]        = full_show["Qty"].map(lambda v: f"{v:,.0f}")
    _render_html_table(full_show)

    st.markdown("<br>", unsafe_allow_html=True)
    sdl1, sdl2, _ = st.columns([1, 1, 2])
    tag_s = f"_{sa_model_input.upper()}" if sa_model_input else "_overall"
    export_df = (
        sa_df_model
        if (sa_model_input and sa_df_model is not None and not sa_df_model.empty)
        else sa_df
    )
    sdl1.download_button(
        "⬇️ CSV",
        export_df.to_csv(index=False).encode("utf-8-sig"),
        dl_name(f"sales{tag_s}", "csv"),
        "text/csv",
        use_container_width=True,
        key=f"sdl_csv{tag_s}"
    )
    sdl2.download_button(
        "⬇️ Excel",
        to_excel_sales(export_df),
        dl_name(f"sales{tag_s}", "xlsx"),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"sdl_xlsx{tag_s}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# SWAG PURCHASE ANALYTICS VIEW
# ─────────────────────────────────────────────────────────────────────────────
def show_purchase_analytics():
    st.markdown(f"### 🛒 {t('SWAG Purchase Analytics','تحليلات مشتريات سواغ')}")
    st.markdown(
        "<div class='info-banner'>📌 "
        + t("Purchase orders from the <b>SWAG</b> system only (state: purchase / done).",
            "أوامر الشراء من نظام <b>سواغ</b> فقط (الحالة: مشترى / منجز).")
        + "</div>",
        unsafe_allow_html=True
    )

    default_from = datetime.now().date() - timedelta(days=365)
    default_to   = datetime.now().date()

    filt_col1, filt_col2, filt_col3, filt_col4 = st.columns([1.2, 1, 1, 1.4])

    with filt_col1:
        po_model_input = st.text_input(
            f"🔖 {t('Model Code (optional)','رمز الموديل (اختياري)')}",
            placeholder=t("e.g. RVT196 — blank = all", "مثال: RVT196 — فارغ = الكل"),
            key="po_model_input_v3"
        ).strip()

    with filt_col2:
        po_date_from = st.date_input(
            f"📅 {t('From','من')}",
            value=default_from,
            key="po_date_from_v3"
        )

    with filt_col3:
        po_date_to = st.date_input(
            f"📅 {t('To','إلى')}",
            value=default_to,
            key="po_date_to_v3"
        )

    with filt_col4:
        cached_po     = st.session_state.get("po_analytics_df")
        vendor_options = []
        if cached_po is not None and not cached_po.empty and "Vendor" in cached_po.columns:
            vendor_options = sorted(cached_po["Vendor"].dropna().unique().tolist())

        all_vendors_label = t("All Vendors", "كل الموردين")
        vendor_choices    = [all_vendors_label] + vendor_options

        po_vendor_sel = st.multiselect(
            f"🏭 {t('Vendor','المورد')}",
            options=vendor_choices,
            default=[],
            placeholder=t("All Vendors (default)", "كل الموردين (افتراضي)"),
            key="po_vendor_sel_v3"
        )

    fetch_po_btn = st.button(
        f"🔍 {t('Fetch Purchase Analytics','جلب تحليلات المشتريات')}",
        type="primary",
        use_container_width=False,
        key="fetch_po_btn_v3"
    )

    if fetch_po_btn:
        with st.spinner(t("⚡ Fetching all purchase data from SWAG…","⚡ جلب بيانات المشتريات من نظام سواغ…")):
            fetched = fetch_swag_purchase_history(
                model_code=None,
                date_from=po_date_from.strftime("%Y-%m-%d"),
                date_to=po_date_to.strftime("%Y-%m-%d"),
            )
        st.session_state.po_analytics_df = fetched
        st.rerun()

    po_full = st.session_state.get("po_analytics_df")

    if po_full is None:
        st.info(t(
            "👆 Set your date range and click **Fetch Purchase Analytics** to load data.",
            "👆 حدد نطاق التاريخ واضغط **جلب تحليلات المشتريات** لتحميل البيانات."
        ))
        return

    if po_full.empty:
        st.info(t("No purchases found for this period.", "لا توجد مشتريات لهذه الفترة."))
        return

    active_vendors = [v for v in po_vendor_sel if v != all_vendors_label]
    if active_vendors:
        pdf_vendor = po_full[po_full["Vendor"].isin(active_vendors)].copy()
    else:
        pdf_vendor = po_full.copy()

    # ── Panel B — Single Model Purchase Detail ────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>🔍 {t('Panel B — Single Model Purchase Detail','لوحة ب — تفاصيل شراء موديل واحد')}</div>",
        unsafe_allow_html=True
    )

    if not po_model_input:
        st.info(t(
            "💡 Enter a **Model Code** in the filter above to see single-model analytics.",
            "💡 أدخل **رمز الموديل** في الفلتر أعلاه لعرض تحليلات موديل واحد."
        ))
    else:
        mc_norm  = po_model_input.upper()
        model_df = po_full[po_full["Model Code"].str.upper() == mc_norm].copy()

        if active_vendors:
            model_df = model_df[model_df["Vendor"].isin(active_vendors)]

        if model_df.empty:
            st.info(t(
                f"No purchase records found for model **{po_model_input}**.",
                f"لا توجد سجلات شراء للموديل **{po_model_input}**."
            ))
        else:
            pb_qty  = float(model_df["Qty"].sum())
            pb_amt  = float(model_df["Subtotal"].sum())
            pb_vend = int(model_df["Vendor"].nunique())

            pb_cards = [
                _premium_kpi_card("📦", f"{pb_qty:,.0f}",  t("Total Qty (this model)","إجمالي الكمية (الموديل)"), "c-blue"),
                _premium_kpi_card("💵", f"{pb_amt:,.2f}",  t("Total Amount (SAR)","إجمالي المبلغ"),               "c-purple"),
                _premium_kpi_card("🏭", str(pb_vend),       t("Vendors","الموردون"),                               "c-green"),
            ]
            _render_kpi_row(pb_cards)

            model_vendors = sorted(model_df["Vendor"].dropna().unique().tolist())
            pb_vendor_sel = st.multiselect(
                f"🏭 {t('Filter vendors for this model','فلتر الموردين لهذا الموديل')}",
                options=model_vendors,
                default=[],
                placeholder=t("All vendors for this model", "كل موردين هذا الموديل"),
                key="pb_vendor_sel_v3"
            )

            model_vendor_df = (
                model_df[model_df["Vendor"].isin(pb_vendor_sel)].copy()
                if pb_vendor_sel else model_df.copy()
            )

            if model_vendor_df.empty:
                st.warning(t("No data for selected vendor(s).", "لا بيانات للموردين المحددين."))
            else:
                st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

                _chart_card_open(t("Purchase Qty Over Time","كمية الشراء عبر الزمن"), "📈")
                ts_df = (
                    model_vendor_df
                    .groupby("Date", as_index=False)["Qty"]
                    .sum()
                    .sort_values("Date")
                )
                if not ts_df.empty:
                    ts_plot = ts_df.copy()
                    ts_plot["Date"] = pd.to_datetime(ts_plot["Date"], errors="coerce")
                    ts_plot = ts_plot.dropna(subset=["Date"])
                    if not ts_plot.empty:
                        st.altair_chart(
                            _alt_line_chart(ts_plot, "Date", "Qty", color="#764ba2"),
                            use_container_width=True
                        )
                _chart_card_close()

                st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

                # Vendor share
                _chart_card_open(t("Vendor Share for this Model","حصة الموردين لهذا الموديل"), "🏭")
                vshare = (
                    model_vendor_df
                    .assign(Vendor=model_vendor_df["Vendor"].replace("", "(No Vendor)").fillna("(No Vendor)"))
                    .groupby("Vendor", as_index=False)["Qty"]
                    .sum()
                    .sort_values("Qty", ascending=False)
                    .reset_index(drop=True)
                )
                vshare["Total Qty"] = vshare["Qty"].map(lambda v: f"{v:,.0f}")
                vs1, vs2 = st.columns([1.6, 1])
                with vs1:
                    st.altair_chart(
                        _alt_bar_chart(vshare, x_field="Vendor", y_field="Qty",
                                       tooltip_fmt=",.0f", color="#9b59b6"),
                        use_container_width=True
                    )
                with vs2:
                    _render_html_table(vshare[["Vendor", "Total Qty"]])
                _chart_card_close()

                # Vendor donut
                if not vshare.empty:
                    _vd1, _vd2, _vd3 = st.columns([1, 1.2, 1])
                    with _vd2:
                        _plotly_donut(
                            vshare["Vendor"].tolist(),
                            vshare["Qty"].tolist(),
                            title=t("Vendor Share", "حصة الموردين"),
                            height=300,
                        )

                st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

                st.markdown(f"#### 📋 {t('Model Detail Table','جدول تفاصيل الموديل')} — {po_model_input}")
                _po_full_table(model_vendor_df)
                st.markdown("<br>", unsafe_allow_html=True)
                _po_download_row(model_vendor_df, tag_suffix=f"_{mc_norm}_v3")

    # ── Panel A — Overall Purchase Analytics ──────────────────────────────────
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>📊 {t('Panel A — Overall Purchase Analytics','لوحة أ — تحليلات المشتريات الإجمالية')}</div>",
        unsafe_allow_html=True
    )

    if pdf_vendor.empty:
        st.warning(t("No data for the selected vendor(s).","لا توجد بيانات للمورد المحدد."))
        return

    _po_kpi_row(pdf_vendor, prefix="pa_v3")
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Vendors by Purchase Amount
    _chart_card_open(t("Top 10 Vendors by Purchase Amount","أعلى 10 موردين حسب مبلغ الشراء"), "🏭")
    vendor_grp = (
        pdf_vendor.copy()
        .assign(Vendor=pdf_vendor["Vendor"].replace("", "(No Vendor)").fillna("(No Vendor)"))
        .groupby("Vendor", as_index=False)["Subtotal"]
        .sum()
        .sort_values("Subtotal", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    vendor_grp["Total Amount (SAR)"] = vendor_grp["Subtotal"].map(lambda v: f"{v:,.2f}")
    vc1, vc2 = st.columns([1.6, 1])
    with vc1:
        st.altair_chart(
            _alt_bar_chart(vendor_grp, x_field="Vendor", y_field="Subtotal",
                           tooltip_fmt=",.2f", color="#667eea"),
            use_container_width=True
        )
    with vc2:
        _render_html_table(vendor_grp[["Vendor", "Total Amount (SAR)"]])
    _chart_card_close()

    # Vendor share donut
    if not vendor_grp.empty:
        _plotly_donut(
            vendor_grp["Vendor"].tolist(),
            vendor_grp["Subtotal"].tolist(),
            title=t("Vendor Share (Top 10)", "حصة الموردين (أعلى 10)")
        )

    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    # Top 10 Products by Qty
    _chart_card_open(t("Top 10 Products by Qty","أعلى 10 منتجات حسب الكمية"), "🏆")
    prod_grp_a = (
        pdf_vendor.copy()
        .assign(**{
            "Model Code": pdf_vendor["Model Code"].replace("", "(No Code)").fillna("(No Code)"),
            "Product":    pdf_vendor["Product"].replace("", "").fillna(""),
        })
        .groupby(["Model Code", "Product"], as_index=False)["Qty"]
        .sum()
        .sort_values("Qty", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    prod_grp_a["Total Qty"] = prod_grp_a["Qty"].map(lambda v: f"{v:,.0f}")
    pc1, pc2 = st.columns([1.6, 1])
    with pc1:
        st.altair_chart(
            _alt_bar_chart(prod_grp_a, x_field="Model Code", y_field="Qty",
                           tooltip_fmt=",.0f", color="#43e97b"),
            use_container_width=True
        )
    with pc2:
        _render_html_table(prod_grp_a[["Model Code", "Product", "Total Qty"]])
    _chart_card_close()

    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)

    _po_top10_altair(
        t("Top 10 Categories by Qty","أعلى 10 فئات حسب الكمية"),
        "Category", "Qty", pdf_vendor,
        color="#4facfe", tooltip_fmt=",.0f"
    )
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    _po_top10_altair(
        t("Top 10 Brand Categories by Qty","أعلى 10 فئات علامة تجارية حسب الكمية"),
        "Brand Category", "Qty", pdf_vendor,
        color="#f093fb", tooltip_fmt=",.0f"
    )

    # Category + Brand category donuts
    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='panel-header'>🥧 {t('Purchase Share Analysis','تحليل حصص المشتريات')}</div>",
        unsafe_allow_html=True
    )
    d1, d2 = st.columns(2)
    with d1:
        cat_share = (
            pdf_vendor.copy()
            .assign(Category=pdf_vendor["Category"].replace("", "(No Category)").fillna("(No Category)"))
            .groupby("Category", as_index=False)["Qty"]
            .sum()
            .sort_values("Qty", ascending=False)
        )
        _plotly_donut(
            cat_share["Category"].tolist(),
            cat_share["Qty"].tolist(),
            title=t("Category Share", "حصة الفئة")
        )
    with d2:
        bc_share_p = (
            pdf_vendor.copy()
            .assign(**{"Brand Category": pdf_vendor["Brand Category"].replace("", "(No Brand)").fillna("(No Brand)")})
            .groupby("Brand Category", as_index=False)["Qty"]
            .sum()
            .sort_values("Qty", ascending=False)
        )
        _plotly_donut(
            bc_share_p["Brand Category"].tolist(),
            bc_share_p["Qty"].tolist(),
            title=t("Brand Category Share", "حصة الفئة التجارية")
        )

    st.markdown("<div class='section-sep'></div>", unsafe_allow_html=True)
    st.markdown(f"#### 📋 {t('Full Purchase Detail','تفاصيل المشتريات الكاملة')}")
    _po_full_table(pdf_vendor)
    st.markdown("<br>", unsafe_allow_html=True)
    _po_download_row(pdf_vendor, tag_suffix="_overall_v3")


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def show_dashboard():
    # ── SIDEBAR ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(f"### ⚙️ {t('Settings','الإعدادات')}")

        lc2 = st.radio(t("🌐 Language","🌐 اللغة"),["EN","AR"],
                       index=0 if get_lang()=="EN" else 1, horizontal=True)
        if lc2!=get_lang(): st.session_state.lang=lc2; st.rerun()

        st.divider()
        st.markdown(f"##### 📊 {t('Analytics','التحليلات')}")

        current_view = st.session_state.get("analytics_view", "purchase")

        sales_type  = "primary" if current_view == "sales"    else "secondary"
        purch_type  = "primary" if current_view == "purchase" else "secondary"

        col_s, col_p = st.columns(2)
        with col_s:
            if st.button(
                f"💰 {t('Sales','المبيعات')}",
                type=sales_type,
                use_container_width=True,
                key="nav_sales_btn"
            ):
                st.session_state.analytics_view = "sales"
                st.rerun()
        with col_p:
            if st.button(
                f"🛒 {t('Purchase','المشتريات')}",
                type=purch_type,
                use_container_width=True,
                key="nav_purchase_btn"
            ):
                st.session_state.analytics_view = "purchase"
                st.rerun()

        if current_view == "sales":
            st.markdown(
                "<div style='background:linear-gradient(90deg,#667eea22,#43e97b22);border-left:3px solid #43e97b;"
                "border-radius:6px;padding:6px 10px;font-size:0.78rem;color:#86efac;margin-top:4px;'>"
                f"✅ {t('Viewing: SWAG Sales','عرض: مبيعات سواغ')}</div>",
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='background:linear-gradient(90deg,#667eea22,#f093fb22);border-left:3px solid #667eea;"
                "border-radius:6px;padding:6px 10px;font-size:0.78rem;color:#c4b5fd;margin-top:4px;'>"
                f"✅ {t('Viewing: SWAG Purchase','عرض: مشتريات سواغ')}</div>",
                unsafe_allow_html=True
            )

        st.divider()

        st.markdown(f"👤 **{st.session_state.user_email}**")
        if st.button(f"🚪 {t('Logout','تسجيل الخروج')}", use_container_width=True):
            do_logout()
        st.divider()

        st.markdown(f"##### 🔬 {t('Search Mode','وضع البحث')}")
        et = st.toggle(t("Exact match only","تطابق تام فقط"),
                       value=st.session_state.search_exact)
        if et!=st.session_state.search_exact:
            st.session_state.search_exact = et
            st.session_state.total_df     = None
            st.session_state.branch_df    = None
            st.session_state.transfers_df = None
            st.rerun()
        st.caption(t("🎯 Exact","🎯 تطابق تام") if st.session_state.search_exact
                   else t("🔍 Variant wildcard","🔍 كل المتغيرات"))
        st.divider()

        st.markdown(f"##### 🔴 {t('Low Stock Alert','تنبيه المخزون')}")
        thr = st.number_input(t("Threshold (qty ≤)","الحد (كمية ≤)"),
                              min_value=0, max_value=1000,
                              value=st.session_state.low_stock_thresh, step=1)
        if thr!=st.session_state.low_stock_thresh:
            st.session_state.low_stock_thresh = int(thr)
        st.divider()

        if st.session_state.last_run:
            st.markdown(f"🕒 **{t('Last Run','آخر تشغيل')}**")
            st.caption(st.session_state.last_run.get("time",""))

    # ── Route to analytics views ───────────────────────────────────────────────
    current_view = st.session_state.get("analytics_view", "purchase")

    if current_view == "sales":
        st.markdown(f"""
        <div class='dash-header'>
            <div class='dash-title'>💰 {t('SWAG Sales Dashboard','لوحة مبيعات سواغ')}</div>
            <div class='dash-subtitle'>{t('Sales analytics from SWAG Odoo system',
                                           'تحليلات المبيعات من نظام سواغ أودو')}</div>
        </div>""", unsafe_allow_html=True)
        st.divider()
        show_sales_analytics()
        return

    if current_view == "purchase":
        st.markdown(f"""
        <div class='dash-header'>
            <div class='dash-title'>🛒 {t('SWAG Purchase Dashboard','لوحة مشتريات سواغ')}</div>
            <div class='dash-subtitle'>{t('Purchase analytics from SWAG Odoo system',
                                           'تحليلات المشتريات من نظام سواغ أودو')}</div>
        </div>""", unsafe_allow_html=True)
        st.divider()
        show_purchase_analytics()
        return

    # ── Default: Stock Comparison Dashboard ───────────────────────────────────
    st.markdown(f"""
    <div class='dash-header'>
        <div class='dash-title'>📊 {t('SWAG Product Comparison','مقارنة منتجات سواغ')}</div>
        <div class='dash-subtitle'>{t('Real-time stock & price across 4 Odoo systems',
                                       'المخزون والسعر الآني عبر 4 أنظمة أودو')}</div>
    </div>""", unsafe_allow_html=True)
    st.divider()

    # ── PDF ───────────────────────────────────────────────────────────────────
    st.markdown(f"### 📄 {t('Upload Invoice PDF','رفع فاتورة PDF')}")
    p1,p2 = st.columns([2.5,1.5])
    with p1:
        updf = st.file_uploader(t("Upload PDF","رفع PDF"),
                                type=["pdf"], label_visibility="collapsed")
    with p2:
        emode = None
        if updf:
            emode = st.radio(t("Extract mode","وضع الاستخراج"),
                             [t("Main models","موديلات رئيسية"),
                              t("With sizes","مع المقاسات")], horizontal=True)

    if updf:
        fbytes = updf.read()
        fhash  = hashlib.md5(fbytes).hexdigest()
        ck     = f"pdf_{fhash}"
        if ck not in st.session_state:
            with st.spinner(t("⚡ Parsing PDF...","⚡ جاري قراءة الفاتورة...")):
                st.session_state[ck] = parse_invoice_pdf_cached(fbytes)

        raw = st.session_state[ck]

        if raw:
            is_main = emode is None or "Main" in emode or "رئيسية" in emode

            if is_main:
                unique = get_unique_base_models(raw)
            else:
                seen_ws, unique = set(), []
                for item in raw:
                    if item["code"] not in seen_ws:
                        seen_ws.add(item["code"])
                        unique.append(item)

            unique_sorted = sorted(unique, key=lambda x: x["sequence"])
            unique_codes  = [item["code"] for item in unique_sorted]

            c1,c2,c3 = st.columns(3)
            c1.metric(t("Raw codes","رموز مستخرجة"), len(raw))
            c2.metric(t("Unique models","موديلات فريدة"), len(unique_codes))
            c3.info(f"📌 {t('Main','رئيسية') if is_main else t('With sizes','مع المقاسات')}")

            with st.expander(t(f"📋 {len(unique_codes)} codes","📋 الرموز"), expanded=False):
                st.code(
                    "\n".join(
                        f"{item['sequence']:>3}. {item['code']}"
                        for item in unique_sorted
                    )
                )

            ca,cb = st.columns(2)
            with ca:
                if st.button(f"🚀 {t('Total Stock','مخزون إجمالي')}",
                             type="primary", use_container_width=True, key="pt"):
                    st.session_state.pdf_codes = unique_codes
                    st.session_state.pdf_mode  = "total"
                    st.rerun()
            with cb:
                if st.button(f"🗺️ {t('Branch-wise','حسب الفرع')}",
                             type="secondary", use_container_width=True, key="pb"):
                    st.session_state.pdf_codes = unique_codes
                    st.session_state.pdf_mode  = "branch"
                    st.rerun()
        else:
            st.warning(t("No codes found in PDF.","لم يتم العثور على رموز."))

    st.divider()

    # ── Manual Search ─────────────────────────────────────────────────────────
    st.markdown(f"### ✍️ {t('Manual Search','بحث يدوي')}")
    L,R = st.columns([1.5,1])
    with L:
        if not st.session_state.search_exact:
            st.markdown(
                "<div class='info-banner'>🔍 <b>Variant mode</b> — XP6013 → XP6013-S/M/L</div>",
                unsafe_allow_html=True)
        else:
            st.markdown(
                "<div class='warn-banner'>🎯 <b>Exact match mode</b> — identical codes only.</div>",
                unsafe_allow_html=True)
        ms   = t("Single Model","موديل واحد")
        mm   = t("Multiple Models","موديلات متعددة")
        mode = st.radio(t("Mode","الوضع"),[ms,mm],
                        horizontal=True, label_visibility="collapsed")
        if mode==mm:
            rt    = st.text_area(t("Codes","الرموز"), height=130,
                                 placeholder="ABC123\nDEF456")
            codes = [c.strip() for c in rt.replace(",","\n").splitlines() if c.strip()]
        else:
            sg    = st.text_input(t("Model Code","رمز الموديل"), placeholder="e.g. XP6013")
            codes = [sg.strip()] if sg.strip() else []

        t1,t2,t3,t4,t5 = st.columns(5)
        sz  = t1.toggle(t("Zero","الصفري"),     value=False)
        sb  = t2.toggle(t("Branch","فروع"),      value=False)
        ss  = t3.toggle(t("Sort","ترتيب"),       value=False)
        st_ = t4.toggle(t("Transfers","نقليات"), value=False)
        sr  = t5.toggle(t("Reorder","طلب"),      value=False)

        if sr:
            with st.expander(f"⚙️ {t('Reorder Settings','إعدادات')}", expanded=True):
                rx,ry = st.columns(2)
                with rx:
                    rm = st.radio(
                        t("Mode","الوضع"),
                        [t("Days cover","تغطية أيام"),t("Max level","مستوى أقصى")],
                        horizontal=True,
                        index=0 if st.session_state.reorder_mode=="days_cover" else 1)
                    st.session_state.reorder_mode = (
                        "days_cover" if "Days" in rm or "تغطية" in rm else "max_level")
                with ry:
                    st.session_state.reorder_point = st.number_input(
                        t("Reorder point","نقطة الطلب"), min_value=0, max_value=9999,
                        value=st.session_state.reorder_point, step=1)
                if st.session_state.reorder_mode=="days_cover":
                    st.session_state.reorder_target_days = st.slider(
                        t("Target days","أيام"), 7, 180,
                        st.session_state.reorder_target_days)
                else:
                    st.session_state.reorder_max_level = st.number_input(
                        t("Max level","الحد"), min_value=1, max_value=99999,
                        value=st.session_state.reorder_max_level, step=1)

        cbtn = st.button(f"🔍 {t('Compare','مقارنة')}",
                         use_container_width=True, type="primary")

    with R:
        st.markdown(f"#### 📋 {t('Last Run','آخر تشغيل')}")
        snap  = st.session_state.last_run
        stats = st.session_state.sys_stats
        if not snap:
            st.info(t("Run a comparison first.","قم بتشغيل مقارنة أولاً."))
        else:
            on = sum(1 for v in stats.values() if v=="OK")
            st.markdown(
                f"<div class='snap-card'>"
                f"🕒 <b>{t('Time','الوقت')}:</b> {snap.get('time','—')}<br>"
                f"📦 <b>{t('Models','الموديلات')}:</b> {snap.get('models','—')}<br>"
                f"🌐 <b>{t('Online','متصل')}:</b> {on}/4<br>"
                f"📊 <b>{t('Rows','الصفوف')}:</b> {snap.get('rows','—')}"
                f"</div>", unsafe_allow_html=True)
            st.markdown("")
            for key in SYSTEM_KEYS:
                s  = stats.get(key,"—")
                bc = "badge-ok" if s=="OK" else "badge-off" if s=="NOT_FOUND" else "badge-err"
                bt = "✅ OK"    if s=="OK" else "🔴 OFF"    if s=="NOT_FOUND" else "⚠️ ERR"
                display_name = get_system_name(key)
                st.markdown(
                    f"<div class='sys-row'>"
                    f"<span style='font-size:.85rem;color:#e8e8ff'>"
                    f"<b>{display_name}</b></span>"
                    f"<span class='{bc}'>{bt}</span></div>",
                    unsafe_allow_html=True)

    # ── Run ───────────────────────────────────────────────────────────────────
    run_codes    = None
    force_branch = False
    if st.session_state.get("pdf_codes"):
        run_codes    = st.session_state.pdf_codes
        force_branch = st.session_state.get("pdf_mode","total")=="branch"
        sb = True
        st.session_state.pdf_codes = None
        st.session_state.pdf_mode  = "total"
    elif cbtn:
        run_codes = codes

    if run_codes is not None:
        if not run_codes:
            st.warning(t("Enter at least one model code.","أدخل رمزاً واحداً.")); st.stop()
        run_codes = list(dict.fromkeys([c.strip() for c in run_codes if c.strip()]))
        ct = tuple(run_codes)
        with st.spinner(t("⚡ Fetching from 4 systems…","⚡ جلب البيانات من 4 أنظمة…")):
            data = fetch_all_data(
                ct, exact=st.session_state.search_exact,
                need_branch=sb or force_branch,
                need_transfers=st_, need_reorder=sr,
                reorder_mode=st.session_state.reorder_mode,
                target_days=st.session_state.reorder_target_days,
                max_level=st.session_state.reorder_max_level,
                reorder_point=st.session_state.reorder_point)

        tdf  = prepare_df(data["total"])
        bdf  = prepare_df(data["branch"])
        trdf = prepare_df(data["transfers"])
        rdf  = prepare_df(data["reorder"])

        sc2 = "System"
        raw_tdf = data["total"]
        ns = {k:"NOT_FOUND" for k in SYSTEM_KEYS}
        if "_status" in raw_tdf.columns and sc2 in raw_tdf.columns:
            for key in SYSTEM_KEYS:
                mask = raw_tdf[sc2] == key
                if mask.any():
                    sv = raw_tdf.loc[mask,"_status"]
                    if   "OK"    in sv.values: ns[key]="OK"
                    elif "ERROR" in sv.values: ns[key]="ERROR"

        qc2 = t("On Hand","متوفر")
        sc2_loc = t("System","النظام")
        if qc2 in tdf.columns:
            zero_mask = pd.to_numeric(tdf[qc2], errors="coerce").fillna(0) == 0
            tdf.loc[zero_mask, "_status"] = "not_available"

        if ss and sc2_loc in tdf.columns:
            tdf = tdf.sort_values(sc2_loc).reset_index(drop=True)
        if not bdf.empty and ss and sc2_loc in bdf.columns:
            bdf = bdf.sort_values(sc2_loc).reset_index(drop=True)

        if sz:
            zero_count = int((pd.to_numeric(tdf[qc2], errors="coerce").fillna(0) == 0).sum())
            if zero_count:
                st.sidebar.info(
                    t(f"ℹ️ {zero_count} rows have zero qty (shown as ❌ Not Available)",
                      f"ℹ️ {zero_count} صف بكمية صفر (معروض كـ ❌ لا يوجد)"))

        st.session_state.total_df       = tdf
        st.session_state.branch_df      = bdf
        st.session_state.transfers_df   = trdf
        st.session_state.reorder_df     = rdf
        st.session_state.show_transfers = st_
        st.session_state.show_reorder   = sr
        st.session_state.sys_stats      = ns
        st.session_state.last_run       = {
            "time"  : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "models": len(run_codes),
            "rows"  : len(tdf),
        }
        record_price_snapshot(tdf)
        st.rerun()

    # ── Results ───────────────────────────────────────────────────────────────
    tdf  = st.session_state.total_df
    bdf  = st.session_state.branch_df
    trdf = st.session_state.transfers_df
    rdf  = st.session_state.reorder_df
    if tdf is None or tdf.empty: return

    st.divider()
    thr   = st.session_state.low_stock_thresh
    qc2   = t("On Hand","متوفر"); pc2 = t("Sale Price","سعر البيع")
    sc2   = t("System","النظام"); stats = st.session_state.sys_stats
    ok    = tdf[tdf["_status"]=="OK"] if "_status" in tdf.columns else tdf
    on    = sum(1 for v in stats.values() if v=="OK")

    if thr>0 and qc2 in ok.columns:
        low = ok[(ok[qc2]>0)&(ok[qc2]<=thr)]
        if not low.empty:
            mc2 = t("Model Code","رمز الموديل")
            det = ", ".join(
                f"{r.get(mc2,'?')}@{r.get(sc2,'?')}({r.get(qc2,0)})"
                for _,r in low.head(8).iterrows())
            if len(low)>8: det+=f" +{len(low)-8}"
            st.markdown(
                f"<div class='alert-banner'>🔴 <b>{t('Low Stock','مخزون منخفض')}:</b> "
                f"{len(low)} ≤{thr} — <span class='mono'>{det}</span></div>",
                unsafe_allow_html=True)

    m1,m2,m3,m4 = st.columns(4)
    m1.metric(t("Total Rows","إجمالي الصفوف"), len(tdf))
    m2.metric(t("Systems Online","الأنظمة"), f"{on}/4")
    if qc2 in ok.columns:
        m3.metric(t("Total Qty","إجمالي الكمية"),
                  int(pd.to_numeric(ok[qc2], errors="coerce").fillna(0).sum()))
    if pc2 in ok.columns:
        vp = ok[ok[pc2]>0][pc2]
        m4.metric(t("Avg Price","متوسط السعر"),
                  f"{vp.mean():.2f} SAR" if not vp.empty else "—")

    hb = bdf  is not None and not bdf.empty
    ht = st.session_state.show_transfers and trdf is not None and not trdf.empty
    hr = st.session_state.show_reorder   and rdf  is not None and not rdf.empty

    tlabels = [f"📦 {t('Total Stock','المخزون الإجمالي')}",
               f"📊 {t('Price History','تاريخ الأسعار')}"]
    if hb: tlabels.append(f"🗺️ {t('Branch Stock','مخزون الفروع')}")
    if ht: tlabels.append(f"🚚 {t('Transfers','النقليات')}")
    if hr: tlabels.append(f"📦 {t('Reorder','إعادة الطلب')}")

    tabs = st.tabs(tlabels); ti = 0

    with tabs[ti]:
        ti+=1
        st.markdown(f"### 📦 {t('Total Stock','المخزون الإجمالي')}")
        display_df(tdf, thr, table_key="total")
        st.markdown("<br>", unsafe_allow_html=True)
        d1,d2,d3,_ = st.columns([1,1,1,1])
        d1.download_button("⬇️ CSV",
            to_csv(tdf), dl_name("total","csv"), "text/csv",
            use_container_width=True)
        d2.download_button("⬇️ Excel",
            to_excel(tdf), dl_name("total","xlsx"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
        d3.download_button("📥 All Systems",
            to_excel_bulk(tdf), dl_name("bulk","xlsx"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    with tabs[ti]:
        ti+=1
        st.markdown(f"### 📈 {t('Price History','تاريخ الأسعار')}")
        hdf = build_price_history_df()
        if hdf.empty:
            st.info(t("Run multiple comparisons to track prices.",
                      "قم بتشغيل مقارنات متعددة لتتبع الأسعار."))
        else:
            st.line_chart(hdf, use_container_width=True)
            if st.button(f"🗑️ {t('Clear History','مسح السجل')}"):
                st.session_state.price_history={}; st.rerun()

    if hb:
        with tabs[ti]:
            ti+=1
            st.markdown(f"### 🗺️ {t('Branch-wise Stock','مخزون حسب الفرع')}")
            display_df(bdf, thr, table_key="branch")
            bc2 = t("Branch","الفرع")
            okb = bdf[bdf["_status"]=="OK"] if "_status" in bdf.columns else bdf
            if not okb.empty and bc2 in okb.columns and qc2 in okb.columns:
                chart = okb.groupby([sc2,bc2])[qc2].sum().reset_index()
                if not chart.empty:
                    st.markdown(f"#### 📊 {t('Qty by Branch','الكميات حسب الفرع')}")
                    st.bar_chart(chart.set_index(bc2)[qc2], use_container_width=True)
            b1,b2,_ = st.columns([1,1,2])
            b1.download_button("⬇️ CSV",
                to_csv(bdf), dl_name("branch","csv"), "text/csv",
                use_container_width=True)
            b2.download_button("⬇️ Excel",
                to_excel(bdf), dl_name("branch","xlsx"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    if ht:
        with tabs[ti]:
            ti+=1
            st.markdown(f"### 🚚 {t('Pending Transfers','النقليات المعلقة')}")
            okt = trdf[trdf["_status"]=="OK"] if "_status" in trdf.columns else trdf
            if not okt.empty:
                k1,k2,k3 = st.columns(3)
                k1.metric(t("Total","إجمالي"), len(okt))
                qd = t("Qty","الكمية")
                if qd  in okt.columns: k2.metric(t("Total Qty","إجمالي الكمية"), int(okt[qd].sum()))
                if sc2 in okt.columns: k3.metric(t("Systems","الأنظمة"), okt[sc2].nunique())
            display_df(trdf, thresh=0, table_key="transfers")
            x1,x2,_ = st.columns([1,1,2])
            x1.download_button("⬇️ CSV",
                to_csv(trdf), dl_name("transfers","csv"), "text/csv",
                use_container_width=True)
            x2.download_button("⬇️ Excel",
                to_excel(trdf), dl_name("transfers","xlsx"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    if hr:
        with tabs[ti]:
            ti+=1
            CPRI  = t("Priority","الأولوية")
            CSUGG = t("Suggest","المقترح")
            st.markdown(f"### 📦 {t('Reorder Suggestions','اقتراحات إعادة الطلب')}")
            okr = rdf[rdf["_status"]=="OK"] if "_status" in rdf.columns else rdf
            if not okr.empty:
                crit = okr[okr[CPRI].str.startswith("🔴")].shape[0] if CPRI in okr.columns else 0
                lo   = okr[okr[CPRI].str.startswith("🟡")].shape[0] if CPRI in okr.columns else 0
                okn  = okr[okr[CPRI].str.startswith("🟢")].shape[0] if CPRI in okr.columns else 0
                sg   = int(okr[CSUGG].sum())                         if CSUGG in okr.columns else 0
                r1,r2,r3,r4 = st.columns(4)
                r1.metric(t("🔴 Critical","🔴 حرج"), crit)
                r2.metric(t("🟡 Low","🟡 منخفض"), lo)
                r3.metric(t("🟢 OK","🟢 كافٍ"), okn)
                r4.metric(t("To Order","للطلب"), sg)
                if crit+lo>0:
                    st.markdown(
                        f"<div class='alert-banner'>🔴 {crit+lo} "
                        f"{t('products need reordering','منتجات تحتاج إعادة طلب')}</div>",
                        unsafe_allow_html=True)
                sa = st.toggle(t("Show all","عرض الكل"), value=False)
                dr = (okr if sa else
                      okr[okr[CPRI].str.startswith(("🔴","🟡"))] if CPRI in okr.columns else okr)
                display_df(dr.reset_index(drop=True), table_key="reorder")
            else:
                st.info(t("No reorder data.","لا بيانات إعادة طلب."))
            o1,o2,_ = st.columns([1,1,2])
            o1.download_button("⬇️ CSV",
                to_csv(rdf), dl_name("reorder","csv"), "text/csv",
                use_container_width=True)
            o2.download_button("⬇️ Excel",
                to_excel(rdf), dl_name("reorder","xlsx"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
restore_session()

if not st.session_state.authenticated:
    show_login()
else:
    show_dashboard()
